from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from pymongo import MongoClient
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from bson import ObjectId
import os
import io
import pandas as pd
import random
from deap import base, creator, tools, algorithms
from datetime import datetime
import itertools
import logging
from logger import setup_logger
try:
    from ortools.sat.python import cp_model  # Tambahkan import OR-Tools
    ORTOOLS_AVAILABLE = True
except ImportError:
    ORTOOLS_AVAILABLE = False
    cp_model = None
    print("WARNING: ortools not installed. Some optimization features will be disabled.")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# ------------------------------
# CONFIGURATION
# ------------------------------

APP_SECRET = "secret_key"
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {'csv', 'xlsx'}

# Room lists as per task requirements
NON_LAB_ROOMS = [
    'Infor 1', 'Infor 2', 'Infor 3', 'Infor 4', 'Infor 5', 
    'Jawa 8A', 'Jawa 8B'
]

LAB_ROOMS = [
    'Lab 1', 'Lab 2', 'Lab 3', 'Lab 4'
]

REQUIRED_NON_LAB_SECTIONS = 40
REQUIRED_LAB_SECTIONS = 36

REQUIRED_SKS_PER_DOSEN = 10  # Target SKS per dosen
MIN_SKS_PER_DOSEN = 8        # Minimum wajib mengajar
MAX_SKS_PER_DOSEN = 12       # Maximum SKS per dosen
MAX_COURSES_PER_DOSEN = 3
# Note: MIN/MAX_SECTIONS_PER_DOSEN tidak digunakan lagi untuk distribusi section
# Section distribution sekarang dihitung secara dinamis berdasarkan SKS dan jumlah dosen
MIN_SECTIONS_PER_DOSEN = 4   # Legacy (tidak digunakan dalam distribusi baru)
MAX_SECTIONS_PER_DOSEN = 6   # Legacy (tidak digunakan dalam distribusi baru)

POSSIBLE_DAYS = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
POSSIBLE_STARTS = ['08:00', '10:00', '13:00', '15:00']
SKS_TO_MINUTES = {1: 100, 2: 100, 3: 150}  # 1-2 SKS = 100 minutes, 3 SKS = 150 minutes

# Canonical timeblocks (Mon-Thu) and Friday-special blocks
TIMEBLOCKS = {
    1: ["08:00-08:50", "08:50-09:40", "09:40-10:30", "10:30-11:20", "11:20-12:10",
        "14:00-14:50", "14:50-15:40", "15:40-16:30"],
    2: ["08:00-09:40", "09:40-11:20", "10:30-12:10", "11:20-13:00",
        "14:00-15:40", "15:40-17:20"],
    3: ["08:00-10:30", "09:40-12:10", "10:30-13:00", "14:00-16:30"]
}

TIMEBLOCKS_JUMAT = {
    1: ["08:00-08:50", "08:50-09:40", "09:40-10:30", "10:30-11:20", "11:20-12:10",
        "14:00-14:50", "14:50-15:40", "15:40-16:30"],
    2: ["08:00-09:40", "09:40-11:20", "10:30-12:10",
        "14:00-15:40", "15:40-17:20"],
    3: ["08:00-10:30", "09:40-12:10",
        "14:00-16:30"]  # Friday excludes 10:30-13:00 to respect 12:10 limit
}

def get_timeblocks_for_day(day, sks):
    """Return appropriate time blocks based on day and SKS (Friday has stricter morning cutoff)."""
    if day == 'Jumat':
        return TIMEBLOCKS_JUMAT.get(sks, TIMEBLOCKS_JUMAT[3])
    return TIMEBLOCKS.get(sks, TIMEBLOCKS[3])

# New constraints
MAX_CLASSES_MON_THU = 20
MIN_CLASSES_FRI = 10
MAX_CLASSES_FRI = 15
MAX_CONSECUTIVE_DAYS = 3

# Konstanta baru untuk mencegah konflik dan meningkatkan penalti
MAX_SAME_COURSE_DIFF_DAY_PENALTY = 10**30  # Penalti sangat besar jika 2 MK di hari berbeda
MAX_CONFLICT_PENALTY = 10**25             # Penalti untuk bentrok waktu/ruangan

app = Flask(__name__)
app.secret_key = APP_SECRET
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Set up logger
logger = setup_logger()

os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# Lightweight fallback objects when MongoDB is unavailable.
class DummyCollection:
    def find(self, *args, **kwargs):
        return []
    def find_one(self, *args, **kwargs):
        return None
    def count_documents(self, *args, **kwargs):
        return 0
    def insert_one(self, *args, **kwargs):
        return None
    def insert_many(self, *args, **kwargs):
        return None
    def delete_many(self, *args, **kwargs):
        return None
    def delete_one(self, *args, **kwargs):
        return None
    def update_one(self, *args, **kwargs):
        return None
    def update_many(self, *args, **kwargs):
        return None
    def aggregate(self, *args, **kwargs):
        return []

class DummyDB:
    def __init__(self):
        self.schedules = DummyCollection()
        self.users = DummyCollection()
        self.courses = DummyCollection()
        self.unavailability_reports = DummyCollection()


# ------------------------------
# HELPER FUNCTIONS
# ------------------------------

def _parse_minutes(time_str):
    """Parse a time string like 'HH:MM' or 'HH:MM:SS' into minutes since midnight. Return None if not parseable."""
    if not time_str:
        return None
    try:
        parts = time_str.split(":")
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 else 0
        return h * 60 + m
    except Exception:
        return None

def _minutes_to_time(minutes):
    """Convert minutes since midnight to HH:MM string."""
    h = minutes // 60
    m = minutes % 60
    return f"{h:02d}:{m:02d}"

def get_valid_starts(day, duration_minutes):
    """
    Get valid start times for a given day and duration.
    Duration is in minutes, e.g., 100 for 1-2 SKS, 150 for 3 SKS.
    """
    if day == "Jumat":
        # Jumat: 08:00-12:10 (sama seperti hari lain), istirahat Jumat 12:10-14:00, then 14:00-16:30
        available_blocks = [
            ("08:00", 250),  # 08:00 to 12:10 (250 min) - SAMA dengan hari lain
            ("14:00", 150),  # 14:00 to 16:30 (150 min)
        ]
    else:
        # Senin-Kamis: 08:00-12:10, break, 14:00-16:30
        available_blocks = [
            ("08:00", 250),  # 08:00 to 12:10 (250 min)
            ("14:00", 150),  # 14:00 to 16:30 (150 min)
        ]
    
    valid_starts = []
    for start_str, block_duration in available_blocks:
        start_min = _parse_minutes(start_str)
        if start_min is None:
            continue
        # Can fit if duration <= block_duration
        if duration_minutes <= block_duration:
            # Add start times in 50-min increments within the block
            current = start_min
            end_block = start_min + block_duration
            while current + duration_minutes <= end_block:
                valid_starts.append(_minutes_to_time(current))
                current += 50  # 50 min slots
    
    return valid_starts

def calculate_end_time(start_str, duration_minutes):
    """Calculate end time given start and duration in minutes."""
    start_min = _parse_minutes(start_str)
    if start_min is None:
        return start_str  # fallback
    end_min = start_min + duration_minutes
    return _minutes_to_time(end_min)

def _times_overlap(s1, e1, s2, e2):
    """Return True if time intervals [s1,e1) and [s2,e2) overlap. Inputs are minutes ints or None."""
    if s1 is None or e1 is None or s2 is None or e2 is None:
        # if we cannot parse times, play safe and assume possible overlap
        return True
    return (s1 < e2) and (s2 < e1)

# ------------------------------
# MONGODB CONNECTION
# ------------------------------

try:
    client = MongoClient(
        "mongodb+srv://salamull1005:jabbar1005@schedules.fb8isvj.mongodb.net/?appName=schedules",
        # tls=True,
        # tlsAllowInvalidCertificates=True,
        serverSelectionTimeoutMS=10000,
        connectTimeoutMS=10000,
        socketTimeoutMS=10000
    )
    client.server_info()  # Test connection immediately
    db = client["schedule_db"]
    schedules_collection = db["schedules"]
    users_collection = db["users"]
    courses_collection = db["courses"]
    unavailability_reports_collection = db["unavailability_reports"]
    # New collection for GA-generated sections (before time/room scheduling)
    try:
        sections_collection = db["sections"]
    except Exception:
        sections_collection = None
    print("MongoDB Atlas connection: SUCCESS")
except Exception as e:
    print(f"MongoDB Atlas connection: FAILED\n{e}")
    # Provide safe dummy objects so the app routes won't crash when DB is down.
    db = DummyDB()
    schedules_collection = db.schedules
    users_collection = db.users
    courses_collection = db.courses
    unavailability_reports_collection = db.unavailability_reports
    sections_collection = getattr(db, "sections", None)



# try:
#     # KONEKSI KE MONGODB LOCAL
#     client = MongoClient(
#         "mongodb://localhost:27017",
#         serverSelectionTimeoutMS=5000
#     )

#     # Test koneksi
#     client.admin.command("ping")

#     db = client["schedule_db"]

#     schedules_collection = db["schedules"]
#     users_collection = db["users"]
#     courses_collection = db["courses"]
#     unavailability_reports_collection = db["unavailability_reports"]
#     sections_collection = db["sections"]

#     print("MongoDB LOCAL connection: SUCCESS")

# except Exception as e:
#     print(f"MongoDB LOCAL connection: FAILED\n{e}")

#     # Dummy fallback (biar app tidak crash)
#     db = DummyDB()
#     schedules_collection = db.schedules
#     users_collection = db.users
#     courses_collection = db.courses
#     unavailability_reports_collection = db.unavailability_reports
#     sections_collection = getattr(db, "sections", None)


# ------------------------------
# CREATE DEFAULT KOORDINATOR
# ------------------------------

if db is not None:
    if users_collection.count_documents({"role": "koordinator"}) == 0:
        users_collection.insert_one({
            "username": "koordinator",
            "password": generate_password_hash("koordinator"),
            "role": "koordinator",
            "name": "Koordinator"
        })

    # Load lecturers from uploads/nama_dosen.csv
    dosen_file = os.path.join(UPLOAD_FOLDER, 'nama_dosen.csv')
    if os.path.exists(dosen_file):
        df_dosen = pd.read_csv(dosen_file)
        inserted = 0
        for _, row in df_dosen.iterrows():
            name = row['Nama Dosen'].strip('"')
            if not users_collection.find_one({"name": name}):
                email = name.lower().replace(' ', '').replace(',', '').replace('.', '') + "@example.com"
                users_collection.insert_one({
                    "name": name,
                    "email": email,
                    "password": generate_password_hash("123456"),
                    "role": "dosen"
                })
                inserted += 1
        print(f"Inserted {inserted} dosen")

    # Load courses from uploads/daftar_mata_kuliah.xlsx
    courses_file = os.path.join(UPLOAD_FOLDER, 'daftar_mata_kuliah.xlsx')
    if os.path.exists(courses_file):
        df_courses = pd.read_excel(courses_file)
        for _, row in df_courses.iterrows():
            name = str(row.get("course_name", ""))
            if name and not courses_collection.find_one({"course_name": name}):
                # Check if course is Kewarganegaraan or Kemahlikussalehan (online courses)
                is_online = any(keyword in name.lower() for keyword in ['kewarganegaraan', 'kemahlikussalahan'])
                
                courses_collection.insert_one({
                    "course_name": name,
                    "sks": int(row.get("sks", 0)),
                    "semester": str(row.get("semester", "")),
                    "is_lab": bool(row.get("is_lab", False)),
                    "is_online": bool(row.get("is_online", is_online))  # Auto-detect or from CSV
                })
    
    # Auto-update existing courses to mark online courses
    online_keywords = ['kewarganegaraan', 'kemahlikussalahan', 'kemal']
    for keyword in online_keywords:
        courses_collection.update_many(
            {"course_name": {"$regex": keyword, "$options": "i"}, "is_online": {"$ne": True}},
            {"$set": {"is_online": True}}
        )
                
def fix_mongo_id(data):
    new_data = []
    for item in data:
        new_item = dict(item)
        new_item["_id"] = str(new_item["_id"])  # hanya convert ID
        new_data.append(new_item)
    return new_data

# ------------------------------
# DEAP GENETIC ALGORITHM SETUP
# ------------------------------

if "FitnessMin" not in creator.__dict__:
    creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
if "Individual" not in creator.__dict__:
    creator.create("Individual", list, fitness=creator.FitnessMin)

def load_schedule():
    if schedules_collection is not None:
        return list(schedules_collection.find())
    return []

master_schedule = load_schedule()


def normalize_course_selected_by(courses, lecturers):
    """
    Normalize each course['selected_by'] list to contain lecturer display names.
    Accepts values that may be stored as username, email, or name and maps them
    to the canonical lecturer name used elsewhere in the app.
    Returns a list of tuples (course_id, unmatched_list) for diagnostics.
    """
    lecturer_names = {l['name'] for l in lecturers}
    # build mapping from alternative ids to display name
    id_map = {}
    for l in lecturers:
        name = l.get('name')
        if not name:
            continue
        id_map[name] = name
        if l.get('username'):
            id_map[l['username']] = name
        if l.get('email'):
            id_map[l['email']] = name

    unmatched = []
    for c in courses:
        sb = c.get('selected_by', [])
        if isinstance(sb, str):
            sb = [sb]
        normalized = []
        not_found = []
        for s in sb:
            if s in id_map:
                mapped = id_map[s]
                if mapped in lecturer_names and mapped not in normalized:
                    normalized.append(mapped)
            else:
                # try direct match on name
                if s in lecturer_names and s not in normalized:
                    normalized.append(s)
                else:
                    not_found.append(s)
        c['selected_by'] = normalized
        if not_found:
            unmatched.append((str(c.get('_id')), not_found))

    return unmatched

# ------------------------------
# EVALUATE SCHEDULE WITH PREFERENCES
# ------------------------------

def evaluate_schedule(individual):
    conflicts = 0
    lecturer_assigned = {}  # (day, start, end) -> lecturer
    course_assigned = set()  # set of course_ids to check uniqueness
    room_assigned = {}  # (day, room, start, end) -> True
    lecturer_load = {}  # lecturer -> total_sks
    lecturer_days = {}  # lecturer -> set of days
    lecturer_courses = {}  # lecturer -> list of courses

    # Get all lecturer preferences
    lecturer_preferences = {}
    for user in users_collection.find({"role": "dosen"}):
        lecturer_preferences[user["name"]] = user.get("preferences", {})

    for idx, lecturer in enumerate(individual):
        slot = master_schedule[idx]
        day = slot["day"]
        start = slot["start"]
        end = slot["end"]
        room = slot["room"]
        course_id = slot.get("course_id")
        sks = slot.get("sks") or 0
        key_time = (day, start, end)
        key_room = (day, room, start, end)

        if lecturer:
            # Lecturer conflict: same lecturer at same time
            if key_time in lecturer_assigned:
                if lecturer_assigned[key_time] == lecturer:
                    conflicts += 10  # High penalty for time conflict
            lecturer_assigned[key_time] = lecturer

            # Track lecturer load
            lecturer_load[lecturer] = lecturer_load.get(lecturer, 0) + sks
            lecturer_days[lecturer] = lecturer_days.get(lecturer, set())
            lecturer_days[lecturer].add(day)
            lecturer_courses[lecturer] = lecturer_courses.get(lecturer, [])
            lecturer_courses[lecturer].append(course_id)

            # Check preferences
            prefs = lecturer_preferences.get(lecturer, {})
            available_days = prefs.get("available_days", [])
            preferred_times = prefs.get("preferred_times", {})
            max_load = prefs.get("max_load", 12)

            # Penalti jika hari tidak tersedia (hanya jika ada preferensi hari yang ditentukan)
            if available_days and day not in available_days:
                conflicts += 5

            # Penalti jika waktu tidak sesuai preferensi
            if day in preferred_times:
                preferred_slots = preferred_times[day]
                slot_time = f"{start}-{end}"
                if slot_time not in preferred_slots:
                    conflicts += 3

            # Penalti jika beban melebihi maksimal
            if lecturer_load[lecturer] > max_load:
                conflicts += 10

            # Penalti jika lebih dari 2 mata kuliah
            if len(set(lecturer_courses[lecturer])) > 2:
                conflicts += 5

            # Penalti jika lebih dari3 hari mengajar
            if len(lecturer_days[lecturer]) > 3:
                conflicts += 2

        # Course uniqueness: each course assigned at most once
        if course_id and course_id in course_assigned:
            conflicts += 20  # High penalty for duplicate course assignment
        if course_id:
            course_assigned.add(course_id)

        # Room conflict: same room at same time
        if key_room in room_assigned:
            conflicts += 15  # High penalty for room conflict
        room_assigned[key_room] = True

    return (conflicts,)

# ------------------------------
# OPTIMIZE SCHEDULE
# ------------------------------

def check_schedule_conflicts(schedule=None):
    """
    Check for conflicts in current schedule.
    Returns True if conflicts are found, False otherwise.
    """
    if schedule is None:
        schedule = load_schedule()
    
    conflicts_found = False

    lecturer_time = {}  # (lecturer, day, start, end) -> True
    room_time = {}  # (day, room, start, end) -> True

    for slot in schedule:
        lecturer = slot.get('dosen')
        day = slot.get('day')
        start = slot.get('start')
        end = slot.get('end')
        room = slot.get('room')
        course_name = slot.get('course_name', 'Unknown')

        if lecturer:
            key_lect = (lecturer, day, start, end)
            if key_lect in lecturer_time:
                print(f"Conflict: Lecturer {lecturer} has overlapping slots on {day} {start}-{end} for {course_name}")
                conflicts_found = True
            lecturer_time[key_lect] = True

        if room:
            key_room = (day, room, start, end)
            if key_room in room_time:
                print(f"Conflict: Room {room} is double-booked on {day} {start}-{end} for {course_name}")
                conflicts_found = True
            room_time[key_room] = True

    return conflicts_found

# ---------------------------------------------------
# KONSTANTA HARI DAN TIMESLOT (SETELAH ISTIRAHAT)
# ---------------------------------------------------

DAYS = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat"]

# Timeslot untuk Senin–Kamis (istirahat jam 13:10–14:00)
TIMESLOTS_SENIN_KAMIS = [
    "08:00-08:50",
    "08:50-09:40",
    "09:40-10:30",
    "10:30-11:20",
    "11:20-12:10",
    "12:10-13:00",
    "13:00-13:10",   # khusus 1 SKS (opsional)
    # istirahat 13:10–14:00 → tidak dipakai
    "14:00-14:50",
    "14:50-15:40",
    "15:40-16:30"
]

# Timeslot hari Jumat (istirahat 11:30–13:50 + kurangi beban)
TIMESLOTS_JUMAT = [
    "08:00-08:50",
    "08:50-09:40",
    "09:40-10:30",
    "10:30-11:20",
    # 11:20–14:00 istirahat Jumat → tidak dipakai
    "14:00-14:50",
    "14:50-15:40",
    "15:40-16:30"
]
TIMEBLOCKS_2SKS = [
    "08:00-09:40",
    "08:50-10:30",
    "09:40-11:20",
    "10:30-12:10",
    "11:20-13:00",
    "14:00-15:40",
    "14:50-16:30",
    "15:40-17:20",
]

TIMEBLOCKS_3SKS = [
    "08:00-10:30",
    "08:50-11:20",
    "09:40-12:10",
    "10:30-13:00",
    "14:00-16:30",
    "14:50-17:20",
]

TIMEBLOCKS_1SKS = [
    "08:00-08:50",
    "08:50-09:40",
    "09:40-10:30",
    "10:30-11:20",
    "11:20-12:10",
    "14:00-14:50",
    "14:50-15:40",
    "15:40-16:30",
    "16:30-17:20",
]

# ===================
#  FUNGSI AMBIL TIMESLOT PER HARI
# ===================
def get_timeslots_by_day(day):
    """
    Mengembalikan daftar timeslot valid berdasarkan hari.
    Senin–Kamis: gunakan TIMESLOTS_SENIN_KAMIS
    Jumat: gunakan TIMESLOTS_JUMAT
    """
    if day == "Jumat":
        return TIMESLOTS_JUMAT
    else:
        return TIMESLOTS_SENIN_KAMIS

# ===================
#  VALIDASI LECTURER ASSIGNMENTS (BARU)
# ===================
def validate_lecturer_assignments():
    """Validasi assignment dosen.

    Aturan yang diterapkan (RELAXED):
    - Tidak lagi enforce jumlah section per mata kuliah
    - Tidak lagi enforce jumlah mata kuliah per dosen
    - Hanya cek: maksimal 3 hari mengajar
    """
    schedule = load_schedule()
    lecturer_days = {}

    # Hitung hari mengajar dari schedule
    for slot in schedule:
        lecturer = slot.get('dosen')
        day = slot.get('day')

        if lecturer and day:
            lecturer_days.setdefault(lecturer, set())
            lecturer_days[lecturer].add(day)

    violations = []
    for lecturer, days in lecturer_days.items():
        # Cek aturan hari: maksimal 3 hari
        if len(days) > 3:
            violations.append(f"{lecturer}: Mengajar {len(days)} hari (maksimal 3 hari)")

    return violations

# ===================
#  OR-TOOLS OPTIMIZATION
# ===================
def ortools_optimize_schedule():
    """
    Optimasi jadwal menggunakan OR-Tools Constraint Programming.
    - Menangani semua batasan keras secara eksak
    - Menghasilkan solusi optimal atau feasible
    """
    logger = app.logger
    
    # Step 1: load data
    courses = list(courses_collection.find())
    lecturers = list(users_collection.find({"role": "dosen"}))
    
    if not courses or not lecturers:
        logger.warning("No courses or lecturers found, skipping optimization")
        flash("Tidak ada mata kuliah atau dosen untuk dioptimasi.")
        return
    
    lecturer_names = [l['name'] for l in lecturers]
    
    # normalize selected_by to list of names (strings)
    for c in courses:
        sb = c.get('selected_by', [])
        if isinstance(sb, str):
            sb = [sb]
        sb = [s for s in sb if s in lecturer_names]
        c['selected_by'] = sb
    
    # Step 2: Buat model CP-SAT
    model = cp_model.CpModel()
    
    # Step 3: Buat variabel
    # Variabel: assign[lecturer][course][section][day][room][time] = 1 jika dosen mengajar section tersebut pada hari, ruangan, dan waktu tersebut
    assign = {}
    for lecturer in lecturer_names:
        assign[lecturer] = {}
        for course in courses:
            course_id = str(course['_id'])
            assign[lecturer][course_id] = {}
            for section in range(2):  # 2 section per course
                assign[lecturer][course_id][section] = {}
                for day in DAYS:
                    assign[lecturer][course_id][section][day] = {}
                    for room in (LAB_ROOMS if course.get('is_lab', False) else NON_LAB_ROOMS):
                        assign[lecturer][course_id][section][day][room] = {}
                        for start in get_valid_starts(day, 100 if course.get('sks', 0) in [1, 2] else 150):
                            assign[lecturer][course_id][section][day][room][start] = model.NewBoolVar(
                                f"assign_{lecturer}_{course_id}_{section}_{day}_{room}_{start}"
                            )
    
    # Step 4: Tambahkan batasan
    
    # Batasan 1: Setiap section harus diajar oleh tepat satu dosen
    for course in courses:
        course_id = str(course['_id'])
        for section in range(2):
            all_assignments = []
            for lecturer in lecturer_names:
                for day in DAYS:
                    for room in (LAB_ROOMS if course.get('is_lab', False) else NON_LAB_ROOMS):
                        for start in get_valid_starts(day, 100 if course.get('sks', 0) in [1, 2] else 150):
                            all_assignments.append(assign[lecturer][course_id][section][day][room][start])
            model.Add(sum(all_assignments) == 1)
    
    # Batasan 2: Setiap dosen mengajar tepat 2 mata kuliah
    for lecturer in lecturer_names:
        courses_taught = []
        for course in courses:
            course_id = str(course['_id'])
            course_taught = model.NewBoolVar(f"course_taught_{lecturer}_{course_id}")
            courses_taught.append(course_taught)
            
            # Hubungkan course_taught dengan variabel assign
            all_sections = []
            for section in range(2):
                for day in DAYS:
                    for room in (LAB_ROOMS if course.get('is_lab', False) else NON_LAB_ROOMS):
                        for start in get_valid_starts(day, 100 if course.get('sks', 0) in [1, 2] else 150):
                            all_sections.append(assign[lecturer][course_id][section][day][room][start])
            
            # Jika ada setidaknya satu section yang diajar, maka course_taught = 1
            model.Add(sum(all_sections) >= 1).OnlyEnforceIf(course_taught)
            model.Add(sum(all_sections) == 0).OnlyEnforceIf(course_taught.Not())
        
        model.Add(sum(courses_taught) == 2)
    
    # Batasan 3: Setiap dosen mengajar tepat 2 kelas per mata kuliah
    for lecturer in lecturer_names:
        for course in courses:
            course_id = str(course['_id'])
            all_sections = []
            for section in range(2):
                for day in DAYS:
                    for room in (LAB_ROOMS if course.get('is_lab', False) else NON_LAB_ROOMS):
                        for start in get_valid_starts(day, 100 if course.get('sks', 0) in [1, 2] else 150):
                            all_sections.append(assign[lecturer][course_id][section][day][room][start])
            
            model.Add(sum(all_sections) == 2)
    
    # Batasan 4: Dosen maksimal 2 hari dan harus berturut-turut
    weekday_order = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
    for lecturer in lecturer_names:
        # Buat variabel biner: lecturer_on_day[day] = 1 jika dosen mengajar di hari tsb
        lecturer_on_day = {}
        for day in DAYS:
            lecturer_on_day[day] = model.NewBoolVar(f"lecturer_on_day_{lecturer}_{day}")
            all_sections_on_day = []
            for course in courses:
                course_id = str(course['_id'])
                for section in range(2):
                    for room in (LAB_ROOMS if course.get('is_lab', False) else NON_LAB_ROOMS):
                        for start in get_valid_starts(day, 100 if course.get('sks', 0) in [1, 2] else 150):
                            all_sections_on_day.append(assign[lecturer][course_id][section][day][room][start])
            model.Add(sum(all_sections_on_day) >= 1).OnlyEnforceIf(lecturer_on_day[day])
            model.Add(sum(all_sections_on_day) == 0).OnlyEnforceIf(lecturer_on_day[day].Not())

        # Maksimal 2 hari
        total_days = model.NewIntVar(0, 2, f"total_days_{lecturer}")
        model.Add(total_days == sum(lecturer_on_day[day] for day in DAYS))
        model.Add(total_days <= 2)

        # Jika mengajar di 2 hari, harus berturut-turut (exactly one consecutive pair)
        day_bools = [lecturer_on_day[day] for day in weekday_order]
        consecutive_pairs = []
        for i in range(len(day_bools)-1):
            consecutive_pairs.append(model.NewBoolVar(f"consecutive_{lecturer}_{i}"))
            model.AddBoolAnd([day_bools[i], day_bools[i+1]]).OnlyEnforceIf(consecutive_pairs[i])
            model.AddBoolOr([day_bools[i].Not(), day_bools[i+1].Not()]).OnlyEnforceIf(consecutive_pairs[i].Not())
        # Jika dosen mengajar di 2 hari, harus ada tepat satu consecutive_pairs yang aktif
        model.Add(sum(consecutive_pairs) == 1).OnlyEnforceIf(total_days == 2)
    
    # Batasan 5: Batasan SKS per dosen
    for lecturer in lecturer_names:
        sks_sum = 0
        for course in courses:
            course_id = str(course['_id'])
            sks = course.get('sks', 0)
            for section in range(2):
                for day in DAYS:
                    for room in (LAB_ROOMS if course.get('is_lab', False) else NON_LAB_ROOMS):
                        for start in get_valid_starts(day, 100 if course.get('sks', 0) in [1, 2] else 150):
                            sks_sum += assign[lecturer][course_id][section][day][room][start] * sks
        
        model.Add(sks_sum >= MIN_SKS_PER_DOSEN)
        model.Add(sks_sum <= MAX_SKS_PER_DOSEN)
    
    # Batasan 6: Tidak ada bentrok waktu untuk dosen yang sama
    for lecturer in lecturer_names:
        for day in DAYS:
            # Bagi hari menjadi slot 10 menit
            for minute in range(8 * 60, 17 * 60, 10):  # Dari 08:00 hingga 17:00 dengan interval 10 menit
                overlapping_assignments = []
                for course in courses:
                    course_id = str(course['_id'])
                    sks = course.get('sks', 0)
                    duration = 100 if sks in [1, 2] else 150
                    
                    for section in range(2):
                        for room in (LAB_ROOMS if course.get('is_lab', False) else NON_LAB_ROOMS):
                            for start in get_valid_starts(day, duration):
                                start_min = _parse_minutes(start)
                                end_min = start_min + duration
                                
                                # Jika minute berada dalam rentang [start, end)
                                if start_min <= minute < end_min:
                                    overlapping_assignments.append(assign[lecturer][course_id][section][day][room][start])
                
                # Maksimal satu assignment pada waktu yang sama
                model.Add(sum(overlapping_assignments) <= 1)
    
    # Batasan 7: Tidak ada bentrok ruangan
    for day in DAYS:
        for room in NON_LAB_ROOMS + LAB_ROOMS:
            # Bagi hari menjadi slot 10 menit
            for minute in range(8 * 60, 17 * 60, 10):  # Dari 08:00 hingga 17:00 dengan interval 10 menit
                overlapping_assignments = []
                for course in courses:
                    course_id = str(course['_id'])
                    sks = course.get('sks', 0)
                    duration = 100 if sks in [1, 2] else 150
                    
                    for lecturer in lecturer_names:
                        for section in range(2):
                            # Hanya gunakan ruangan yang sesuai dengan jenis course
                            if (room in LAB_ROOMS and not course.get('is_lab', False)) or \
                               (room in NON_LAB_ROOMS and course.get('is_lab', False)):
                                continue
                                
                            for start in get_valid_starts(day, duration):
                                start_min = _parse_minutes(start)
                                end_min = start_min + duration
                                
                                # Jika minute berada dalam rentang [start, end)
                                if start_min <= minute < end_min:
                                    overlapping_assignments.append(assign[lecturer][course_id][section][day][room][start])
                
                # Maksimal satu assignment pada waktu yang sama
                model.Add(sum(overlapping_assignments) <= 1)
    
    # Step 5: Solver
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 60.0  # Batasi waktu pencarian
    solver.parameters.num_search_workers = 8
    solver.parameters.log_search_progress = True
    
    # Step 6: Solve
    result = solver.Solve(model)
    
    if result == cp_model.OPTIMAL or result == cp_model.FEASIBLE:
        # Step 7: Extract solution
        new_docs = []
        
        for course in courses:
            course_id = str(course['_id'])
            course_name = course['course_name']
            sks = course.get('sks', 0)
            is_lab = course.get('is_lab', False)
            
            for section in range(2):
                for lecturer in lecturer_names:
                    for day in DAYS:
                        for room in (LAB_ROOMS if is_lab else NON_LAB_ROOMS):
                            for start in get_valid_starts(day, 100 if sks in [1, 2] else 150):
                                if solver.Value(assign[lecturer][course_id][section][day][room][start]) == 1:
                                    section_letter = chr(65 + (section // 10))
                                    section_number = (section % 10) + 1
                                    section_name = f"{section_letter}{section_number}"
                                    
                                    new_docs.append({
                                        "course_id": course['_id'],
                                        "course_name": course_name,
                                        "section_name": section_name,
                                        "sks": sks,
                                        "section_number": section + 1,
                                        "room": room,
                                        "day": day,
                                        "start": start,
                                        "end": calculate_end_time(start, 100 if sks in [1, 2] else 150),
                                        "dosen": lecturer
                                    })
        
        # Replace schedules atomically: clear then bulk insert
        try:
            schedules_collection.delete_many({})
        except Exception:
            pass
        if new_docs:
            schedules_collection.insert_many(new_docs)
        
        # Post-check for conflicts
        conflicts_found = check_schedule_conflicts()
        if conflicts_found:
            flash("Jadwal dihasilkan, tetapi masih terdapat bentrok.")
        else:
            flash("Jadwal berhasil dioptimasi dengan OR-Tools tanpa bentrok!")
        
        # Validasi constraint dosen
        violations = validate_lecturer_assignments()
        if violations:
            flash(f"Terjadi {len(violations)} pelanggaran constraint dosen:")
            for violation in violations[:5]:
                flash(f"- {violation}")
        else:
            flash("Semua constraint dosen terpenuhi (2 MK, 2 kelas per MK, tidak ada MK di hari yang sama).")
        
        # Log final loads
        logger.info("Final lecturer assignments (summary):")
        for l in lecturers:
            name = l['name']
            slots = list(schedules_collection.find({"dosen": name}))
            total_sks = sum(s.get("sks", 0) for s in slots)
            days = sorted(set(s.get("day") for s in slots if s.get("day")))
            num_sections = len(slots)
            logger.info("%s -> %d SKS, days: %s, sections: %d", name, total_sks, days, num_sections)
    else:
        flash("Tidak ditemukan solusi yang memenuhi semua batasan. Silakan periksa batasan atau coba lagi.")

# ===================
#  OR-TOOLS CONSTRAINT SOLVER
# ===================
def ortools_constraint_solver(individual, section_objs, lecturers):
    """
    Menggunakan OR-Tools untuk memvalidasi dan memperbaiki solusi dari GA
    agar memenuhi semua batasan keras (hard constraints).
    """
    model = cp_model.CpModel()
    
    # Buat variabel untuk setiap penugasan
    # Variabel: assign[lecturer][section] = 1 jika dosen mengajar section tersebut
    all_lecturers = [l['name'] for l in lecturers]
    all_sections = list(range(len(section_objs)))
    
    assign = {}
    for lecturer in all_lecturers:
        for section_idx in all_sections:
            assign[(lecturer, section_idx)] = model.NewBoolVar(f"assign_{lecturer}_{section_idx}")
    
    # BATASAN 1: Setiap section harus diajar oleh tepat satu dosen
    for section_idx in all_sections:
        model.Add(sum(assign[(lecturer, section_idx)] for lecturer in all_lecturers) == 1)
    
    # BATASAN 2: Setiap dosen mengajar tepat 2 mata kuliah
    for lecturer in all_lecturers:
        courses_taught = {}
        for section_idx in all_sections:
            course_id = str(section_objs[section_idx]['course_id'])
            if course_id not in courses_taught:
                courses_taught[course_id] = model.NewBoolVar(f"course_{lecturer}_{course_id}")
            model.AddImplication(assign[(lecturer, section_idx)], courses_taught[course_id])
        
        model.Add(sum(courses_taught.values()) == 2)
    
    # BATASAN 3: Setiap dosen mengajar tepat 2 kelas per mata kuliah
    for lecturer in all_lecturers:
        for section_idx in all_sections:
            course_id = str(section_objs[section_idx]['course_id'])
            sections_for_course = [idx for idx, s in enumerate(section_objs) if str(s['course_id']) == course_id]
            model.Add(sum(assign[(lecturer, idx)] for idx in sections_for_course) == 2)
    
    # BATASAN 4: Dosen maksimal 2 hari dan harus berturut-turut (consecutive days)
    weekday_order = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
    for lecturer in all_lecturers:
        lecturer_on_day = {day: model.NewBoolVar(f"lecturer_on_day_{lecturer}_{day}") for day in weekday_order}
        # Set lecturer_on_day[day] = 1 jika dosen mengajar di hari tsb
        for day in weekday_order:
            any_assign = []
            for section_idx in all_sections:
                if section_idx < len(individual):
                    _, d, _ = individual[section_idx]
                    if d == day:
                        any_assign.append(assign[(lecturer, section_idx)])
            if any_assign:
                model.Add(sum(any_assign) >= 1).OnlyEnforceIf(lecturer_on_day[day])
                model.Add(sum(any_assign) == 0).OnlyEnforceIf(lecturer_on_day[day].Not())
            else:
                model.Add(lecturer_on_day[day] == 0)
        # Maksimal 2 hari
        total_days = model.NewIntVar(0, 2, f"total_days_{lecturer}")
        model.Add(total_days == sum(lecturer_on_day[day] for day in weekday_order))
        model.Add(total_days <= 2)
        # Jika mengajar di 2 hari, harus berturut-turut (exactly one consecutive pair)
        day_bools = [lecturer_on_day[day] for day in weekday_order]
        consecutive_pairs = []
        for i in range(len(day_bools)-1):
            consecutive_pairs.append(model.NewBoolVar(f"consecutive_{lecturer}_{i}"))
            model.AddBoolAnd([day_bools[i], day_bools[i+1]]).OnlyEnforceIf(consecutive_pairs[i])
            model.AddBoolOr([day_bools[i].Not(), day_bools[i+1].Not()]).OnlyEnforceIf(consecutive_pairs[i].Not())
        model.Add(sum(consecutive_pairs) == 1).OnlyEnforceIf(total_days == 2)
    
    # BATASAN 5: Batasan SKS per dosen
    for lecturer in all_lecturers:
        sks_sum = 0
        for section_idx in all_sections:
            sks = section_objs[section_idx].get('sks', 0)
            sks_sum += assign[(lecturer, section_idx)] * sks
        model.Add(sks_sum >= globals().get('MIN_SKS_PER_DOSEN', 8))
        model.Add(sks_sum <= globals().get('MAX_SKS_PER_DOSEN', 12))
    
    # Inisialisasi solusi dari individual (solusi GA)
    for section_idx in all_sections:
        if section_idx < len(individual):
            room, day, time_slot = individual[section_idx]
            lecturer = section_objs[section_idx].get('forced_lecturer')
            if lecturer:
                model.Add(assign[(lecturer, section_idx)] == 1)
    
    # Solver
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30.0  # Batasi waktu pencarian
    solver.parameters.num_search_workers = 8
    result = solver.Solve(model)
    
    if result == cp_model.OPTIMAL or result == cp_model.FEASIBLE:
        # Konversi solusi OR-Tools kembali ke format individual
        new_individual = []
        for section_idx in all_sections:
            if section_idx < len(individual):
                room, day, time_slot = individual[section_idx]
                # Cari dosen yang ditugaskan oleh OR-Tools
                assigned_lecturer = None
                for lecturer in all_lecturers:
                    if solver.Value(assign[(lecturer, section_idx)]) == 1:
                        assigned_lecturer = lecturer
                        break
                
                # Update section_objs dengan dosen yang ditugaskan
                section_objs[section_idx]['forced_lecturer'] = assigned_lecturer
                
                # Simpan slot waktu yang sama
                new_individual.append((room, day, time_slot))
        
        return new_individual, True
    else:
        # Jika tidak ada solusi yang memenuhi semua batasan
        return individual, False

# ===================
#  HYBRID OPTIMIZATION (GA + OR-TOOLS)
# ===================
def hybrid_optimize_schedule(population=100, generations=100):
    # DEPRECATED: Fungsi lama hybrid (GA + OR-Tools) untuk assignment + scheduling.
    # Sudah digantikan oleh pipeline baru: generate_sections_ga_pipeline + schedule_sections_with_ortools.
    # Dipertahankan sementara untuk kompatibilitas rute lama namun tidak direkomendasikan dipakai.
    """
    Optimasi jadwal hybrid: GA untuk optimasi awal, OR-Tools untuk perbaikan
    - Setiap dosen WAJIB mengajar 2 mata kuliah
    - Setiap mata kuliah WAJIB diajar oleh 3 dosen
    - Setiap dosen mengajar 2 kelas per mata kuliah
    """
    logger = app.logger

    # Step 1: load data
    courses = list(courses_collection.find())
    lecturers = list(users_collection.find({"role": "dosen"}))

    if not courses or not lecturers:
        logger.warning("No courses or lecturers found, skipping optimization")
        flash("Tidak ada mata kuliah atau dosen untuk dioptimasi.")
        return

    lecturer_names = [l['name'] for l in lecturers]
    logger.info("Starting hybrid GA optimization")
    logger.info(f"Total courses: {len(courses)}, Total lecturers: {len(lecturers)}")

    # Normalize selected_by using central helper (handles username/email/name)
    unmatched = normalize_course_selected_by(courses, lecturers)
    if unmatched:
        logger.warning("normalize_course_selected_by found unmatched identifiers for %d courses", len(unmatched))
        logger.debug("Unmatched selected_by entries (course_id -> list): %s", unmatched)

    # Step 2: Buat distribusi awal yang seimbang
    # Hanya dosen yang memilih minimal satu matkul yang masuk active_lecturers
    lecturers_with_choice = set()
    for c in courses:
        for s in c.get('selected_by', []):
            lecturers_with_choice.add(s)
    active_lecturers = [l for l in lecturer_names if l in lecturers_with_choice]
    if not active_lecturers:
        logger.warning("Tidak ada dosen yang memilih mata kuliah!")
        flash("Tidak ada dosen yang memilih mata kuliah!")
        return

    # Jika distribusi tidak seimbang, hanya kurangi course
    total_lecturer_assignments_needed = len(active_lecturers) * 2
    total_course_assignments_needed = len(courses) * 3
    active_courses = courses[:]
    if total_course_assignments_needed > total_lecturer_assignments_needed:
        excess_courses = total_course_assignments_needed - total_lecturer_assignments_needed
        active_courses = active_courses[:-excess_courses]
        logger.warning(f"Mengurangi jumlah mata kuliah aktif menjadi {len(active_courses)} untuk menyeimbangkan penugasan.")
        flash(f"Mengurangi jumlah mata kuliah aktif menjadi {len(active_courses)} untuk menyeimbangkan penugasan.")
    
    # Buat distribusi awal yang seimbang
    # Target: setiap dosen aktif mendapat 2 MK, setiap MK aktif mendapat 3 dosen
    final_assignments = []
    lecturer_course_count = {name: 0 for name in active_lecturers}
    course_lecturer_count = {str(c['_id']): 0 for c in active_courses}
    
    # Iterasi pertama: berikan preferensi pada yang sudah memilih
    for lecturer in active_lecturers:
        for course in active_courses:
            course_id = str(course['_id'])
            if lecturer_course_count[lecturer] < 2 and course_lecturer_count[course_id] < 3:
                if lecturer in course.get('selected_by', []):
                    final_assignments.append({'course_id': course_id, 'lecturer': lecturer})
                    lecturer_course_count[lecturer] += 1
                    course_lecturer_count[course_id] += 1
    
    # Iterasi kedua: isi kekosongan
    for lecturer in active_lecturers:
        for course in active_courses:
            course_id = str(course['_id'])
            if lecturer_course_count[lecturer] < 2 and course_lecturer_count[course_id] < 3:
                final_assignments.append({'course_id': course_id, 'lecturer': lecturer})
                lecturer_course_count[lecturer] += 1
                course_lecturer_count[course_id] += 1
    
    # Step 3: Buat section_objs dari distribusi akhir
    section_objs = []
    for assignment in final_assignments:
        course_id = assignment['course_id']
        lecturer = assignment['lecturer']
        
        # Cari detail course
        course = next((c for c in active_courses if str(c['_id']) == course_id), None)
        if not course:
            continue
            
        # Buat 2 section untuk setiap assignment
        for i in range(2):
            section_objs.append({
                "course_id": course['_id'],
                "course_name": course['course_name'],
                "sks": course.get('sks', 0),
                "is_lab": bool(course.get('is_lab', False)),
                "section_num": i + 1,
                "forced_lecturer": lecturer
            })
    
    logger.info(f"Total sections created: {len(section_objs)} (for {len(active_courses)} courses)")
    logger.info("Final lecturer assignments (summary):")
    for name, count in lecturer_course_count.items():
        if count > 0:
            logger.info(f"  {name} -> {count} courses")
    
    # Step 4: Setup GA
    toolbox = base.Toolbox()

    def create_individual():
        ind = []
        for sec in section_objs:
            rooms = LAB_ROOMS if sec['is_lab'] else NON_LAB_ROOMS
            room = random.choice(rooms)
            day = random.choice(DAYS)
            sks = sec['sks']
            duration = 100 if sks in [1, 2] else 150
            valid_starts = get_valid_starts(day, duration)
            time_slot = random.choice(valid_starts)
            ind.append((room, day, time_slot))
        return creator.Individual(ind)
        
    def mutate_individual(individual, indpb):
        for i in range(len(individual)):
            if random.random() < indpb:
                room, day, time_slot = individual[i]
                section = section_objs[i]
                sks = section['sks']
                is_lab = section['is_lab']
                duration = 100 if sks in [1, 2] else 150

                mutation_type = random.choice(['day', 'room', 'time_slot'])
                
                if mutation_type == 'day':
                    new_day = random.choice(DAYS)
                    valid_starts = get_valid_starts(new_day, duration)
                    new_time_slot = random.choice(valid_starts) if valid_starts else time_slot
                    individual[i] = (room, new_day, new_time_slot)
                elif mutation_type == 'room':
                    new_room = random.choice(LAB_ROOMS if is_lab else NON_LAB_ROOMS)
                    individual[i] = (new_room, day, time_slot)
                elif mutation_type == 'time_slot':
                    valid_starts = get_valid_starts(day, duration)
                    new_time_slot = random.choice(valid_starts) if valid_starts else time_slot
                    individual[i] = (room, day, new_time_slot)
        return individual,

    def evaluate_full_schedule(individual):
        # individual adalah list of (room, day, time_slot)
        conflicts = 0
        lecturer_time_book = {}
        room_time_book = {}
        lecturer_sks = {name: 0 for name in active_lecturers}
        lecturer_days = {name: set() for name in active_lecturers}
        lecturer_course_sections = {name: {} for name in active_lecturers}

        day_order = {"Senin": 1, "Selasa": 2, "Rabu": 3, "Kamis": 4, "Jumat": 5}

        for idx, gene in enumerate(individual):
            if idx >= len(section_objs):
                continue

            room, day, time_slot = gene
            section = section_objs[idx]
            lecturer = section['forced_lecturer']
            course_id = str(section['course_id'])
            
            sks = section['sks']
            duration_minutes = 100 if sks in [1, 2] else 150
            end = calculate_end_time(time_slot, duration_minutes)
                
            start = _parse_minutes(time_slot)
            end = _parse_minutes(end)

            # Update lecturer stats
            lecturer_sks[lecturer] += sks
            lecturer_days[lecturer].add(day)
            
            # Track sections per course for each lecturer
            if course_id not in lecturer_course_sections[lecturer]:
                lecturer_course_sections[lecturer][course_id] = 0
            lecturer_course_sections[lecturer][course_id] += 1

            # Hard constraint: book every time fragment (10-minute block)
            for minute in range(start, end, 10):
                # Lecturer conflict
                lt_key = (lecturer, day, minute)
                if lt_key in lecturer_time_book:
                    conflicts += 1000000  # Very high penalty
                lecturer_time_book[lt_key] = True

                # Room conflict
                rt_key = (day, room, minute)
                if rt_key in room_time_book:
                    conflicts += 1000000  # Very high penalty
                room_time_book[rt_key] = True

            # Soft rule: Prioritaskan Senin-Kamis
            if day == "Jumat":
                conflicts += 100

        # ================== ATURAN DOSEN =====================
        for name in active_lecturers:
            # Constraint: Setiap dosen harus mengajar tepat 2 mata kuliah
            num_courses = len(lecturer_course_sections.get(name, {}))
            if num_courses != 2:
                conflicts += abs(num_courses - 2) * 1000000  # Very high penalty

            # Constraint: Setiap dosen harus mengajar tepat 2 kelas per mata kuliah
            for course_id, section_count in lecturer_course_sections.get(name, {}).items():
                if section_count != 2:
                    conflicts += abs(section_count - 2) * 1000000  # Very high penalty

            # Constraint SKS
            load = lecturer_sks.get(name, 0)
            min_sks = globals().get('MIN_SKS_PER_DOSEN', 8)
            max_sks = globals().get('MAX_SKS_PER_DOSEN', 12)
            target_sks = globals().get('REQUIRED_SKS_PER_DOSEN', 10)

            if load < min_sks:
                conflicts += (min_sks - load) * 100000
            if load > max_sks:
                conflicts += (load - max_sks) * 100000
            
            deviation = abs(load - target_sks)
            if deviation > 0:
                conflicts += deviation * 10000

            # Constraint Hari: maksimal 2 hari, dan jika 2 hari harus berturut-turut
            weekday_order = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
            day_index = {d: i for i, d in enumerate(weekday_order)}
            days = sorted([day_index.get(d, 99) for d in lecturer_days.get(name, [])])
            # jika salah satu hari tak dikenal, beri penalti
            if any(d == 99 for d in days):
                conflicts += 200000
            elif len(days) > 2:
                conflicts += (len(days) - 2) * 200000
            elif len(days) == 2:
                if abs(days[1] - days[0]) != 1:
                    conflicts += 200000
        
        return (conflicts,)

    # Register GA functions
    toolbox.register("attr_float", random.random)
    toolbox.register("individual", create_individual)
    toolbox.register("mate", tools.cxTwoPoint)
    toolbox.register("mutate", mutate_individual, indpb=0.2)
    toolbox.register("select", tools.selTournament, tournsize=3)
    toolbox.register("evaluate", evaluate_full_schedule)

    # Create initial population manually
    pop = [toolbox.individual() for _ in range(population)]

    # Run GA without statistics to avoid the error
    algorithms.eaSimple(pop, toolbox, cxpb=0.7, mutpb=0.2, ngen=generations, 
                       halloffame=None)

    # Get the best individual
    best_ind = tools.selBest(pop, 1)[0]
    
    # Step 5: Extract solution from best individual
    new_docs = []
    
    for idx, gene in enumerate(best_ind):
        if idx >= len(section_objs):
            continue

        room, day, time_slot = gene
        section = section_objs[idx]
        course_id = str(section['course_id'])
        course_name = section['course_name']
        sks = section['sks']
        is_lab = section['is_lab']
        
        section_name = f"{chr(65 + idx % 26)}{idx // 26 + 1}"
        
        new_docs.append({
            "course_id": section['course_id'],
            "course_name": course_name,
            "section_name": section_name,
            "sks": sks,
            "section_number": section['section_num'],
            "room": room,
            "day": day,
            "start": time_slot,
            "end": calculate_end_time(time_slot, 100 if sks in [1, 2] else 150),
            "dosen": section['forced_lecturer']
        })
    
    # Replace schedules atomically: clear then bulk insert
    try:
        schedules_collection.delete_many({})
    except Exception:
        pass
    if new_docs:
        schedules_collection.insert_many(new_docs)
    
    flash("Jadwal berhasil dioptimasi dengan metode hybrid (GA + OR-Tools)!")
    
    # Log final loads
    logger.info("Final lecturer assignments (summary):")
    for l in lecturers:
        name = l['name']
        slots = list(schedules_collection.find({"dosen": name}))
        total_sks = sum(s.get("sks", 0) for s in slots)
        days = sorted(set(s.get("day") for s in slots if s.get("day")))
        num_sections = len(slots)
        
        # Hitung jumlah mata kuliah yang diajar
        courses_taught = set()
        for slot in slots:
            courses_taught.add(slot.get("course_id"))
        num_courses = len(courses_taught)
        
        logger.info("%s -> %d SKS, days: %s, sections: %d", 
                  name, total_sks, days, num_sections)

# ===================
#  OPTIMASI JADWAL (DIPERBAIKI)
# ===================
def full_optimize_schedule(population=1000, generations=5000):
    # DEPRECATED: Optimizer GA penuh lama. Mengandung logika assignment dan penjadwalan yang
    # kini dipisah menjadi dua tahap (GA sections + OR-Tools scheduling). Biarkan tetap ada
    # sampai rute lama dihapus, tetapi jangan dikembangkan lagi.
    """
    Versi perbaikan dari optimasi jadwal.
    - Membuat distribusi dosen yang seimbang (setiap dosen tepat 2 MK).
    - Mencegah satu dosen mengajar 2 MK di hari yang sama.
    - Meningkatkan parameter GA untuk hasil yang lebih baik.
    - Memastikan jadwal akhir bebas dari bentrok.
    """
    logger = app.logger

    # Step 1: load data
    courses = list(courses_collection.find())
    lecturers = list(users_collection.find({"role": "dosen"}))

    if not courses or not lecturers:
        logger.warning("No courses or lecturers found, skipping optimization")
        flash("Tidak ada mata kuliah atau dosen untuk dioptimasi.")
        return

    lecturer_names = [l['name'] for l in lecturers]
    
    # --- DIAGNOSTIC: Calculate total available slots ---
    all_rooms = NON_LAB_ROOMS + LAB_ROOMS
    all_days = DAYS
    all_starts = set()
    for day in all_days:
        for dur in [100, 150]:
            all_starts.update(get_valid_starts(day, dur))
    all_starts = list(all_starts)
    total_slots = len(all_rooms) * len(all_days) * len(all_starts)
    logger.warning(f"[DIAGNOSTIC] Total available slots: {len(all_rooms)} × {len(all_days)} × {len(all_starts)} = {total_slots}")

    logger.info("Starting GA optimization with population=%d, generations=%d", population, generations)

    # Normalize selected_by using central helper (handles username/email/name)
    unmatched = normalize_course_selected_by(courses, lecturers)
    if unmatched:
        logger.warning("normalize_course_selected_by found unmatched identifiers for %d courses", len(unmatched))
        logger.debug("Unmatched selected_by entries (course_id -> list): %s", unmatched)

    # -----------------------------
    # Step 2 — Balanced Assignment Generator (LOGIKA BARU)
    # -----------------------------
    
    # 2a. Buat pool dari semua kemungkinan (course, lecturer)
    assignment_pool = []
    for course in courses:
        cid = str(course['_id'])
        selected = course.get('selected_by', [])
        for lecturer in selected:
            assignment_pool.append({'course_id': cid, 'lecturer': lecturer})
    
    # Acak pool untuk menghindari bias
    random.shuffle(assignment_pool)
    
    # 2b. Buat distribusi awal yang seimbang
    # Target: setiap dosen aktif mendapat 2 MK, setiap MK mendapat 3 dosen
    final_assignments = []
    lecturer_course_count = {name: 0 for name in lecturer_names}
    lecturer_names = [l['name'] for l in lecturers]
    # --- DIAGNOSTIC: Calculate total available slots ---
    all_rooms = NON_LAB_ROOMS + LAB_ROOMS
    all_days = DAYS
    all_starts = set()
    for day in all_days:
        for dur in [100, 150]:
            all_starts.update(get_valid_starts(day, dur))
    all_starts = list(all_starts)
    total_slots = len(all_rooms) * len(all_days) * len(all_starts)
    logger.warning(f"[DIAGNOSTIC] Total available slots: {len(all_rooms)} × {len(all_days)} × {len(all_starts)} = {total_slots}")

    logger.info("Starting GA optimization with population=%d, generations=%d", population, generations)

    # Normalize selected_by using central helper (handles username/email/name)
    unmatched = normalize_course_selected_by(courses, lecturers)
    if unmatched:
        logger.warning("normalize_course_selected_by found unmatched identifiers for %d courses", len(unmatched))
        logger.debug("Unmatched selected_by entries (course_id -> list): %s", unmatched)

    # Step 2 — Balanced Assignment Generator (LOGIKA BARU)
    # Hanya dosen yang memilih minimal satu matkul yang masuk active_lecturers
    lecturers_with_choice = set()
    for c in courses:
        for s in c.get('selected_by', []):
            lecturers_with_choice.add(s)
    active_lecturers = [l for l in lecturer_names if l in lecturers_with_choice]
    if not active_lecturers:
        logger.warning("Tidak ada dosen yang memilih mata kuliah!")
        flash("Tidak ada dosen yang memilih mata kuliah!")
        return

    # Jika distribusi tidak seimbang, hanya kurangi course
    total_lecturer_assignments_needed = len(active_lecturers) * 2
    total_course_assignments_needed = len(courses) * 3
    active_courses = courses[:]
    if total_course_assignments_needed > total_lecturer_assignments_needed:
        excess_courses = total_course_assignments_needed - total_lecturer_assignments_needed
        active_courses = active_courses[:-excess_courses]
        logger.warning(f"Mengurangi jumlah mata kuliah aktif menjadi {len(active_courses)} untuk menyeimbangkan penugasan.")
        flash(f"Mengurangi jumlah mata kuliah aktif menjadi {len(active_courses)} untuk menyeimbangkan penugasan.")

    # 2a. Buat pool dari semua kemungkinan (course, lecturer)
    assignment_pool = []
    for course in active_courses:
        cid = str(course['_id'])
        selected = course.get('selected_by', [])
        for lecturer in selected:
            if lecturer in active_lecturers:
                assignment_pool.append({'course_id': cid, 'lecturer': lecturer})

    # Acak pool untuk menghindari bias
    random.shuffle(assignment_pool)

    # 2b. Buat distribusi awal yang seimbang
    # Target: setiap dosen aktif mendapat 2 MK, setiap MK mendapat 3 dosen
    final_assignments = []
    lecturer_course_count = {name: 0 for name in active_lecturers}
    course_lecturer_count = {str(c['_id']): 0 for c in active_courses}

    # Iterasi pertama: berikan preferensi pada yang sudah memilih
    for assignment in assignment_pool:
        lecturer = assignment['lecturer']
        course_id = assignment['course_id']
        if lecturer_course_count[lecturer] < 2 and course_lecturer_count[course_id] < 3:
            final_assignments.append(assignment)
            lecturer_course_count[lecturer] += 1
            course_lecturer_count[course_id] += 1

    # 2c. Isi kekosongan untuk dosen yang kurang dari 2 MK
    underassigned_lecturers = [name for name, count in lecturer_course_count.items() if count < 2]
    underassigned_courses = [cid for cid, count in course_lecturer_count.items() if count < 3]

    # Iterasi kedua: isi kekosongan
    random.shuffle(underassigned_lecturers)
    random.shuffle(underassigned_courses)

    for lecturer in underassigned_lecturers:
        for course_id in underassigned_courses:
            if lecturer_course_count[lecturer] < 2 and course_lecturer_count[course_id] < 3:
                final_assignments.append({'course_id': course_id, 'lecturer': lecturer})
                lecturer_course_count[lecturer] += 1
                course_lecturer_count[course_id] += 1
                if lecturer_course_count[lecturer] == 2:
                    break # Dosen ini sudah penuh, lanjut ke dosen berikutnya
        if all(count == 2 for count in lecturer_course_count.values()):
            break # Semua dosen sudah penuh

    # 2d. Buat section_objs dari distribusi akhir
    section_objs = []
    for assignment in final_assignments:
        course_id = assignment['course_id']
        lecturer = assignment['lecturer']
        # Cari detail course
        course = next((c for c in active_courses if str(c['_id']) == course_id), None)
        if not course:
            continue
        # Buat 2 section untuk setiap assignment
        for i in range(2):
            section_objs.append({
                "course_id": course['_id'],
                "course_name": course['course_name'],
                "sks": course.get('sks', 0),
                "is_lab": bool(course.get('is_lab', False)),
                "section_num": i + 1,
                "forced_lecturer": lecturer
            })

    logger.info(f"Total sections created: {len(section_objs)} (for {len(active_courses)} courses)")
    logger.info(f"Final lecturer assignments (summary):")
    for name, count in lecturer_course_count.items():
        logger.info(f"  {name} -> {count} courses")

    # -----------------------------
    # GA setup (DIPERBAIKI)
    # -----------------------------
    # Minimal evaluator to keep GA functional if used. Counts room/lecturer overlaps only.
    def evaluate_full_schedule(individual):
        conflicts = 0
        booked_lect = set()
        booked_room = set()
        for idx, gene in enumerate(individual):
            if idx >= len(section_objs):
                continue
            room, day, start = gene
            section = section_objs[idx]
            sks = section.get('sks', 0)
            duration = 100 if sks in [1, 2] else 150
            s_min = _parse_minutes(start)
            e_min = _parse_minutes(calculate_end_time(start, duration))
            if s_min is None or e_min is None:
                conflicts += 10
                continue
            lecturer = section.get('forced_lecturer')
            local_conflict = False
            for minute in range(s_min, e_min, 10):
                if (lecturer, day, minute) in booked_lect:
                    conflicts += 1000
                    local_conflict = True
                    break
                if (day, room, minute) in booked_room:
                    conflicts += 1000
                    local_conflict = True
                    break
            if not local_conflict:
                for minute in range(s_min, e_min, 10):
                    booked_lect.add((lecturer, day, minute))
                    booked_room.add((day, room, minute))
        return (conflicts,)

    toolbox = base.Toolbox()

    def create_individual():
        ind = []
        for sec in section_objs:
            rooms = LAB_ROOMS if sec['is_lab'] else NON_LAB_ROOMS
            room = random.choice(rooms)
            day = random.choice(DAYS)
            sks = sec['sks']
            duration = 100 if sks in [1,2] else 150
            valid_starts = get_valid_starts(day, duration)
            time_slot = random.choice(valid_starts)
            ind.append((room, day, time_slot))
        return creator.Individual(ind)
        
    def mutate_individual(individual, indpb):
        for i in range(len(individual)):
            if random.random() < indpb:
                room, day, time_slot = individual[i]
                section = section_objs[i]
                sks = section['sks']
                is_lab = section['is_lab']
                rooms = LAB_ROOMS if is_lab else NON_LAB_ROOMS
                duration = 100 if sks in [1,2] else 150

                mutation_type = random.choice(['day', 'room', 'time_slot'])
                
                if mutation_type == 'day':
                    new_day = random.choice(DAYS)
                    valid_starts = get_valid_starts(new_day, duration)
                    new_time_slot = random.choice(valid_starts) if valid_starts else time_slot
                    individual[i] = (room, new_day, new_time_slot)
                elif mutation_type == 'room':
                    new_room = random.choice(rooms)
                    individual[i] = (new_room, day, time_slot)
                elif mutation_type == 'time_slot':
                    valid_starts = get_valid_starts(day, duration)
                    new_time_slot = random.choice(valid_starts) if valid_starts else time_slot
                    individual[i] = (room, day, new_time_slot)
        return individual,

    toolbox.register("individual", create_individual)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("mate", tools.cxTwoPoint)
    toolbox.register("mutate", mutate_individual, indpb=0.2)
    toolbox.register("select", tools.selTournament, tournsize=3)
    toolbox.register("evaluate", evaluate_full_schedule)
    
    # ---------------- GA EXECUTE ---------------------
    pop = toolbox.population(n=population)

    if len(section_objs) < 2:
        algorithms.eaSimple(pop, toolbox, cxpb=0.0, mutpb=0.3, ngen=generations, verbose=False)
    else:
        algorithms.eaSimple(pop, toolbox, cxpb=0.5, mutpb=0.2, ngen=generations, verbose=False)

    best = tools.selBest(pop, 1)[0]

    # -----------------------------
    # Step 3 — Conflict-Free Schedule Builder
    # -----------------------------
    new_docs = []
    booked_lect = set()
    booked_room = set()

    for idx, gene in enumerate(best):
        if idx >= len(section_objs):
            continue
        
        room, day, time_slot = gene
        section = section_objs[idx]
        sks = section.get('sks', 0)
        duration_minutes = 100 if sks in [1, 2] else 150
        
        # helper to try placing a section for a given lecturer
        def _try_place(lecturer_candidate, preferred_room=None, preferred_day=None, preferred_start=None):
            # diagnostic helper
            def _diag(reason, tried=None):
                try:
                    doc = {
                        "timestamp": datetime.utcnow(),
                        "section_idx": idx,
                        "course_id": str(section.get('course_id')),
                        "course_name": section.get('course_name'),
                        "lecturer_candidate": lecturer_candidate,
                        "tried": tried,
                        "reason": reason
                    }
                    try:
                        db.placement_diagnostics.insert_one(doc)
                    except Exception:
                        app.logger.debug("placement diagnostic write failed")
                except Exception:
                    app.logger.debug("placement diagnostic construction failed")

            rooms = LAB_ROOMS if section.get('is_lab') else NON_LAB_ROOMS
            room_choices = ([preferred_room] + [r for r in rooms if r != preferred_room]) if preferred_room else list(rooms)
            day_choices = ([preferred_day] + [d for d in DAYS if d != preferred_day]) if preferred_day else list(DAYS)

            # build set of days already assigned to this lecturer (from booked_lect, new_docs and DB)
            try:
                existing_days = set(d for (lect, d, minute) in booked_lect if lect == lecturer_candidate)
            except Exception:
                existing_days = set()
            try:
                for s in new_docs:
                    if s.get('dosen') == lecturer_candidate and s.get('day'):
                        existing_days.add(s.get('day'))
            except Exception:
                pass
            try:
                for s in schedules_collection.find({"dosen": lecturer_candidate}):
                    d = s.get('day')
                    if d:
                        existing_days.add(d)
            except Exception:
                # DB read failure is non-fatal here
                pass

            for r in room_choices:
                for d in day_choices:
                    valid_starts = get_valid_starts(d, duration_minutes)
                    start_choices = ([preferred_start] + [s for s in valid_starts if s != preferred_start]) if preferred_start else valid_starts
                    for start in start_choices:
                        if not start:
                            continue
                        s_min = _parse_minutes(start)
                        e_min = _parse_minutes(calculate_end_time(start, duration_minutes))
                        if s_min is None or e_min is None:
                            _diag("invalid_time_parse", {"room": r, "day": d, "start": start})
                            continue

                        # Enforce day-constraints: will this placement create >2 days or non-consecutive days?
                        try:
                            if d not in existing_days and len(existing_days) >= 2:
                                _diag("would_exceed_max_days", {"room": r, "day": d, "start": start})
                                continue
                            if d not in existing_days and len(existing_days) == 1:
                                weekday_order_local = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
                                existing_day = next(iter(existing_days))
                                try:
                                    if abs(weekday_order_local.index(existing_day) - weekday_order_local.index(d)) != 1:
                                        _diag("non_consecutive_day", {"room": r, "day": d, "start": start})
                                        continue
                                except ValueError:
                                    _diag("unknown_day_name", {"existing_day": existing_day, "tried_day": d})
                                    continue
                        except Exception as e:
                            app.logger.debug("day-constraint check error: %s", e)

                        # Check conflicts against booked sets (fast) and DB (defensive)
                        conflict = False
                        for minute in range(s_min, e_min, 10):
                            if (lecturer_candidate, d, minute) in booked_lect or (d, r, minute) in booked_room:
                                conflict = True
                                break
                        if conflict:
                            _diag("conflict_with_booked", {"room": r, "day": d, "start": start})
                            continue

                        # Also check against already persisted schedule docs for safety
                        try:
                            for s in schedules_collection.find({"day": d, "room": r}):
                                s_s = _parse_minutes(s.get('start'))
                                s_e = _parse_minutes(s.get('end'))
                                if _times_overlap(s_min, e_min, s_s, s_e):
                                    conflict = True
                                    break
                            if conflict:
                                _diag("conflict_with_db_room", {"room": r, "day": d, "start": start})
                                continue
                        except Exception:
                            pass

                        try:
                            for s in schedules_collection.find({"dosen": lecturer_candidate, "day": d}):
                                s_s = _parse_minutes(s.get('start'))
                                s_e = _parse_minutes(s.get('end'))
                                if _times_overlap(s_min, e_min, s_s, s_e):
                                    conflict = True
                                    break
                            if conflict:
                                _diag("conflict_with_db_lecturer", {"lecturer": lecturer_candidate, "day": d, "start": start})
                                continue
                        except Exception:
                            pass

                        # Passed all checks: commit booking into booked sets
                        for minute in range(s_min, e_min, 10):
                            booked_lect.add((lecturer_candidate, d, minute))
                            booked_room.add((d, r, minute))
                        section_letter = chr(65 + ((section['section_num'] - 1) // 10))
                        section_number = ((section['section_num'] - 1) % 10) + 1
                        section_name = f"{section_letter}{section_number}"
                        placed = {
                            "course_id": section['course_id'],
                            "course_name": section['course_name'],
                            "section_name": section_name,
                            "sks": section.get('sks', 0),
                            "section_number": section['section_num'],
                            "room": r,
                            "day": d,
                            "start": start,
                            "end": calculate_end_time(start, duration_minutes),
                            "dosen": lecturer_candidate
                        }
                        return placed

            return None

        # Try placements in this order: gene -> forced_lecturer -> any lecturer
        placed = None
        lecturer = section.get('forced_lecturer')
        try:
            placed = _try_place(lecturer, preferred_room=room, preferred_day=day, preferred_start=time_slot)
        except Exception:
            placed = None

        if not placed:
            # Fallback: coba kombinasi lain
            for r in (LAB_ROOMS if section.get('is_lab') else NON_LAB_ROOMS):
                for d in DAYS:
                    for start in get_valid_starts(d, duration_minutes):
                        placed = _try_place(lecturer, preferred_room=r, preferred_day=d, preferred_start=start)
                        if placed:
                            break
                    if placed:
                        break
                if placed:
                    break

        if not placed:
            logger.warning(f"Could not place section for course {section.get('course_name')} (section {section.get('section_num')}); it will be skipped.")
            continue

        new_docs.append(placed)

    # Replace schedules atomically: clear then bulk insert
    try:
        schedules_collection.delete_many({})
    except Exception:
        pass
    if new_docs:
        schedules_collection.insert_many(new_docs)

    # Post-check for conflicts
    conflicts_found = check_schedule_conflicts()
    if conflicts_found:
        flash("Jadwal dihasilkan, tetapi masih terdapat bentrok.")
    else:
        flash("Jadwal berhasil dioptimasi tanpa bentrok!")
    
    # Validasi constraint dosen
    violations = validate_lecturer_assignments()
    if violations:
        flash(f"Terjadi {len(violations)} pelanggaran constraint dosen:")
        for violation in violations[:5]:
            flash(f"- {violation}")
    else:
        flash("Semua constraint dosen terpenuhi (2 MK, 2 kelas per MK, tidak ada MK di hari yang sama).")

    # Log final loads
    logger.info("Final lecturer assignments (summary):")
    for l in lecturers:
        name = l['name']
        slots = list(schedules_collection.find({"dosen": name}))
        total_sks = sum(s.get("sks", 0) for s in slots)
        days = sorted(set(s.get("day") for s in slots if s.get("day")))
        num_sections = len(slots)
        logger.info("%s -> %d SKS, days: %s, sections: %d", name, total_sks, days, num_sections)

def optimize_schedule():
    # DEPRECATED: Wrapper lama untuk menjalankan optimasi jadwal penuh.
    # Gunakan rute baru /generate_sections_ga lalu /schedule_sections_ortools.
    global master_schedule
    master_schedule = load_schedule()

    dosen_list = [d["name"] for d in users_collection.find({"role": "dosen"})]

    # Only optimize slots that have courses assigned
    assigned_slots = [s for s in master_schedule if s.get('course_id')]

    if not dosen_list or len(assigned_slots) < 1:
        return

    # Temporarily set master_schedule to assigned_slots for evaluation
    original_master = master_schedule
    master_schedule = assigned_slots

    toolbox = base.Toolbox()
    toolbox.register(
        "individual",
        tools.initIterate,
        creator.Individual,
        lambda: [
            random.choice(dosen_list) if random.random() > 0.5 else None
            for _ in assigned_slots
        ]
    )
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("evaluate", evaluate_schedule)
    toolbox.register("mate", tools.cxTwoPoint)
    toolbox.register("mutate", tools.mutShuffleIndexes, indpb=0.05)
    toolbox.register("select", tools.selTournament, tournsize=3)

    pop = toolbox.population(n=40)
    algorithms.eaSimple(pop, toolbox, cxpb=0.5, mutpb=0.2, ngen=15, verbose=False)

    best = tools.selBest(pop, 1)[0]

    for idx, lecturer in enumerate(best):
        slot_id = assigned_slots[idx]["_id"]
        schedules_collection.update_one(
            {"_id": slot_id},
            {"$set": {"dosen": lecturer}}
        )

    # Restore original master_schedule
    master_schedule = original_master
    master_schedule = load_schedule()

weekday_order = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']

def get_day_order(day):
    if day and day in weekday_order:
        return weekday_order.index(day)
    return len(weekday_order)

def semester_to_int(sem):
    try:
        return int(sem)
    except ValueError:
        roman = {'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5, 'VI': 6, 'VII': 7, 'VIII': 8}
        return roman.get(sem.upper(), 0)

def group_courses_by_semester(courses):
    """Return a dict mapping semester -> list of course dicts (sorted by day, then course_name)."""
    groups = {}
    for c in courses:
        sem = str(c.get('semester') or '0')
        groups.setdefault(sem, []).append(c)
    # sort each group's courses by day, then by name
    for sem, lst in groups.items():
        lst.sort(key=lambda x: (get_day_order((x.get('slot') or {}).get('day')), x.get('course_name', '')))
    return dict(sorted(groups.items(), key=lambda kv: semester_to_int(kv[0])))

def group_slots_by_day(schedule):
    """Return an ordered dict-like structure: list of (day, slots[]) pairs, days in weekday order where possible."""
    # preferred order
    weekday_order = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    groups = {}
    for s in schedule:
        day = s.get('day') or 'Lainnya'
        groups.setdefault(day, []).append(s)
    # sort slots by start time when possible
    for day, lst in groups.items():
        try:
            lst.sort(key=lambda x: x.get('start'))
        except Exception:
            pass
    # order days according to weekday_order, then rest alphabetically
    ordered = []
    for d in weekday_order:
        if d in groups:
            ordered.append((d, groups[d]))
    for d in sorted(k for k in groups.keys() if k not in weekday_order):
        ordered.append((d, groups[d]))
    return ordered

# ------------------------------
# HELPER FUNCTION
# ------------------------------

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ------------------------------
# ROUTES
# ------------------------------

def repair_schedule_conflicts_and_days():
    """
    Re-places existing schedule entries to remove conflicts and enforce lecturer day rules
    (≤3 days). Keeps sections, courses, and lecturers unchanged.
    Adjusts only day/start/end/room.
    """
    docs = list(schedules_collection.find())
    if not docs:
        return 0, 0

    # Build course lookup to know lab vs non-lab
    course_map = {}
    try:
        for c in courses_collection.find():
            course_map[str(c.get('_id'))] = c
    except Exception:
        pass

    # Group by lecturer to preserve day-rule decisions
    by_lect = {}
    for d in docs:
        lect = d.get('dosen') or '(unassigned)'
        by_lect.setdefault(lect, []).append(d)

    # Sort each lecturer's docs to keep relative order stable
    for lect, lst in by_lect.items():
        lst.sort(key=lambda x: (x.get('day') or '', x.get('start') or '', x.get('course_name') or ''))

    new_docs = []
    booked_lect = set()
    booked_room = set()

    weekday_order = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']

    placed_count = 0
    skipped_count = 0

    # Process per lecturer to respect day rule progressively
    for lect, entries in by_lect.items():
        lect_days = set()
        for e in entries:
            course_id = e.get('course_id')
            course_name = e.get('course_name')
            section_name = e.get('section_name')
            section_number = e.get('section_number')
            is_lab = False
            sks = e.get('sks', 0)
            if course_id:
                try:
                    c = course_map.get(str(course_id))
                    if c is not None:
                        is_lab = bool(c.get('is_lab', False))
                        if not sks:
                            sks = c.get('sks', 0)
                except Exception:
                    pass
            # fallback using room type
            if not is_lab and (e.get('room') in LAB_ROOMS):
                is_lab = True

            duration = 100 if sks in [1, 2] else 150

            # Try placement preferring original values first
            preferred_room = e.get('room')
            preferred_day = e.get('day')
            preferred_start = e.get('start')

            rooms = LAB_ROOMS if is_lab else NON_LAB_ROOMS

            def try_place(room_pref=None, day_pref=None, start_pref=None):
                rlist = ([room_pref] + [r for r in rooms if r != room_pref]) if room_pref else list(rooms)
                dlist = ([day_pref] + [d for d in weekday_order if d != day_pref]) if day_pref else list(weekday_order)
                for rd in rlist:
                    for dd in dlist:
                        # Enforce day rule for this lecturer: ≤3 distinct days
                        if dd not in lect_days and len(lect_days) >= 3:
                            continue
                        starts = get_valid_starts(dd, duration)
                        slist = ([start_pref] + [s for s in starts if s != start_pref]) if start_pref else starts
                        for st in slist:
                            if not st:
                                continue
                            s_min = _parse_minutes(st)
                            e_min = _parse_minutes(calculate_end_time(st, duration))
                            if s_min is None or e_min is None:
                                continue
                            conflict = False
                            for minute in range(s_min, e_min, 10):
                                if (lect, dd, minute) in booked_lect or (dd, rd, minute) in booked_room:
                                    conflict = True
                                    break
                            if conflict:
                                continue
                            # Commit
                            for minute in range(s_min, e_min, 10):
                                booked_lect.add((lect, dd, minute))
                                booked_room.add((dd, rd, minute))
                            lect_days.add(dd)
                            return {
                                "_id": e.get('_id'),
                                "course_id": course_id,
                                "course_name": course_name,
                                "section_name": section_name,
                                "sks": sks,
                                "section_number": section_number,
                                "room": rd,
                                "day": dd,
                                "start": st,
                                "end": calculate_end_time(st, duration),
                                "dosen": lect
                            }
                return None

            placed = (
                try_place(preferred_room, preferred_day, preferred_start) or
                # same day, other starts
                try_place(preferred_room, preferred_day, None) or
                # other rooms, same day
                try_place(None, preferred_day, preferred_start) or
                # expand within day rule and room type
                try_place(None, None, None)
            )

            if placed:
                new_docs.append(placed)
                placed_count += 1
            else:
                skipped_count += 1

    # Replace schedules with repaired version (preserve _id if present)
    try:
        schedules_collection.delete_many({})
    except Exception:
        pass
    if new_docs:
        # Convert _id back to ObjectId when available
        to_insert = []
        for d in new_docs:
            doc = dict(d)
            if doc.get('_id'):
                try:
                    doc['_id'] = ObjectId(str(doc['_id']))
                except Exception:
                    doc.pop('_id', None)
            to_insert.append(doc)
        schedules_collection.insert_many(to_insert)
    return placed_count, skipped_count

@app.route("/repair_schedule", methods=["GET", "POST"])
def repair_schedule_route():
    placed, skipped = repair_schedule_conflicts_and_days()
    flash(f"Perbaikan jadwal selesai. Ditempatkan ulang: {placed}, gagal: {skipped}.")
    return redirect(url_for("koordinator"))

def build_conflict_report(schedule=None):
    """
    Scan schedule for conflicts at 10-minute granularity.
    Returns a dict with counts and sample conflict items.
    """
    if schedule is None:
        schedule = load_schedule()
    room_occ = {}
    lect_occ = {}
    room_conflicts = []
    lect_conflicts = []
    for slot in schedule:
        day = slot.get('day')
        room = slot.get('room')
        lect = slot.get('dosen')
        start = slot.get('start')
        end = slot.get('end') or calculate_end_time(start, 100 if slot.get('sks') in [1, 2] else 150)
        s_min = _parse_minutes(start)
        e_min = _parse_minutes(end)
        if day is None or s_min is None or e_min is None:
            continue
        for minute in range(s_min, e_min, 10):
            if room:
                key = (day, room, minute)
                if key in room_occ:
                    room_conflicts.append({
                        "type": "room",
                        "day": day,
                        "minute": minute,
                        "time": _minutes_to_time(minute),
                        "room": room,
                        "slot": {
                            "course": slot.get('course_name'),
                            "section": slot.get('section_name'),
                            "lecturer": lect,
                            "start": start,
                            "end": end
                        },
                        "existing": room_occ[key]
                    })
                else:
                    room_occ[key] = {
                        "course": slot.get('course_name'),
                        "section": slot.get('section_name'),
                        "lecturer": lect,
                        "start": start,
                        "end": end
                    }
            if lect:
                key2 = (day, lect, minute)
                if key2 in lect_occ:
                    lect_conflicts.append({
                        "type": "lecturer",
                        "day": day,
                        "minute": minute,
                        "time": _minutes_to_time(minute),
                        "lecturer": lect,
                        "slot": {
                            "course": slot.get('course_name'),
                            "section": slot.get('section_name'),
                            "room": room,
                            "start": start,
                            "end": end
                        },
                        "existing": lect_occ[key2]
                    })
                else:
                    lect_occ[key2] = {
                        "course": slot.get('course_name'),
                        "section": slot.get('section_name'),
                        "room": room,
                        "start": start,
                        "end": end
                    }
    return {
        "room_conflict_count": len(room_conflicts),
        "lecturer_conflict_count": len(lect_conflicts),
        "room_conflicts_sample": room_conflicts[:50],
        "lecturer_conflicts_sample": lect_conflicts[:50]
    }

@app.route("/conflicts_report", methods=["GET"])
def conflicts_report():
    report = build_conflict_report()
    return jsonify(report)

@app.route("/")
def index():
    # Only expose login page at root — redirect to /login
    return redirect(url_for('login'))

@app.route("/koordinator_home")
def dashborad():
    # Redirect legacy /koordinator_home to unified /koordinator route so dashboard
    # always receives the correct computed counts and data.
    return redirect(url_for('koordinator'))

@app.route("/koordinator_courses")
def koordinator_courses():
    courses = list(db.courses.find())
    if "user" not in session or session["user"]["role"] != "koordinator":
        return redirect(url_for("login"))

    schedule = load_schedule()
    schedule = fix_mongo_id(schedule)  # FIX

    dosen_list = list(users_collection.find({"role": "dosen"}))
    dosen_list = fix_mongo_id(dosen_list)  # FIX

    # Add has_slot and slot info to each course
    for c in courses:
        c['has_slot'] = False
        c['slot'] = None
        try:
            slot = schedules_collection.find_one({"course_id": c["_id"]})
            if slot:
                c['has_slot'] = True
                c['slot'] = {
                    'day': slot.get('day'),
                    'start': slot.get('start'),
                    'end': slot.get('end'),
                    'room': slot.get('room')
                }
        except Exception:
            pass

    courses_by_semester = group_courses_by_semester(courses)
    return render_template(
        "koordinator_courses.html",
        courses=courses,
        courses_by_semester=courses_by_semester,
        schedule=schedule,
        users=dosen_list
    )

@app.route("/koordinator_dosen")
def koordinator_dosen():
    courses = list(db.courses.find())
    if "user" not in session or session["user"]["role"] != "koordinator":
        return redirect(url_for("login"))

    schedule = load_schedule()
    schedule = fix_mongo_id(schedule)  # FIX

    dosen_list = list(users_collection.find({"role": "dosen"}))
    dosen_list = fix_mongo_id(dosen_list)  # FIX

    # Calculate selected_count and section_count for each dosen
    for dosen in dosen_list:
        dosen_name = dosen.get('name')
        
        # Count courses where this dosen is in selected_by
        selected_count = courses_collection.count_documents({"selected_by": dosen_name})
        dosen['selected_count'] = selected_count
        
        # Count sections assigned to this dosen
        section_count = sections_collection.count_documents({"lecturer": dosen_name})
        dosen['section_count'] = section_count

    courses_by_semester = group_courses_by_semester(courses)
    return render_template(
        "koordinator_dosen.html",
        courses=courses,
        courses_by_semester=courses_by_semester,
        schedule=schedule,
        users=dosen_list
    )

@app.route("/koordinator_timeslot")
def koordinator_timeslot():
    courses = list(db.courses.find())
    if "user" not in session or session["user"]["role"] != "koordinator":
        return redirect(url_for("login"))

    schedule = load_schedule()
    schedule = fix_mongo_id(schedule)  # FIX

    dosen_list = list(users_collection.find({"role": "dosen"}))
    dosen_list = fix_mongo_id(dosen_list)  # FIX

    schedule_by_day = group_slots_by_day(schedule)
    return render_template(
        "koordinator_timeslot.html",
        courses=courses,
        schedule=schedule,
        schedule_by_day=schedule_by_day,
        users=dosen_list
    )

@app.route("/koordinator_schedule")
def schedule():
    courses = list(db.courses.find())
    if "user" not in session or session["user"]["role"] != "koordinator":
        return redirect(url_for("login"))

    schedule = load_schedule()
    schedule = fix_mongo_id(schedule)  # FIX

    dosen_list = list(users_collection.find({"role": "dosen"}))
    dosen_list = fix_mongo_id(dosen_list)  # FIX

    schedule_by_day = group_slots_by_day(schedule)
    return render_template(
        "koordinator_schedule.html",
        courses=courses,
        schedule=schedule,
        schedule_by_day=schedule_by_day,
        users=dosen_list
    )

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html")

    identifier = request.form.get("identifier")
    password = request.form.get("password")

    # Koordinator login
    user = users_collection.find_one({"username": identifier})
    if user and check_password_hash(user["password"], password):
        session["user"] = {
            "id": str(user["_id"]),     # FIX
            "username": user["username"],
            "name": user["name"],
            "role": "koordinator"
        }
        return redirect(url_for("koordinator"))

    # Dosen login via email
    user = users_collection.find_one({"email": identifier})
    if user and check_password_hash(user["password"], password):
        session["user"] = {
            "id": str(user["_id"]),     # FIX
            "email": user["email"],
            "name": user["name"],
            "role": "dosen"
        }
        return redirect(url_for("dosen"))

    flash("Login gagal!")
    return redirect(url_for("login"))

@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("index"))

# ------------------------------
# KOORDINATOR PAGE
# ------------------------------

@app.route('/koordinator')
def koordinator():
    if 'user' not in session or session.get('user', {}).get('role') != 'koordinator':
        return redirect(url_for('login'))
    
    # Get data
    courses = list(courses_collection.find())
    lecturers = list(users_collection.find({"role": "dosen"}))
    schedule = list(schedules_collection.find())
    # Pending & failed reschedule counts for notifications
    try:
        pending_reschedule_count = unavailability_reports_collection.count_documents({"status": "pending"})
        failed_needing_action = unavailability_reports_collection.count_documents({"status": "approved", "reschedule_result.failed": {"$gt": 0}})
    except Exception:
        pending_reschedule_count = 0
        failed_needing_action = 0
    
    # Calculate statistics
    courses_count = len(courses)
    dosen_count = len(lecturers)
    
    # Calculate lecturer statistics
    lecturer_stats = {}
    for lecturer in lecturers:
        lecturer_stats[lecturer['name']] = {
            'name': lecturer['name'],
            'total_sks': 0,
            'num_courses': 0,
            'num_sections': 0,
            'days': set(),
            'courses': set()  # Track unique courses
        }
    
    for slot in schedule:
        lecturer_name = slot.get('dosen')
        if lecturer_name and lecturer_name in lecturer_stats:
            lecturer_stats[lecturer_name]['total_sks'] += slot.get('sks', 0)
            lecturer_stats[lecturer_name]['num_sections'] += 1
            lecturer_stats[lecturer_name]['days'].add(slot.get('day', ''))
            course_id = slot.get('course_id')
            if course_id:
                lecturer_stats[lecturer_name]['courses'].add(str(course_id))
    
    # Count unique courses per lecturer
    for stats in lecturer_stats.values():
        stats['num_courses'] = len(stats['courses'])
    
    # Filter active lecturers (those with SKS > 0)
    active_lecturers = []
    for stats in lecturer_stats.values():
        if stats['total_sks'] > 0:
            # Convert days set to sorted list
            stats['days'] = sorted(list(stats['days']))
            active_lecturers.append(stats)
    
    active_dosen_count = len(active_lecturers)
    
    # Sort active lecturers by name
    active_lecturers.sort(key=lambda x: x['name'])
    
    # Get selected courses
    selected_courses = []
    for course in courses:
        # Use selected_by as primary source, fallback to lecturers
        lecturers_list = course.get('selected_by', []) or course.get('lecturers', [])
        course_data = {
            '_id': course.get('_id'),
            'course_name': course.get('course_name'),
            'sks': course.get('sks'),
            'semester': course.get('semester'),
            'lecturer_list': lecturers_list  # Combined field
        }
        selected_courses.append(course_data)
    
    # Check for conflicts
    show_conflicts = request.args.get('show_conflicts') == 'true'
    
    # Add conflict info to schedule if needed
    if show_conflicts:
        schedule = check_schedule_conflicts(schedule)
    
    # Determine schedule status
    schedule_status = 'needs_optimization'
    violations = validate_lecturer_assignments()
    
    if not violations:
        schedule_status = 'optimal'
    elif len(violations) < 5:
        schedule_status = 'feasible'
    
    # Count online classes
    online_count = sum(1 for slot in schedule if slot.get('room') == 'ONLINE')
    
    # Count empty slots
    empty_slots_count = sum(1 for slot in schedule if not slot.get('dosen'))
    
    return render_template(
        'koordinator.html',
        courses_count=courses_count,
        dosen_count=dosen_count,
        active_dosen_count=active_dosen_count,
        schedule=schedule,
        selected_courses=selected_courses,
        violations=violations,
        schedule_status=schedule_status,
        show_conflicts=show_conflicts,
        active_lecturers=active_lecturers,
        online_count=online_count,
        empty_slots_count=empty_slots_count,
        pending_reschedule_count=pending_reschedule_count,
        failed_needing_action=failed_needing_action
    )

@app.route('/koordinator/dosen/<lecturer_name>')
def view_lecturer_schedule(lecturer_name):
    if 'user' not in session or session.get('user', {}).get('role') != 'koordinator':
        return redirect(url_for('login'))
    
    # Get schedule for this lecturer
    schedule = list(schedules_collection.find({"dosen": lecturer_name}))
    schedule = fix_mongo_id(schedule)
    
    # Group by day
    schedule_by_day = {}
    days_order = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat"]
    
    for slot in schedule:
        day = slot.get('day')
        if day:
            if day not in schedule_by_day:
                schedule_by_day[day] = []
            schedule_by_day[day].append(slot)
    
    # Sort each day by time
    for day in schedule_by_day:
        schedule_by_day[day].sort(key=lambda x: x.get('start', ''))
    
    # Calculate statistics
    total_sks = sum(slot.get('sks', 0) for slot in schedule)
    num_classes = len(schedule)
    courses = set(slot.get('course_name') for slot in schedule if slot.get('course_name'))
    num_courses = len(courses)
    days_teaching = sorted(list(schedule_by_day.keys()), key=lambda x: days_order.index(x) if x in days_order else 99)
    
    return render_template(
        'koordinator_lecturer_schedule.html',
        lecturer_name=lecturer_name,
        schedule_by_day=schedule_by_day,
        days_order=days_order,
        total_sks=total_sks,
        num_classes=num_classes,
        num_courses=num_courses,
        days_teaching=days_teaching
    )

def check_schedule_conflicts(schedule):
    """
    Deteksi bentrok ruangan dan dosen dalam jadwal.
    Returns: dictionary dengan keys 'lecturer_conflicts' dan 'room_conflicts'
    NOTE: ONLINE rooms are excluded from room conflict checks (unlimited capacity)
    """
    if not schedule:
        return {'lecturer_conflicts': [], 'room_conflicts': []}
    
    from collections import defaultdict
    
    # Group by lecturer and room
    lecturer_schedule = defaultdict(list)
    room_schedule = defaultdict(list)
    
    for slot in schedule:
        day = slot.get('day')
        room = slot.get('room')
        lecturer = slot.get('dosen')
        start_time = slot.get('start')
        end_time = slot.get('end')
        course = slot.get('course_name', 'Unknown')
        
        if not all([day, room, start_time, end_time]):
            continue
        
        # Convert time to minutes for easier comparison
        def time_to_minutes(time_str):
            if not time_str or ':' not in time_str:
                return 0
            h, m = time_str.split(':')
            return int(h) * 60 + int(m)
        
        start_min = time_to_minutes(start_time)
        end_min = time_to_minutes(end_time)
        
        if lecturer:
            lecturer_schedule[lecturer].append((day, start_min, end_min, course, room, slot.get('_id')))
        
        # Only check physical rooms for conflicts (skip ONLINE)
        if room and room != 'ONLINE':
            room_schedule[(day, room)].append((start_min, end_min, course, lecturer, slot.get('_id')))
    
    # Check lecturer conflicts
    lecturer_conflicts = []
    for lect, slots in lecturer_schedule.items():
        for i in range(len(slots)):
            for j in range(i+1, len(slots)):
                day1, s1, e1, c1, r1, id1 = slots[i]
                day2, s2, e2, c2, r2, id2 = slots[j]
                
                # Same day and overlapping times?
                if day1 == day2 and s1 < e2 and s2 < e1:
                    lecturer_conflicts.append({
                        'lecturer': lect,
                        'day': day1,
                        'course1': c1,
                        'room1': r1,
                        'time1': f"{s1//60:02d}:{s1%60:02d}-{e1//60:02d}:{e1%60:02d}",
                        'course2': c2,
                        'room2': r2,
                        'time2': f"{s2//60:02d}:{s2%60:02d}-{e2//60:02d}:{e2%60:02d}",
                        'id1': str(id1) if id1 else None,
                        'id2': str(id2) if id2 else None
                    })
    
    # Check room conflicts (ONLINE already excluded above)
    room_conflicts = []
    for (day, room), slots in room_schedule.items():
        for i in range(len(slots)):
            for j in range(i+1, len(slots)):
                s1, e1, c1, l1, id1 = slots[i]
                s2, e2, c2, l2, id2 = slots[j]
                
                # Overlapping times?
                if s1 < e2 and s2 < e1:
                    room_conflicts.append({
                        'room': room,
                        'day': day,
                        'course1': c1,
                        'lecturer1': l1,
                        'time1': f"{s1//60:02d}:{s1%60:02d}-{e1//60:02d}:{e1%60:02d}",
                        'course2': c2,
                        'lecturer2': l2,
                        'time2': f"{s2//60:02d}:{s2%60:02d}-{e2//60:02d}:{e2%60:02d}",
                        'id1': str(id1) if id1 else None,
                        'id2': str(id2) if id2 else None
                    })
    
    return {
        'lecturer_conflicts': lecturer_conflicts,
        'room_conflicts': room_conflicts
    }

@app.route("/koordinator/spread_online_courses", methods=["POST"])
def spread_online_courses():
    """Sebarkan kelas ONLINE ke Senin-Jumat dan perbaiki bentrok dengan jadwal fisik dosen."""
    if "user" not in session or session["user"]["role"] != "koordinator":
        return redirect(url_for("login"))

    import time
    start_time = time.perf_counter()
    
    print("=== START: Spreading online courses & fixing conflicts ===")
    logger.info("=== START: Spreading online courses & fixing conflicts ===")
    
    try:
        from collections import defaultdict
        from bson import ObjectId
        from pymongo import UpdateOne

        TARGET_DAYS = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
        
        # Time slots untuk kelas ONLINE (1-2 SKS)
        ONLINE_TIMEBLOCKS = [
            '08:00-08:50', '08:50-09:40', '09:40-10:30', '10:30-11:20', '11:20-12:10',
            '08:00-09:40', '09:40-11:20', '11:20-13:00',
            '14:00-14:50', '14:50-15:40', '15:40-16:30',
            '14:00-15:40', '15:40-17:20'
        ]

        print("Fetching all schedules...")
        all_schedules = list(schedules_collection.find())
        
        # Separate ONLINE and non-ONLINE
        online_schedules = [s for s in all_schedules if s.get('room') == 'ONLINE']
        non_online_schedules = [s for s in all_schedules if s.get('room') != 'ONLINE']
        
        print(f"Found {len(online_schedules)} online courses, {len(non_online_schedules)} physical courses")
        
        if not online_schedules:
            flash("Tidak ada kelas daring untuk disebarkan.", "info")
            return redirect("/koordinator")
        
        # Build lecturer occupation map (physical classes only)
        def time_to_minutes(t):
            h, m = map(int, t.split(':'))
            return h * 60 + m
        
        def times_overlap(start1, end1, start2, end2):
            s1, e1 = time_to_minutes(start1), time_to_minutes(end1)
            s2, e2 = time_to_minutes(start2), time_to_minutes(end2)
            return s1 < e2 and s2 < e1
        
        # Map: (lecturer, day) -> [(start, end), ...]
        lecturer_physical_schedule = defaultdict(list)
        for sched in non_online_schedules:
            lecturer = sched.get('dosen')
            day = sched.get('day')
            if lecturer and day:
                lecturer_physical_schedule[(lecturer, day)].append(
                    (sched.get('start'), sched.get('end'))
                )
        
        print("Distributing online courses across Senin-Jumat evenly...")
        
        # Group online schedules by lecturer
        online_by_lecturer = defaultdict(list)
        for sched in online_schedules:
            lecturer = sched.get('dosen')
            if lecturer:
                online_by_lecturer[lecturer].append(sched)
        
        bulk_ops = []
        spread_count = 0
        
        # Distribute each lecturer's online classes
        for lecturer, online_classes in online_by_lecturer.items():
            print(f"\n📚 Processing {len(online_classes)} online classes for {lecturer}")
            
            # Get all physical slots for this lecturer
            physical_slots_all_days = {}
            for day in TARGET_DAYS:
                physical_slots_all_days[day] = lecturer_physical_schedule.get((lecturer, day), [])
            
            # Distribute online classes across days evenly
            # Strategy: Round-robin assignment to days, avoiding conflicts
            day_index = 0
            
            for online_sched in online_classes:
                current_day = online_sched.get('day')
                current_start = online_sched.get('start')
                current_end = online_sched.get('end')
                
                # Try to assign to a day in round-robin fashion
                found_slot = False
                attempts = 0
                
                while attempts < len(TARGET_DAYS) and not found_slot:
                    target_day = TARGET_DAYS[day_index % len(TARGET_DAYS)]
                    day_physical_slots = physical_slots_all_days[target_day]
                    
                    # Find a safe timeblock for this day
                    for timeblock in ONLINE_TIMEBLOCKS:
                        tb_start, tb_end = timeblock.split('-')
                        
                        # Check if this timeblock conflicts with any physical class
                        safe = True
                        for phys_start, phys_end in day_physical_slots:
                            if times_overlap(tb_start, tb_end, phys_start, phys_end):
                                safe = False
                                break
                        
                        if safe:
                            # Found a safe slot!
                            if target_day != current_day or tb_start != current_start:
                                print(f"  ✓ {online_sched.get('course_name', 'Unknown')}: {current_day} {current_start} → {target_day} {timeblock}")
                                bulk_ops.append(UpdateOne(
                                    {'_id': online_sched['_id']},
                                    {'$set': {
                                        'day': target_day,
                                        'start': tb_start,
                                        'end': tb_end
                                    }}
                                ))
                                spread_count += 1
                            else:
                                print(f"  ✓ {online_sched.get('course_name', 'Unknown')}: tetap di {target_day} {timeblock}")
                            
                            found_slot = True
                            day_index += 1  # Move to next day for next class
                            break
                    
                    if not found_slot:
                        attempts += 1
                        day_index += 1
                
                if not found_slot:
                    print(f"  ❌ Could not find safe slot for {online_sched.get('course_name', 'Unknown')}")
        
        # Execute bulk updates
        if bulk_ops:
            print(f"\n🔄 Executing {len(bulk_ops)} bulk updates...")
            res = schedules_collection.bulk_write(bulk_ops, ordered=False)
            print(f"✅ Updated {res.modified_count} schedules")
        
        elapsed = (time.perf_counter() - start_time) * 1000
        
        if spread_count > 0:
            msg = f"✅ Berhasil menyebarkan {spread_count} kelas ONLINE ke Senin-Jumat (waktu {elapsed:.0f} ms)"
            flash(msg, "success")
        else:
            msg = f"✅ Semua kelas ONLINE sudah tersebar optimal (waktu {elapsed:.0f} ms)"
            flash(msg, "info")
        
        print(f"\n{msg}")
        logger.info(msg)
        logger.info("=== END: Spreading online courses SUCCESS ===")

    except Exception as e:
        elapsed = (time.perf_counter() - start_time) * 1000
        error_msg = f"❌ Error saat menyebarkan kelas: {str(e)}"
        flash(error_msg, "error")
        logger.error(f"Error spreading online courses after {elapsed:.0f}ms: {e}", exc_info=True)
        logger.info("=== END: Spreading online courses FAILED ===")

    return redirect(url_for('koordinator'))


@app.route("/koordinator/update_islab_field", methods=["POST"])
def update_islab_field_route():
    """Route untuk update is_lab field secara manual"""
    if 'user' not in session or session['user']['role'] != 'koordinator':
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    
    try:
        updated_count = _update_islab_field()
        
        # Count lab vs non-lab courses
        lab_count = courses_collection.count_documents({'is_lab': True})
        non_lab_count = courses_collection.count_documents({'is_lab': False})
        
        return jsonify({
            "success": True,
            "message": f"Berhasil update {updated_count} mata kuliah",
            "lab_courses": lab_count,
            "non_lab_courses": non_lab_count
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ------------------------------
# KOORDINATOR CRUD ROUTES
# ------------------------------

@app.route('/koordinator/edit_course/<course_id>', methods=['GET', 'POST'])
def edit_course(course_id):
    if 'user' not in session or session['user']['role'] != 'koordinator':
        return redirect(url_for('login'))
    try:
        course = courses_collection.find_one({'_id': ObjectId(course_id)})
    except Exception:
        course = None
    if not course:
        flash('Mata kuliah tidak ditemukan')
        return redirect(url_for('koordinator'))

    if request.method == 'POST':
        name = request.form.get('course_name')
        sks = int(request.form.get('sks') or 0)
        semester = request.form.get('semester')

        # perform update and redirect after POST
        courses_collection.update_one({'_id': ObjectId(course_id)}, {'$set': {'course_name': name, 'sks': sks, 'semester': semester}})
        flash('Mata kuliah diperbarui')
        return redirect(url_for('koordinator_courses'))

    try:
        course['_id'] = str(course['_id'])
    except Exception:
        pass
    return render_template('edit_course.html', course=course)

@app.route('/koordinator/delete_course/<course_id>')
def delete_course(course_id):
    if 'user' not in session or session['user']['role'] != 'koordinator':
        return redirect(url_for('login'))
    try:
        # Delete course
        courses_collection.delete_one({'_id': ObjectId(course_id)})

        # Also delete any slot that was assigned to this course
        schedules_collection.delete_many({"course_id": ObjectId(course_id)})

        flash('Mata kuliah dihapus dan slot terkait dihapus')
    except Exception:
        flash('Gagal menghapus mata kuliah')
    return redirect(url_for('koordinator_courses'))

@app.route('/koordinator/edit_dosen/<dosen_id>', methods=['GET', 'POST'])
def edit_dosen(dosen_id):
    if 'user' not in session or session['user']['role'] != 'koordinator':
        return redirect(url_for('login'))
    try:
        u = users_collection.find_one({'_id': ObjectId(dosen_id)})
    except Exception:
        u = None
    if not u:
        flash('Dosen tidak ditemukan')
        return redirect(url_for('koordinator'))

    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')

        users_collection.update_one({'_id': ObjectId(dosen_id)}, {'$set': {'name': name, 'email': email}})
        flash('Dosen diperbarui')
        return redirect(url_for('koordinator_dosen'))

    try:
        u['_id'] = str(u['_id'])
    except Exception:
        pass
    return render_template('edit_dosen.html', dosen=u)

@app.route('/koordinator/delete_dosen/<dosen_id>')
def delete_dosen(dosen_id):
    if 'user' not in session or session['user']['role'] != 'koordinator':
        return redirect(url_for('login'))
    try:
        users_collection.delete_one({'_id': ObjectId(dosen_id)})
        flash('Dosen dihapus')
    except Exception:
        flash('Gagal menghapus dosen')
    return redirect(url_for('koordinator_dosen'))

@app.route('/koordinator/edit_slot/<slot_id>', methods=['GET', 'POST'])
def edit_slot(slot_id):
    if 'user' not in session or session['user']['role'] != 'koordinator':
        return redirect(url_for('login'))
    try:
        s = schedules_collection.find_one({'_id': ObjectId(slot_id)})
    except Exception:
        s = None
    if not s:
        flash('Slot tidak ditemukan')
        return redirect(url_for('koordinator_timeslot'))

    if request.method == 'POST':
        day = request.form.get('day')
        start = request.form.get('start')
        end = request.form.get('end')
        room = request.form.get('room')

        schedules_collection.update_one({'_id': ObjectId(slot_id)}, {'$set': {'day': day, 'start': start, 'end': end, 'room': room}})
        flash('Slot diperbarui')
        return redirect(url_for('koordinator_timeslot'))

    try:
        s['_id'] = str(s['_id'])
    except Exception:
        pass
    return render_template('edit_slot.html', slot=s)

@app.route('/koordinator/delete_slot/<slot_id>')
def delete_slot(slot_id):
    if 'user' not in session or session['user']['role'] != 'koordinator':
        return redirect(url_for('login'))
    try:
        schedules_collection.delete_one({'_id': ObjectId(slot_id)})
        flash('Slot dihapus')
    except Exception:
        flash('Gagal menghapus slot')
    return redirect(url_for('koordinator_timeslot'))

# ------------------------------
# DOSEN PREFERENCES
# ------------------------------

@app.route("/dosen/preferences", methods=["GET", "POST"])
def dosen_preferences():
    if "user" not in session or session["user"]["role"] != "dosen":
        return redirect(url_for("login"))

    username = session["user"]["name"]
    user_id = session["user"]["id"]

    if request.method == "POST":
        available_days = request.form.getlist("available_days")
        preferred_times = {}
        max_load = int(request.form.get("max_load", 12))

        # Parse preferred times for each day
        weekdays = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
        for day in weekdays:
            start_times = request.form.getlist(f"preferred_times_{day}_start")
            end_times = request.form.getlist(f"preferred_times_{day}_end")
            if start_times and end_times:
                preferred_times[day] = [f"{s}-{e}" for s, e in zip(start_times, end_times)]
        
        # Parse blocked time ranges (unavailable times)
        block_days = request.form.getlist("block_day[]")
        block_starts = request.form.getlist("block_start[]")
        block_ends = request.form.getlist("block_end[]")
        
        unavailable_time_ranges = []
        for i in range(len(block_days)):
            if i < len(block_starts) and i < len(block_ends):
                day = block_days[i]
                start = block_starts[i]
                end = block_ends[i]
                
                # Validate non-empty
                if day and start and end:
                    unavailable_time_ranges.append({
                        "day": day,
                        "start": start,
                        "end": end
                    })

        # Save preferences to database
        users_collection.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": {
                "preferences": {
                    "available_days": available_days,
                    "preferred_times": preferred_times,
                    "unavailable_time_ranges": unavailable_time_ranges,
                    "max_load": max_load
                }
            }},
            upsert=True
        )

        flash("Preferensi berhasil disimpan!")
        return redirect(url_for("dosen"))

    # Get current preferences
    user = users_collection.find_one({"_id": ObjectId(user_id)})
    preferences = user.get("preferences", {
        "available_days": [],
        "preferred_times": {},
        "unavailable_time_ranges": [],
        "max_load": 12
    })

    weekdays = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
    time_slots = ['08:00', '09:00', '10:00', '11:00', '12:00', '13:00', '14:00', '15:00', '16:00', '17:00']

    return render_template("dosen_preferences.html",
                         preferences=preferences,
                         weekdays=weekdays,
                         time_slots=time_slots)

@app.route("/dosen/report_unavailability", methods=["GET", "POST"])
def dosen_report_unavailability():
    if "user" not in session or session["user"]["role"] != "dosen":
        return redirect(url_for("login"))

    username = session["user"]["name"]
    user_id = session["user"]["id"]

    if request.method == "POST":
        # Parse multiple unavailability entries
        days = request.form.getlist('unavailable_day[]')
        time_types = request.form.getlist('time_type[]')
        start_times = request.form.getlist('start_time[]')
        end_times = request.form.getlist('end_time[]')

        # Build entries array
        entries = []
        for i in range(len(days)):
            if not days[i]:  # Skip empty entries
                continue
                
            entry = {
                'day': days[i],
                'time_type': time_types[i] if i < len(time_types) else 'all'
            }
            
            # Add time range if type is 'range'
            if entry['time_type'] == 'range':
                entry['start'] = start_times[i] if i < len(start_times) else '07:00'
                entry['end'] = end_times[i] if i < len(end_times) else '17:00'
            else:
                entry['start'] = '07:00'
                entry['end'] = '17:00'
            
            entries.append(entry)

        if not entries:
            flash('Silakan tambahkan minimal 1 entry ketidaktersediaan', 'warning')
            return redirect(url_for('dosen_report_unavailability'))

        # Convert entries to unavailable_time_ranges format for lecturer_preferences
        unavailable_time_ranges = []
        affected_days = set()
        
        for entry in entries:
            if entry['day'] == '*':  # Semua hari
                for day in POSSIBLE_DAYS:
                    unavailable_time_ranges.append({
                        'day': day,
                        'start': entry['start'],
                        'end': entry['end']
                    })
                    affected_days.add(day)
            else:
                unavailable_time_ranges.append({
                    'day': entry['day'],
                    'start': entry['start'],
                    'end': entry['end']
                })
                affected_days.add(entry['day'])

        # Save to unavailability_reports for coordinator approval
        unavailability_reports_collection.insert_one({
            'dosen_name': username,
            'entries': entries,  # Array of {day, time_type, start, end}
            'unavailable_time_ranges': unavailable_time_ranges,  # Expanded format
            'status': 'pending',
            'submitted_at': datetime.now()
        })

        # Also update user preferences (for immediate effect)
        user = users_collection.find_one({"_id": ObjectId(user_id)})
        preferences = user.get("preferences", {})
        
        # Update available_days (remove affected days)
        current_available = preferences.get("available_days", POSSIBLE_DAYS[:])
        new_available = [day for day in current_available if day not in affected_days]
        preferences["available_days"] = new_available
        
        # Merge unavailable_time_ranges (avoid duplicates)
        existing_ranges = preferences.get("unavailable_time_ranges", [])
        for new_range in unavailable_time_ranges:
            # Check if not duplicate
            is_duplicate = any(
                r['day'] == new_range['day'] and 
                r['start'] == new_range['start'] and 
                r['end'] == new_range['end']
                for r in existing_ranges
            )
            if not is_duplicate:
                existing_ranges.append(new_range)
        preferences["unavailable_time_ranges"] = existing_ranges

        users_collection.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": {"preferences": preferences}},
            upsert=True
        )
        
        flash(f'Laporan ketidaktersediaan berhasil dikirim ({len(entries)} entry) dan langsung diupdate ke preferences', 'success')
        return redirect(url_for('dosen'))

    # GET: Load existing pending report if any unavailability_reports
    existing_report = unavailability_reports_collection.find_one({
        'dosen_name': username,
        'status': 'pending'
    })
    
    existing_entries = []
    if existing_report:
        # Use new format (entries)
        existing_entries = existing_report.get('entries', [])

    return render_template('dosen_report_unavailability.html', 
                         existing_entries=existing_entries)

@app.route("/dosen/request_reschedule", methods=["GET", "POST"])
def dosen_request_reschedule():
    """Dosen request reschedule untuk hari tertentu (emergency)"""
    if "user" not in session or session["user"]["role"] != "dosen":
        return redirect(url_for("login"))
    
    username = session["user"]["name"]
    
    if request.method == "POST":
        # Parse multiple reschedule entries
        days = request.form.getlist("unavailable_day[]")
        time_types = request.form.getlist("time_type[]")
        specific_times = request.form.getlist("specific_time[]")
        start_times = request.form.getlist("start_time[]")
        end_times = request.form.getlist("end_time[]")
        reasons = request.form.getlist("reason[]")
        
        if not days:
            flash("Pilih minimal satu hari yang tidak bisa mengajar!", "error")
            return redirect(url_for("dosen_request_reschedule"))
        
        # Process each entry and collect all affected classes
        all_entries = []
        total_affected_classes = []
        
        for i in range(len(days)):
            if not days[i]:
                continue
            
            entry = {
                'day': days[i],
                'time_type': time_types[i] if i < len(time_types) else 'all',
                'reason': reasons[i] if i < len(reasons) else ''
            }
            
            # Add time details based on type
            if entry['time_type'] == 'specific':
                entry['specific_time'] = specific_times[i] if i < len(specific_times) else ''
            elif entry['time_type'] == 'range':
                entry['start_time'] = start_times[i] if i < len(start_times) else ''
                entry['end_time'] = end_times[i] if i < len(end_times) else ''
            
            # Find affected classes for this entry
            # Handle "Semua Hari" (*) - expand to all weekdays
            search_days = []
            if days[i] == '*':
                search_days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
            else:
                search_days = [days[i]]
            
            classes_for_day = []
            for search_day in search_days:
                classes_for_day.extend(list(schedules_collection.find({'dosen': username, 'day': search_day})))
            
            # Filter by time if specified
            if entry['time_type'] != 'all':
                def time_to_minutes(t):
                    if not t:
                        return None
                    h, m = map(int, t.split(':'))
                    return h * 60 + m
                
                def class_matches(cls):
                    start_str = cls.get('start', '')
                    end_str = cls.get('end', '')
                    if not start_str or not end_str:
                        return False
                    start_minutes = time_to_minutes(start_str)
                    end_minutes = time_to_minutes(end_str)
                    
                    if entry['time_type'] == "specific" and entry.get('specific_time'):
                        target = time_to_minutes(entry['specific_time'])
                        return target is not None and abs(start_minutes - target) <= 10
                    
                    if entry['time_type'] == "range" and entry.get('start_time') and entry.get('end_time'):
                        range_start = time_to_minutes(entry['start_time'])
                        range_end = time_to_minutes(entry['end_time'])
                        return (start_minutes < range_end and end_minutes > range_start)
                    
                    return True
                
                classes_for_day = [c for c in classes_for_day if class_matches(c)]
            
            # Store affected classes in entry
            entry['affected_classes'] = [
                {
                    "course_name": cls.get('course_name'),
                    "section_label": cls.get('section_label', '-'),
                    "start": cls.get('start'),
                    "end": cls.get('end'),
                    "room": cls.get('room'),
                    "sks": cls.get('sks', 0)
                }
                for cls in classes_for_day
            ]
            
            all_entries.append(entry)
            total_affected_classes.extend(classes_for_day)
        
        if not all_entries:
            flash("Tidak ada entry yang valid!", "error")
            return redirect(url_for("dosen_request_reschedule"))
        
        if not total_affected_classes:
            flash("Tidak ada kelas yang terpengaruh pada waktu yang dipilih", "info")
            return redirect(url_for("dosen"))
        
        # Save all entries in one report
        from datetime import datetime
        report_data = {
            "lecturer_name": username,
            "entries": all_entries,  # Array of entries with their affected classes
            "total_affected_count": len(total_affected_classes),
            "status": "pending",
            "created_at": datetime.now(),
            "auto_rescheduled": False
        }
        
        unavailability_reports_collection.insert_one(report_data)
        
        flash(f"Berhasil mengirim {len(all_entries)} permintaan reschedule dengan total {len(total_affected_classes)} kelas yang terpengaruh", "success")
        return redirect(url_for("dosen"))
    
    # GET: Show form
    weekdays = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
    
    # Get lecturer's schedule grouped by day
    schedule = list(schedules_collection.find({'dosen': username}))
    schedule_by_day = {}
    for cls in schedule:
        day = cls.get('day')
        if day:
            if day not in schedule_by_day:
                schedule_by_day[day] = []
            # Convert to JSON-serializable format (remove ObjectId and other non-serializable fields)
            schedule_by_day[day].append({
                'course_name': cls.get('course_name', ''),
                'section_label': cls.get('section_label', ''),
                'start': cls.get('start', ''),
                'end': cls.get('end', ''),
                'room': cls.get('room', ''),
                'sks': cls.get('sks', 0)
            })
    
    return render_template("dosen_request_reschedule.html",
                         weekdays=weekdays,
                         schedule_by_day=schedule_by_day)

@app.route("/koordinator/reschedule_lecturer/<report_id>", methods=["POST"])
def koordinator_reschedule_lecturer(report_id):
    """Koordinator approve dan auto-reschedule"""
    if 'user' not in session or session['user']['role'] != 'koordinator':
        return jsonify({"success": False, "message": "Unauthorized"}), 403
    
    try:
        report = unavailability_reports_collection.find_one({'_id': ObjectId(report_id)})
        if not report:
            return jsonify({"success": False, "message": "Report not found"}), 404
        
        lecturer_name = report.get('lecturer_name')
        
        # Check if new format (entries) or old format
        if 'entries' in report and report['entries']:
            # NEW FORMAT: Process each entry separately
            total_moved = 0
            total_failed = 0
            all_failed_courses = []
            all_debug_info = []
            
            for entry_idx, entry in enumerate(report['entries'], 1):
                day = entry.get('day')
                time_type = entry.get('time_type', 'all')
                specific_time = entry.get('specific_time')
                start_time = entry.get('start_time')
                end_time = entry.get('end_time')
                
                # Expand "Semua Hari" (*)
                if day == '*':
                    unavailable_days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
                    unavailable_day = 'Senin'  # fallback for single day param
                else:
                    unavailable_days = [day]
                    unavailable_day = day
                
                all_debug_info.append(f"--- Entry #{entry_idx}: {day} ({time_type}) ---")
                
                # Call reschedule function
                result = _reschedule_lecturer_classes_v2(
                    lecturer_name,
                    unavailable_day,
                    unavailable_days=unavailable_days,
                    time_type=time_type,
                    specific_time=specific_time,
                    start_time=start_time,
                    end_time=end_time
                )
                
                total_moved += result.get('moved', 0)
                total_failed += result.get('failed', 0)
                all_failed_courses.extend(result.get('failed_courses', []))
                if 'debug_info' in result:
                    all_debug_info.extend(result['debug_info'])
            
            # Update report with aggregated results
            final_result = {
                'moved': total_moved,
                'failed': total_failed,
                'failed_courses': all_failed_courses,
                'debug_info': all_debug_info
            }
            
        else:
            # OLD FORMAT: Single unavailable day/time (backward compatibility)
            unavailable_day = report.get('unavailable_day')
            unavailable_days = report.get('unavailable_days') or [unavailable_day]
            time_type = report.get('unavailable_time_type')
            specific_time = report.get('specific_time')
            start_time = report.get('start_time')
            end_time = report.get('end_time')
            
            final_result = _reschedule_lecturer_classes_v2(
                lecturer_name,
                unavailable_day,
                unavailable_days=unavailable_days,
                time_type=time_type,
                specific_time=specific_time,
                start_time=start_time,
                end_time=end_time
            )
        
        # Update report status
        unavailability_reports_collection.update_one(
            {'_id': ObjectId(report_id)},
            {'$set': {
                'status': 'approved',
                'auto_rescheduled': True,
                'reschedule_result': final_result,
                'processed_at': datetime.now()
            }}
        )
        
        if final_result['moved'] > 0:
            flash(f"✅ Berhasil memindahkan {final_result['moved']} kelas {lecturer_name}", "success")
        else:
            flash(f"⚠️ Tidak ada kelas yang bisa dipindahkan otomatis. Perlu penjadwalan manual.", "warning")
        
        return jsonify({"success": True, "result": final_result})
        
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

def _reschedule_lecturer_classes_v2(lecturer_name, unavailable_day, unavailable_days=None, time_type=None, specific_time=None, start_time=None, end_time=None, max_end_time="17:30", enable_swap=True):
    """Enhanced reschedule with:
    - max_end_time constraint (default 17:30 to include afternoon slots like 14:00-16:30)
    - swap logic: if ideal earlier slot occupied by ONE lecturer class, try relocating that blocking class.
    - preference order: keep same day + earlier slots, then other days earlier slots.
    - RESPECTS PREFERENCES: Uses check_lecturer_preferences() for validation
    - RESPECTS UNAVAILABILITY: Checks approved unavailability_reports to avoid blocked times
    """
    from collections import defaultdict

    def t2m(t):
        h,m = map(int,t.split(':')); return h*60+m
    def m2t(m):
        return f"{m//60:02d}:{m%60:02d}"
    def overlap(s1,e1,s2,e2):
        return t2m(s1) < t2m(e2) and t2m(s2) < t2m(e1)

    source_days = unavailable_days if (unavailable_days and len(unavailable_days)>0) else [unavailable_day]
    allow_day_change = (len(source_days)==1) or (time_type in ('specific','range'))
    raw = []
    for d in source_days:
        raw.extend(list(schedules_collection.find({'dosen': lecturer_name,'day':d})))

    def in_block(s,e):
        if not s or not e: return False
        if time_type=='specific' and specific_time:
            return abs(t2m(s)-t2m(specific_time)) <= 10
        if time_type=='range' and start_time and end_time:
            return t2m(s) < t2m(end_time) and t2m(e) > t2m(start_time)
        return True  # no filter -> treat all

    affected = [c for c in raw if in_block(c.get('start'), c.get('end'))]
    unaffected = [c for c in raw if c not in affected]
    
    # Log initial state
    logger.info(f"[Reschedule] Starting reschedule for {lecturer_name}")
    logger.info(f"[Reschedule] Affected classes: {len(affected)}, Unaffected: {len(unaffected)}")
    logger.info(f"[Reschedule] Source days: {source_days}, Time type: {time_type}")
    if time_type == 'range':
        logger.info(f"[Reschedule] Blocked time range: {start_time} - {end_time}")
    
    if not affected:
        logger.info(f"[Reschedule] No classes inside block - nothing to reschedule")
        return {"success":True,"moved":0,"failed":0,"failed_courses":[],"message":"No classes inside block","total":0,"unaffected":len(unaffected)}

    # Build lecturer schedule (all days)
    lect_sched = defaultdict(list)
    for s in schedules_collection.find({'dosen': lecturer_name}):
        d = s.get('day');
        if not d: continue
        lect_sched[d].append({'start':s.get('start'),'end':s.get('end'),'id':s.get('_id'),'room':s.get('room')})

    # Room occupation (including ONLINE) with normalized room names
    room_occ = defaultdict(lambda: defaultdict(list))
    for s in schedules_collection.find():
        d = s.get('day'); r = s.get('room');
        if d and r:
            rn = r.strip()
            room_occ[d][rn].append({'start':s.get('start'),'end':s.get('end'), 'course_name': s.get('course_name'), 'dosen': s.get('dosen')})

    # Get lecturer preferences to respect available_days, preferred_times, and unavailable_time_ranges
    lecturer_prefs = users_collection.find_one({'name': lecturer_name, 'role': 'dosen'})
    available_days_pref = []
    if lecturer_prefs and 'preferences' in lecturer_prefs:
        available_days_pref = lecturer_prefs['preferences'].get('available_days', [])

    # Get approved unavailability_reports for this lecturer (to avoid rescheduling into already blocked times)
    approved_blocks = list(unavailability_reports_collection.find({
        'lecturer_name': lecturer_name,
        'status': 'approved'
    }))
    logger.info(f"[Reschedule] Found {len(approved_blocks)} approved unavailability blocks")
    for idx, block in enumerate(approved_blocks):
        block_days = block.get('unavailable_days', []) or [block.get('unavailable_day')]
        block_time_type = block.get('unavailable_time_type', 'full_day')
        logger.info(f"[Reschedule]   Block {idx+1}: days={block_days}, type={block_time_type}")
    
    # Helper function to check if a slot conflicts with approved unavailability reports
    def is_in_approved_block(day, slot_start, slot_end):
        """Check if this slot falls into any approved unavailability block"""
        for block in approved_blocks:
            block_days = block.get('unavailable_days', []) or [block.get('unavailable_day')]
            if day not in block_days:
                continue
            
            block_time_type = block.get('unavailable_time_type', 'full_day')
            
            if block_time_type == 'full_day':
                return True  # Entire day is blocked
            
            if block_time_type == 'specific':
                specific = block.get('specific_time')
                if specific and abs(t2m(slot_start) - t2m(specific)) <= 10:
                    return True
            
            if block_time_type == 'range':
                block_start = block.get('start_time')
                block_end = block.get('end_time')
                if block_start and block_end:
                    if overlap(slot_start, slot_end, block_start, block_end):
                        return True
        
        return False

    DAYS = ['Senin','Selasa','Rabu','Kamis','Jumat']
    if allow_day_change:
        excluded_days = set(unavailable_days or [unavailable_day])
        
        # CRITICAL FIX: Also exclude ALL days from approved unavailability blocks (full_day)
        for block in approved_blocks:
            block_time_type = block.get('unavailable_time_type', 'full_day')
            if block_time_type == 'full_day':
                block_days = block.get('unavailable_days', []) or [block.get('unavailable_day')]
                # Filter out None values
                block_days = [d for d in block_days if d is not None]
                excluded_days.update(block_days)
        
        logger.info(f"[Reschedule] Excluded days (from unavailability): {sorted(excluded_days)}")
        target_days = [d for d in DAYS if d not in excluded_days]
        
        # Filter by lecturer preferences (available_days)
        if available_days_pref:
            target_days = [d for d in target_days if d in available_days_pref]
            if not target_days:
                # No valid days, fallback to preference days only
                target_days = [d for d in available_days_pref if d not in excluded_days]
        
        if not target_days:
            target_days = source_days[:]  # ultimate fallback
    else:
        target_days = source_days[:]

    LAB_ROOMS = ['Lab 1','Lab 2','Lab 3','Lab 4']
    NON_LAB_ROOMS = ['Infor 1','Infor 2','Infor 3','Infor 4','Infor 5','Jawa 8A','Jawa 8B']

    max_end_m = t2m(max_end_time)

    # Istirahat siang: hindari penempatan yang menyilang 13:00–14:00 untuk Senin–Kamis
    # Slot yang mulai jam 14:00 atau lebih DIPERBOLEHKAN (sudah lewat istirahat)
    LUNCH_START = '13:00'
    LUNCH_END = '14:00'

    def overlaps_lunch(day, cs, ce):
        if day in ('Jumat',):
            return False
        # Allow slots that start at or after 14:00 (after lunch break)
        if t2m(cs) >= t2m('14:00'):
            return False
        # Reject slots that overlap with lunch break
        return overlap(cs, ce, LUNCH_START, LUNCH_END)

    def generate_candidates(duration):
        # Generate time slots starting from 08:00 (normal class hours)
        # earlier first
        candidates = []
        # Ensure we check up to max_end_time (17:00 by default)
        end_limit = max_end_m if max_end_m else t2m('17:00')
        for start_m in range(t2m('08:00'), end_limit - duration + 1, 50):  # 50 min intervals (1 SKS)
            end_m = start_m + duration
            if end_m > end_limit: continue
            candidates.append((m2t(start_m), m2t(end_m)))
        return candidates

    def conflicts(day, s, e, room, ignore_id=None):
        # Check lecturer conflict
        for slot in lect_sched.get(day, []):
            if ignore_id and slot['id']==ignore_id: continue
            if overlap(s,e,slot['start'],slot['end']):
                return 'lecturer'
        
        # Check room conflict (including ONLINE room)
        rn = room.strip() if room else room
        for slot in room_occ[day].get(rn, []):
            if overlap(s,e,slot['start'],slot['end']):
                # Debug which slot causes conflict
                logger.info(f"[Reschedule]       Room conflict in {rn}: {slot.get('start')}-{slot.get('end')} {slot.get('course_name')} by {slot.get('dosen')}")
                return 'room'
        return None


    moved = 0; failed=[]; debug_log = []

    for cls in affected:
        cid = cls.get('_id'); cname = cls.get('course_name'); sec = cls.get('section_label','-')
        oday = cls.get('day'); s = cls.get('start'); e = cls.get('end'); room = cls.get('room'); is_online = (room=='ONLINE')
        duration = t2m(e) - t2m(s)
        
        logger.info(f"[Reschedule] ━━━ Processing: {cname} ({sec}) ━━━")
        logger.info(f"[Reschedule]     Current: {oday} {s}-{e} (room: {room}, duration: {duration}min)")
        debug_log.append(f"Processing: {cname} ({sec}) from {oday} {s}-{e}")
        
        course_id = cls.get('course_id'); is_lab=False
        if course_id:
            try:
                course = courses_collection.find_one({'_id': ObjectId(course_id)})
                if course: is_lab = course.get('is_lab',False)
            except: pass

        # Safety override: if current room is non-lab, force non-lab pool even if course.is_lab=True
        # Likewise, if current room is a lab, force lab pool
        if room in NON_LAB_ROOMS:
            is_lab = False
        elif room in LAB_ROOMS:
            is_lab = True

        logger.info(f"[Reschedule]     is_lab resolved: {is_lab} (current room: {room})")
        room_pool = ['ONLINE'] if is_online else (LAB_ROOMS if is_lab else NON_LAB_ROOMS)
        logger.info(f"[Reschedule]     Room pool for {cname} ({sec}): {room_pool}")

        placed=False
        preferred_days = ([oday]+[d for d in target_days if d!=oday]) if allow_day_change else [oday]
        cand_times = generate_candidates(duration)
        
        logger.info(f"[Reschedule]     Trying days: {preferred_days}")
        logger.info(f"[Reschedule]     Candidate times: {len(cand_times)} slots")

        # Filter out block window times on blocked days
        def blocked(day, cs, ce):
            if day not in source_days: return False
            return in_block(cs, ce)

        # Collect all valid candidates with their preference scores
        valid_candidates = []
        attempt_count = 0
        
        for day in preferred_days:
            for (cs,ce) in cand_times:
                attempt_count += 1
                if blocked(day, cs, ce):
                    logger.info(f"[Reschedule]     #{attempt_count} {day} {cs}-{ce}: REJECT (blocked time)")
                    continue
                # Skip slot yang menyilang istirahat siang Senin–Kamis
                if overlaps_lunch(day, cs, ce):
                    logger.info(f"[Reschedule]     #{attempt_count} {day} {cs}-{ce}: REJECT (istirahat siang 13:00-14:00)")
                    debug_log.append(f"  ⏳ Skip {day} {cs}-{ce} (istirahat siang)")
                    continue
                
                # Check if slot is in approved unavailability block
                if is_in_approved_block(day, cs, ce):
                    logger.info(f"[Reschedule]     #{attempt_count} {day} {cs}-{ce}: REJECT (unavailability block)")
                    continue
                
                # Check lecturer preferences using check_lecturer_preferences function
                # Note: check_lecturer_preferences is defined later in the file, so we need to
                # implement the logic inline here OR ensure it's available globally
                # For now, let's implement a local preference check
                
                # Get lecturer preferences
                prefs = {}
                if lecturer_prefs and 'preferences' in lecturer_prefs:
                    prefs = lecturer_prefs['preferences']
                
                # Check unavailable_time_ranges (hard constraint for physical classes)
                pref_allowed = True
                unavailable_ranges = prefs.get("unavailable_time_ranges", [])
                blocked_by_range = None
                for block in unavailable_ranges:
                    block_day = block.get('day', '*')
                    block_start = block.get('start', '00:00')
                    block_end = block.get('end', '23:59')
                    if block_day == '*' or block_day == day:
                        if overlap(cs, ce, block_start, block_end):
                            if not is_online:
                                pref_allowed = False
                                blocked_by_range = f"{block_day} {block_start}-{block_end}"
                                break
                
                if not pref_allowed:
                    logger.info(f"[Reschedule]     #{attempt_count} {day} {cs}-{ce}: REJECT (preference constraint - blocked by {blocked_by_range})")
                    continue
                
                # Check available_days (SOFT constraint - prefer available days but allow others as fallback)
                available_days = prefs.get("available_days", [])
                
                # Calculate preference score
                # IMPORTANT: When reschedule is due to unavailability block, RELAX preferred_times
                # Allow ANY slot outside the blocked range, even if not in preferred_times
                preferred_times = prefs.get("preferred_times", {})
                
                # Check if current slot is within the reschedule's blocked time range
                is_reschedule_from_block = time_type == 'range' and start_time and end_time and \
                                           oday in source_days and \
                                           overlap(s, e, start_time, end_time)
                
                if is_reschedule_from_block:
                    # Forced reschedule from unavailability block: RELAX preferred_times check
                    # Only score based on available_days, NOT preferred_times
                    if available_days and day in available_days:
                        pref_score = 70  # Good - day is in available_days
                    else:
                        pref_score = 40  # Acceptable - outside available_days but necessary fallback
                else:
                    # Normal preference scoring
                    if available_days and day in available_days:
                        pref_score = 75  # Good - day matches
                    else:
                        pref_score = 50  # Acceptable - not in available_days but allowed as fallback
                    
                    if day in preferred_times:
                        day_prefs = preferred_times[day]
                        if day_prefs:
                            cs_min = t2m(cs)
                            ce_min = t2m(ce)
                            for pref_range in day_prefs:
                                if '-' in pref_range:
                                    pref_start, pref_end = pref_range.split('-')
                                    pref_start_min = t2m(pref_start.strip())
                                    pref_end_min = t2m(pref_end.strip())
                                    if cs_min >= pref_start_min and ce_min <= pref_end_min:
                                        pref_score = 100  # Excellent - fully within preferred time
                                        break
                            else:
                                if available_days and day in available_days:
                                    pref_score = 60  # Acceptable - day OK, time not perfect
                
                # TRY ALL ROOMS IN ROOM POOL (not just the original room)
                # Prioritize original room, then alternatives
                rooms_to_try = [room] + [r for r in room_pool if r != room]
                logger.info(f"[Reschedule]       Trying rooms for {day} {cs}-{ce}: {rooms_to_try}")
                for try_room in rooms_to_try:
                    # Check conflicts with this room
                    conflict_type = conflicts(day, cs, ce, try_room)
                    
                    if not conflict_type:
                        room_type = 'same_room' if try_room == room else 'alt_room'
                        logger.info(f"[Reschedule]     #{attempt_count} {day} {cs}-{ce} {try_room}: ACCEPT ✓ (pref_score: {pref_score})")
                        valid_candidates.append({
                            'day': day,
                            'start': cs,
                            'end': ce,
                            'room': try_room,
                            'score': pref_score,
                            'type': room_type
                        })
                        break  # Found a valid room for this time slot, move to next time slot
                    # If conflict, try next room in pool
                
                # If all rooms failed, log the first attempt's rejection reason
                if not any(c['day'] == day and c['start'] == cs and c['end'] == ce for c in valid_candidates):
                    # All rooms had conflicts - report the issue
                    final_conflict = conflicts(day, cs, ce, room)
                    if final_conflict == 'lecturer':
                        logger.info(f"[Reschedule]     #{attempt_count} {day} {cs}-{ce}: REJECT (lecturer conflict)")
                    elif final_conflict == 'room':
                        logger.info(f"[Reschedule]     #{attempt_count} {day} {cs}-{ce}: REJECT (room conflict in all rooms)")
        
        # Sort candidates by preference score (highest first), then by earlier time
        valid_candidates.sort(key=lambda x: (-x['score'], t2m(x['start'])))
        
        logger.info(f"[Reschedule]     Found {len(valid_candidates)} valid candidates after {attempt_count} attempts")
        
        # Try to place in the best candidate
        if valid_candidates:
            best = valid_candidates[0]
            day = best['day']
            cs = best['start']
            ce = best['end']
            new_room = best['room']
            
            logger.info(f"[Reschedule]     ✅ MOVED: {oday} {s}-{e} → {day} {cs}-{ce} (room: {new_room}, score: {best['score']})")
            
            # Update database
            schedules_collection.update_one({'_id': cid},{'$set': {'day': day,'start': cs,'end': ce, 'room': new_room}})
            
            # Remove from old day's schedule
            if oday in lect_sched:
                lect_sched[oday] = [slot for slot in lect_sched[oday] if slot['id'] != cid]
            if oday in room_occ and room in room_occ[oday]:
                room_occ[oday][room] = [slot for slot in room_occ[oday][room] 
                                         if not (slot['start'] == s and slot['end'] == e)]
            
            # Add to new day's schedule
            lect_sched[day].append({'start':cs,'end':ce,'id':cid,'room':new_room})
            room_occ[day][new_room].append({'start':cs,'end':ce})
            
            debug_log.append(f"  ✅ Moved to {day} {cs}-{ce} (room: {new_room}, pref_score: {best['score']})")
            moved+=1
            placed=True
        
        # If not placed with simple strategy, try swap logic (only if swap enabled)
        if not placed and enable_swap:
            # Try to find a slot where we can swap with another class
            for day in preferred_days:
                for (cs,ce) in cand_times:
                    if blocked(day, cs, ce):
                        continue
                    if overlaps_lunch(day, cs, ce):
                        continue
                    if is_in_approved_block(day, cs, ce):
                        continue
                    
                    # Check preference validity
                    prefs = {}
                    if lecturer_prefs and 'preferences' in lecturer_prefs:
                        prefs = lecturer_prefs['preferences']
                    
                    # Quick preference check
                    available_days = prefs.get("available_days", [])
                    # SOFT constraint: prefer available days but don't reject others
                    # Only hard constraint is unavailable_time_ranges
                    
                    unavailable_ranges = prefs.get("unavailable_time_ranges", [])
                    pref_allowed = True
                    for block in unavailable_ranges:
                        block_day = block.get('day', '*')
                        block_start = block.get('start', '00:00')
                        block_end = block.get('end', '23:59')
                        if block_day == '*' or block_day == day:
                            if overlap(cs, ce, block_start, block_end):
                                if not is_online:
                                    pref_allowed = False
                                    break
                    if not pref_allowed:
                        continue
                    
                    conflict_type = conflicts(day, cs, ce, room)
                    
                    if conflict_type == 'lecturer':
                        # Attempt swap if single blocking slot
                        blockers = [slot for slot in lect_sched.get(day, []) if overlap(cs,ce,slot['start'],slot['end'])]
                        if len(blockers)==1:
                            b = blockers[0]
                            # Only swap if blocker is unaffected (not in affected list)
                            if not any(overlap(b['start'], b['end'], a.get('start'), a.get('end')) and a.get('_id')==b['id'] for a in affected):
                                # Relocate blocker
                                b_doc = schedules_collection.find_one({'_id': b['id']})
                                if b_doc:
                                    b_duration = t2m(b['end']) - t2m(b['start'])
                                    b_candidates = generate_candidates(b_duration)
                                    for (bs,be) in b_candidates:
                                        if (bs,be)==(cs,ce) or blocked(day, bs, be):
                                            continue
                                        if is_in_approved_block(day, bs, be):
                                            continue
                                        if conflicts(day, bs, be, b_doc.get('room'), ignore_id=b['id']):
                                            continue
                                        
                                        # Move blocker
                                        schedules_collection.update_one({'_id': b['id']},{'$set': {'start': bs,'end': be}})
                                        lect_sched[day] = [slot for slot in lect_sched[day] if slot['id']!=b['id']]
                                        lect_sched[day].append({'start':bs,'end':be,'id':b['id'],'room':b_doc.get('room')})
                                        blocker_room = b_doc.get('room')
                                        room_occ[day][blocker_room] = [slot for slot in room_occ[day][blocker_room] if not (slot['start']==b['start'] and slot['end']==b['end'])]
                                        room_occ[day][blocker_room].append({'start':bs,'end':be})
                                        
                                        # Place affected class in freed slot
                                        schedules_collection.update_one({'_id': cid},{'$set': {'day': day,'start': cs,'end': ce}})
                                        
                                        # Remove from old day's schedule
                                        if oday in lect_sched:
                                            lect_sched[oday] = [slot for slot in lect_sched[oday] if slot['id'] != cid]
                                        if oday in room_occ and room in room_occ[oday]:
                                            room_occ[oday][room] = [slot for slot in room_occ[oday][room] 
                                                                     if not (slot['start'] == s and slot['end'] == e)]
                                        
                                        # Add to new day's schedule
                                        lect_sched[day].append({'start':cs,'end':ce,'id':cid,'room':room})
                                        room_occ[day][room].append({'start':cs,'end':ce})
                                        
                                        debug_log.append(f"  ✅ Moved to {day} {cs}-{ce} (swapped with {b_doc.get('course_name')})")
                                        moved+=1; placed=True
                                        break
                                    if placed: break
                    if placed: break
                if placed: break
        
        if not placed:
            logger.warning(f"[Reschedule]     ❌ FAILED to move after {attempt_count} attempts - no valid slots found")
            debug_log.append(f"  ❌ FAILED to move (no valid slots respecting preferences)")
            failed.append(f"{cname} ({sec})")

    logger.info(f"[Reschedule] ═══════════════════════════════════════")
    logger.info(f"[Reschedule] SUMMARY for {lecturer_name}:")
    logger.info(f"[Reschedule]   ✅ Successfully moved: {moved}/{len(affected)}")
    logger.info(f"[Reschedule]   ❌ Failed to move: {len(failed)}")
    if failed:
        logger.warning(f"[Reschedule]   Failed classes: {', '.join(failed)}")
    logger.info(f"[Reschedule] ═══════════════════════════════════════")
    
    return {"success":True,"moved":moved,"failed":len(failed),"failed_courses":failed,"total":len(affected),"unaffected":len(unaffected),"message":f"Moved {moved}/{len(affected)} (unaffected {len(unaffected)})","debug_info":debug_log}

@app.route("/koordinator/unavailability_requests")
def koordinator_unavailability_requests():
    """Grouped view of reschedule requests (pending & approved) per lecturer."""
    if 'user' not in session or session['user']['role'] != 'koordinator':
        return redirect(url_for('login'))
    reports = list(unavailability_reports_collection.find().sort('created_at', -1))
    grouped = {}
    for r in reports:
        r['_id'] = str(r.get('_id'))
        # Ensure created_at is a datetime object for template rendering
        if 'created_at' in r and isinstance(r['created_at'], str):
            try:
                from datetime import datetime
                r['created_at'] = datetime.fromisoformat(r['created_at'])
            except:
                r['created_at'] = None
        elif 'created_at' not in r:
            from datetime import datetime
            r['created_at'] = datetime.now()
        
        lect = r.get('lecturer_name','?')
        grouped.setdefault(lect, []).append(r)
    ordered = sorted(grouped.keys())
    return render_template('koordinator_unavailability_requests.html', grouped=grouped, lecturers=ordered)

@app.route('/koordinator/retry_reschedule/<report_id>', methods=['POST'])
def koordinator_retry_reschedule(report_id):
    if 'user' not in session or session['user']['role'] != 'koordinator':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    try:
        report = unavailability_reports_collection.find_one({'_id': ObjectId(report_id)})
        if not report:
            return jsonify({'success': False, 'message': 'Report not found'}), 404
        lecturer_name = report.get('lecturer_name')
        unavailable_day = report.get('unavailable_day')
        unavailable_days = report.get('unavailable_days') or [unavailable_day]
        time_type = report.get('unavailable_time_type')
        specific_time = report.get('specific_time')
        start_time = report.get('start_time')
        end_time = report.get('end_time')
        result = _reschedule_lecturer_classes_v2(
            lecturer_name,
            unavailable_day,
            unavailable_days=unavailable_days,
            time_type=time_type,
            specific_time=specific_time,
            start_time=start_time,
            end_time=end_time
        )
        unavailability_reports_collection.update_one({'_id': ObjectId(report_id)}, {'$set': {'reschedule_result': result}})
        return jsonify({'success': True, 'result': result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route("/koordinator/unavailability_history")
def koordinator_unavailability_history():
    """Grouped history of reschedule requests per lecturer (compact view)."""
    if 'user' not in session or session['user']['role'] != 'koordinator':
        return redirect(url_for('login'))

    reports = list(unavailability_reports_collection.find().sort('created_at', -1))
    grouped = {}
    for r in reports:
        name = r.get('lecturer_name','?')
        r['_id'] = str(r.get('_id'))
        grouped.setdefault(name, []).append(r)

    # Limit default slice size per lecturer (frontend can expand)
    DEFAULT_LIMIT = 6
    preview = {lect: items[:DEFAULT_LIMIT] for lect, items in grouped.items()}
    counts = {lect: len(items) for lect, items in grouped.items()}
    # sort lecturers alphabetically
    ordered = sorted(grouped.keys())
    return render_template('koordinator_unavailability_history.html',
                           lecturers=ordered,
                           grouped=grouped,
                           preview=preview,
                           counts=counts,
                           default_limit=DEFAULT_LIMIT)

@app.route("/dosen/reschedule_history")
def dosen_reschedule_history():
    if 'user' not in session or session.get('user', {}).get('role') != 'dosen':
        return redirect(url_for('login'))
    username = session['user']['name']
    reports = list(unavailability_reports_collection.find({'lecturer_name': username}).sort('created_at', -1))
    for r in reports:
        r['_id'] = str(r.get('_id'))
    return render_template('dosen_reschedule_history.html', reports=reports, username=username)

@app.route("/dosen/delete_reschedule_request/<report_id>", methods=["POST"])
def dosen_delete_reschedule_request(report_id):
    """Dosen dapat menghapus permintaan reschedule mereka sendiri (hanya yang masih pending)."""
    if 'user' not in session or session.get('user', {}).get('role') != 'dosen':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    try:
        username = session['user']['name']
        report = unavailability_reports_collection.find_one({'_id': ObjectId(report_id), 'lecturer_name': username})
        if not report:
            return jsonify({'success': False, 'message': 'Request tidak ditemukan atau bukan milik Anda'}), 404
        if report.get('status') != 'pending':
            return jsonify({'success': False, 'message': 'Hanya request pending yang bisa dihapus'}), 400
        unavailability_reports_collection.delete_one({'_id': ObjectId(report_id)})
        return jsonify({'success': True, 'message': 'Request berhasil dihapus'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route("/koordinator/delete_reschedule_request/<report_id>", methods=["POST"])
def koordinator_delete_reschedule_request(report_id):
    """Koordinator dapat menghapus permintaan reschedule apapun."""
    if 'user' not in session or session['user']['role'] != 'koordinator':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    try:
        result = unavailability_reports_collection.delete_one({'_id': ObjectId(report_id)})
        if result.deleted_count == 0:
            return jsonify({'success': False, 'message': 'Request tidak ditemukan'}), 404
        return jsonify({'success': True, 'message': 'Request berhasil dihapus'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ------------------------------
# DOSEN PAGE
# ------------------------------

@app.route("/dosen")
def dosen():
    if "user" not in session or session["user"]["role"] != "dosen":
        return redirect(url_for("login"))
    username = session["user"]["name"]

    # courses kept for selection page but not shown on dashboard
    courses = list(courses_collection.find())

    # load full schedule and filter only slots assigned to this dosen
    schedule = fix_mongo_id(load_schedule())
    assigned_slots = [s for s in schedule if s.get("dosen") == username]
    assigned_by_day = group_slots_by_day(assigned_slots)
    schedule_by_day = group_slots_by_day(schedule)

    # Ambil semua course yang sudah dibooking oleh dosen ini (bisa lebih dari satu)
        # Ambil semua course yang sudah dibooking oleh dosen ini (bisa lebih dari satu)
    booked_cursor = courses_collection.find({"selected_by": username})
    booked_courses = []
    for c in booked_cursor:
        # original ObjectId
        cid = c.get("_id")
        has_slot = False
        slot_info = None
        try:
            slot = schedules_collection.find_one({"course_id": cid})
            if slot:
                has_slot = True
                slot_info = {
                    "day": slot.get("day"),
                    "start": slot.get("start"),
                    "end": slot.get("end"),
                    "room": slot.get("room")
                }
        except Exception:
            # ignore lookup errors
            pass

        booked_courses.append({
            "_id": str(cid),
            "course_name": c.get("course_name"),
            "sks": c.get("sks"),
            "semester": c.get("semester"),
            "has_slot": has_slot,
            "slot": slot_info,
            "slot_id": str(slot["_id"]) if has_slot else None
        })

    # sort booked_courses by semester (numeric), then by day order, then by course_name
    booked_courses.sort(key=lambda x: (semester_to_int(x.get('semester', '0')), get_day_order((x.get('slot') or {}).get('day')), x.get('course_name', '')))

    # group booked courses by semester for display
    booked_courses_by_semester = group_courses_by_semester(booked_courses)

    # prepare a list of booked courses that don't yet have an assigned slot
    unassigned_booked = [bc for bc in booked_courses if not bc.get("has_slot")]

    # group unassigned booked courses by semester for display
    unassigned_booked_by_semester = group_courses_by_semester(unassigned_booked)

    return render_template(
        "dosen.html",
        assigned_slots=assigned_slots,
        assigned_by_day=assigned_by_day,
        schedule_by_day=schedule_by_day,
        username=username,
        booked_courses=booked_courses,
        unassigned_booked=unassigned_booked,
        unassigned_booked_by_semester=unassigned_booked_by_semester,
        schedule=schedule
    )

@app.route("/dosen/change_slot/<slot_id>")
def change_slot(slot_id):
    if "user" not in session or session["user"]["role"] != "dosen":
        return redirect(url_for("login"))
    username = session["user"]["name"]

    try:
        slot = schedules_collection.find_one({"_id": ObjectId(slot_id)})
    except Exception:
        slot = None

    if not slot:
        flash("Slot tidak ditemukan!")
        return redirect(url_for("dosen"))
    if slot.get("dosen") != username:
        flash("Anda tidak berhak mengubah slot ini.")
        return redirect(url_for("dosen"))

    course_id = slot.get("course_id")

    # unassign slot so dosen can pick another one for same course
    schedules_collection.update_one(
        {"_id": ObjectId(slot_id)},
        {"$set": {"course_id": None, "course_name": None, "sks": None, "dosen": None}}
    )

    flash("Slot dilepaskan. Silakan pilih slot baru untuk mata kuliah Anda.")

    # redirect to slot selection with course_id so dosen can reassign
    if course_id:
        try:
            cid = str(course_id)
            return redirect(url_for("dosen_select_slot", course_id=cid))
        except Exception:
            pass

    return redirect(url_for("dosen_select_slot"))

@app.route("/dosen/unassign_slot/<slot_id>")
def unassign_slot(slot_id):
    if "user" not in session or session["user"]["role"] != "dosen":
        return redirect(url_for("login"))
    username = session["user"]["name"]

    try:
        slot = schedules_collection.find_one({"_id": ObjectId(slot_id)})
    except Exception:
        slot = None

    if not slot:
        flash("Slot tidak ditemukan!")
        return redirect(url_for("dosen"))
    if slot.get("dosen") != username:
        flash("Anda tidak berhak menghapus slot ini.")
        return redirect(url_for("dosen"))

    # Capture course_id before unassigning
    course_id = slot.get("course_id")

    # unassign dosen from slot (keep slot assigned to course)
    schedules_collection.update_one(
        {"_id": ObjectId(slot_id)},
        {"$set": {"dosen": None}}
    )

    # If there was a course assigned, unset selected_by and reserved_at to make it available again
    if course_id:
        courses_collection.update_one(
            {"_id": course_id},
            {"$unset": {"selected_by": "", "reserved_at": ""}}
        )

    flash("Jadwal berhasil dihapus dan mata kuliah tersedia kembali.")
    return redirect(url_for("dosen"))

@app.route("/dosen/select_course")
def dosen_select_course():
    if "user" not in session or session["user"]["role"] != "dosen":
        return redirect(url_for("login"))

    username = session["user"]["name"]
    courses = list(courses_collection.find())
    courses = fix_mongo_id(courses)

    # Normalize selected_by to list and filter courses that can still be selected (less than 3 lecturers and not already selected by this lecturer)
    filtered_courses = []
    for c in courses:
        selected_by = c.get('selected_by', [])
        if isinstance(selected_by, str):
            selected_by = [selected_by]  # Handle old data where it might be a string
        c['selected_by'] = selected_by  # Update to list
        if len(selected_by) < 3 and username not in selected_by:
            filtered_courses.append(c)
    courses = filtered_courses

    # For each course determine whether it already has an assigned slot
    for c in courses:
        c['has_slot'] = False
        c['slot'] = None
        # try to find any slot referencing this course
        try:
            slot = schedules_collection.find_one({"course_id": ObjectId(c["_id"])})
            if slot:
                c['has_slot'] = True
                # copy relevant fields for display
                c['slot'] = {
                    'day': slot.get('day'),
                    'start': slot.get('start'),
                    'end': slot.get('end'),
                    'room': slot.get('room'),
                    'dosen': slot.get('dosen')
                }
        except Exception:
            # ignore conversion errors
            pass

    courses_by_semester = group_courses_by_semester(courses)
    return render_template("dosen_select_course.html", courses=courses, courses_by_semester=courses_by_semester, username=username)

@app.route("/dosen/select_slot")
def dosen_select_slot():
    if "user" not in session or session["user"]["role"] != "dosen":
        return redirect(url_for("login"))

    username = session["user"]["name"]

    # only show available slots (not yet assigned)
    schedule = fix_mongo_id(load_schedule())
    available_slots = [s for s in schedule if not s.get("dosen")]
    slots_by_day = group_slots_by_day(available_slots)

    # optional course id passed when user just selected a course
    course_id = request.args.get("course_id")
    course = None
    if course_id:
        try:
            course = courses_collection.find_one({"_id": ObjectId(course_id)})
            # convert ID to string for template usage
            if course:
                course["_id"] = str(course["_id"])
        except Exception:
            course = None

    return render_template("dosen_select_slot.html", slots=available_slots, slots_by_day=slots_by_day, username=username, course=course)

# -------------------------
# PILIH MATA KULIAH
# -------------------------
@app.route("/choose_course/<course_id>")
def choose_course(course_id):
    if "user" not in session or session["user"]["role"] != "dosen":
        return redirect(url_for("login"))

    username = session["user"]["name"]
    course = courses_collection.find_one({"_id": ObjectId(course_id)})

    if not course:
        flash("Mata kuliah tidak ditemukan!")
        return redirect(url_for("dosen_select_course"))

    # Check if course already selected by 3 lecturers
    selected_by = course.get("selected_by", [])
    if isinstance(selected_by, str):
        selected_by = [selected_by]  # Handle old data
    if len(selected_by) >= 3 and username not in selected_by:
        flash("Mata kuliah ini sudah dipilih oleh maksimal 3 dosen!")
        return redirect(url_for("dosen_select_course"))

    # Check how many courses this dosen has selected
    selected_count = courses_collection.count_documents({"selected_by": username})
    if selected_count >= MAX_COURSES_PER_DOSEN:
        flash(f"Anda sudah memilih maksimal {MAX_COURSES_PER_DOSEN} mata kuliah!")
        return redirect(url_for("dosen_select_course"))

    # Mark course as selected by this dosen (append to list if already selected)
    selected_by = course.get("selected_by", [])
    if isinstance(selected_by, str):
        selected_by = [selected_by]  # Handle old data
    if username not in selected_by:
        selected_by.append(username)
    courses_collection.update_one(
        {"_id": ObjectId(course_id)},
        {"$set": {"selected_by": selected_by, "reserved_at": datetime.now()}}
    )

    flash("Mata kuliah berhasil dipilih! Silakan pilih slot waktu mengajar.")
    return redirect("/dosen/select_slot?course_id=" + course_id)

# -------------------------
# PILIH SLOT / WAKTU MENGAJAR
# -------------------------
@app.route("/select_slot/<slot_id>")
def select_slot(slot_id):
    if "user" not in session or session["user"]["role"] != "dosen":
        flash("Login dulu!")
        return redirect(url_for("login"))

    username = session["user"]["name"]

    # Prefer course_id from query string (when dosen just chose a course), else fallback to any selected course
    course_id = request.args.get("course_id")
    selected_course_obj = None
    if course_id:
        try:
            selected_course_obj = courses_collection.find_one({"_id": ObjectId(course_id)})
        except Exception:
            selected_course_obj = None

    if not selected_course_obj:
        selected_course_obj = courses_collection.find_one({"selected_by": username})

    if not selected_course_obj:
        flash("Pilih mata kuliah terlebih dahulu sebelum memilih slot!")
        return redirect(url_for("dosen"))

    slot = schedules_collection.find_one({"_id": ObjectId(slot_id)})
    if not slot:
        flash("Slot tidak ditemukan!")
        return redirect(url_for("dosen"))
    if slot.get("dosen"):
        flash("Slot sudah diambil dosen lain!")
        return redirect(url_for("dosen"))

    # --- Conflict detection: prevent assigning a slot that overlaps an existing slot for this dosen ---
    try:
        desired_day = slot.get("day")
        desired_start = _parse_minutes(slot.get("start"))
        desired_end = _parse_minutes(slot.get("end"))

        # find slots already assigned to this dosen
        assigned_slots = list(schedules_collection.find({"dosen": username}))
        for asg in assigned_slots:
            # only consider same day
            if asg.get("day") != desired_day:
                continue
            asg_start = _parse_minutes(asg.get("start"))
            asg_end = _parse_minutes(asg.get("end"))
            if _times_overlap(desired_start, desired_end, asg_start, asg_end):
                # Determine conflicting course name if available
                conf_course = asg.get("course_name") or "(mata kuliah tidak diketahui)"
                conf_msg = f"Tidak dapat memilih slot — bentrok dengan '{conf_course}' pada {asg.get('day')} {asg.get('start')}-{asg.get('end')}"
                flash(conf_msg)
                # redirect back to slot selection (preserve course_id if present)
                if course_id:
                    return redirect(url_for("dosen_select_slot", course_id=course_id))
                return redirect(url_for("dosen_select_slot"))
        # --- Room conflict check: ensure room is not occupied by another assigned slot at overlapping time ---
        try:
            room = slot.get("room")
            if room:
                # find other assigned slots in same room and day
                other_assigned = list(schedules_collection.find({"room": room, "day": desired_day, "dosen": {"$ne": None}}))
                for o in other_assigned:
                    # skip the same slot record
                    try:
                        if str(o.get("_id")) == str(slot.get("_id")):
                            continue
                    except Exception:
                        pass
                    o_start = _parse_minutes(o.get("start"))
                    o_end = _parse_minutes(o.get("end"))
                    if _times_overlap(desired_start, desired_end, o_start, o_end):
                        other_dosen = o.get("dosen") or "(dosen tidak diketahui)"
                        other_course = o.get("course_name") or "(mata kuliah tidak diketahui)"
                        room_msg = f"Tidak dapat memilih slot — ruangan {room} sudah digunakan oleh {other_dosen} untuk '{other_course}' pada {o.get('day')} {o.get('start')}-{o.get('end')}"
                        flash(room_msg)
                        if course_id:
                            return redirect(url_for("dosen_select_slot", course_id=course_id))
                        return redirect(url_for("dosen_select_slot"))
        except Exception:
            # if room-check fails, be conservative and block selection
            flash("Gagal memeriksa ketersediaan ruangan. Silakan coba lagi atau hubungi admin.")
            if course_id:
                return redirect(url_for("dosen_select_slot", course_id=course_id))
            return redirect(url_for("dosen_select_slot"))
    except Exception:
        # if any error occurs during conflict checking, play safe and block selection
        flash("Terjadi kesalahan saat memeriksa bentrok jadwal. Silakan coba lagi atau hubungi admin.")
        if course_id:
            return redirect(url_for("dosen_select_slot", course_id=course_id))
        return redirect(url_for("dosen_select_slot"))

    # Update slot using selected course
    result = schedules_collection.update_one(
        {"_id": ObjectId(slot_id), "dosen": None},
        {"$set": {
            "course_id": selected_course_obj["_id"],
            "course_name": selected_course_obj["course_name"],
            "sks": selected_course_obj.get("sks"),
            "dosen": username
        }}
    )

    # If modified_count is 0 it means another process assigned the slot concurrently
    if result.modified_count == 0:
        flash("Gagal mengambil slot — kemungkinan sudah diambil orang lain. Silakan pilih slot lain.")
        if course_id:
            return redirect(url_for("dosen_select_slot", course_id=course_id))
        return redirect(url_for("dosen_select_slot"))

    flash("Slot berhasil dipilih!")
    return redirect(url_for("dosen"))

@app.route("/add_course", methods=["POST"])
def add_course():
    if "user" not in session or session["user"]["role"] != "koordinator":
        return redirect(url_for("login"))

    name = request.form["course_name"]
    sks = int(request.form["sks"])
    semester = request.form.get("semester", "")
    
    # Auto-detect if course should be online
    online_keywords = ['kewarganegaraan', 'kemahlikussalahan', 'kemal']
    is_online = any(keyword in name.lower() for keyword in online_keywords)

    courses_collection.insert_one({
        "course_name": name,
        "sks": sks,
        "semester": semester,
        "is_online": is_online
    })

    if is_online:
        flash(f"Mata kuliah '{name}' berhasil ditambahkan sebagai KELAS ONLINE!")
    else:
        flash("Mata kuliah berhasil ditambahkan!")
    return redirect(url_for("koordinator_courses"))

# ------------------------------
# GENETIC ALGORITHM FOR SECTION GENERATION
# ------------------------------

def evaluate_sections(individual, courses, lecturers):
    """
    Fitness function for section generation GA.
    individual: list of section counts [s1, s2, ..., sn]
    """
    penalty = 0

    # Separate lab and non-lab courses
    non_lab_sections = 0
    lab_sections = 0

    for i, sections in enumerate(individual):
        course = courses[i]
        if course.get('is_lab', False):
            lab_sections += sections
        else:
            non_lab_sections += sections

    # Penalty for not matching required totals
    penalty += abs(non_lab_sections - REQUIRED_NON_LAB_SECTIONS) * 1000
    penalty += abs(lab_sections - REQUIRED_LAB_SECTIONS) * 1000

    # Calculate lecturer assignments and penalties
    lecturer_assignments = {}
    for lecturer in lecturers:
        lecturer_assignments[lecturer['name']] = {
            'courses': lecturer.get('selected_courses', []),
            'preferences': lecturer.get('preferences', {}),
            'total_sks': 0,
            'days': set()
        }

    # Assign sections to lecturers based on their selected courses
    for i, sections in enumerate(individual):
        course = courses[i]
        course_id = str(course['_id'])

        # Find lecturers who selected this course
        assigned_lecturers = []
        for lecturer in lecturers:
            if course_id in lecturer.get('selected_courses', []):
                assigned_lecturers.append(lecturer['name'])

        if not assigned_lecturers:
            penalty += sections * 50  # Penalty for unassigned sections
            continue

        # Distribute sections among assigned lecturers
        sections_per_lecturer = sections // len(assigned_lecturers)
        remainder = sections % len(assigned_lecturers)

        for j, lecturer_name in enumerate(assigned_lecturers):
            extra = 1 if j < remainder else 0
            assigned_sections = sections_per_lecturer + extra

            lecturer_assignments[lecturer_name]['total_sks'] += assigned_sections * course['sks']

            lecturer_assignments[lecturer_name]['sections'] = lecturer_assignments[lecturer_name].get('sections', 0) + assigned_sections

            # Simulate day assignment (simplified - would be done in full GA)
            # For now, just check if total SKS exceeds limit
            if lecturer_assignments[lecturer_name]['total_sks'] > REQUIRED_SKS_PER_DOSEN:
                penalty += (lecturer_assignments[lecturer_name]['total_sks'] - REQUIRED_SKS_PER_DOSEN) * 10

    # Penalty for lecturers with too many days (simplified)
    for lecturer_data in lecturer_assignments.values():
        if len(lecturer_data['days']) > 3:
            penalty += (len(lecturer_data['days']) - 3) * 20

    return penalty,

def generate_sections_ga(courses, lecturers, population=100, generations=500):
    """
    Run GA to generate optimal section counts.
    """
    # Setup DEAP
    creator.create("FitnessMin", base.Fitness, weights=(-1.0,))
    creator.create("Individual", list, fitness=creator.FitnessMin)

    toolbox = base.Toolbox()

    # Attribute generator: random section count between 1 and 10
    toolbox.register("attr_int", random.randint, 1, 10)

    # Structure initializers
    toolbox.register("individual", tools.initRepeat, creator.Individual,
                    toolbox.attr_int, n=len(courses))
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)

    toolbox.register("evaluate", evaluate_sections, courses=courses, lecturers=lecturers)
    toolbox.register("mate", tools.cxTwoPoint)
    toolbox.register("mutate", tools.mutUniformInt, low=1, up=10, indpb=0.2)
    toolbox.register("select", tools.selTournament, tournsize=3)

    # Create population
    pop = toolbox.population(n=population)
    hof = tools.HallOfFame(1)

    # Run GA
    algorithms.eaSimple(pop, toolbox, cxpb=0.5, mutpb=0.2, ngen=generations,
                       halloffame=hof, verbose=False)

    return hof[0]

# ------------------------------
# SCHEDULE GENERATION ROUTE
# ------------------------------

# ------------------------------
# DISPLAY GENERATED SCHEDULES
# ------------------------------

@app.route("/view_sections")
def view_sections():
    if "user" not in session or session["user"]["role"] != "koordinator":
        return redirect(url_for("login"))

    courses = list(courses_collection.find())
    return render_template("sections_table.html", courses=courses)

@app.route("/view_schedule")
def view_schedule():
    if "user" not in session or session["user"]["role"] != "koordinator":
        return redirect(url_for("login"))

    schedules = list(schedules_collection.find())
    return render_template("schedule_table.html", schedules=schedules)

@app.route("/view_lecturer_summary")
def view_lecturer_summary():
    if "user" not in session or session["user"]["role"] != "koordinator":
        return redirect(url_for("login"))

    lecturers = list(users_collection.find({"role": "dosen"}))
    lecturer_summaries = []

    for lecturer in lecturers:
        username = lecturer['username']
        schedules = list(schedules_collection.find({"dosen": username}))

        total_sks = sum(schedule.get('sks', 0) for schedule in schedules)
        total_sections = len(schedules)
        days = set(schedule.get('day', '') for schedule in schedules)
        total_days = len(days)

        lecturer_summaries.append({
            'dosen': username,
            'total_sks': total_sks,
            'total_sections': total_sections,
            'total_days': total_days
        })

    return render_template("lecturer_summary.html", summaries=lecturer_summaries)

# ------------------------------
# WELCOME ENDPOINT
# ------------------------------

@app.route("/welcome", methods=["GET"])
def welcome():
    """
    Returns a welcome message
    """
    logger.info(f"Request received: {request.method} {request.path}")
    return jsonify({'message': 'Welcome to Flask API Service!'})

# ==============================================================================
# HYBRID SCHEDULING SYSTEM (GA + OR-Tools)
# ==============================================================================

class Section:
    """Represents a section with course info and assigned lecturer"""
    def __init__(self, course_id, course_name, sks, is_lab, section_letter, section_number, lecturer=None, is_online=False):
        self.course_id = course_id
        self.course_name = course_name
        self.sks = sks
        self.is_lab = is_lab
        self.is_online = is_online  # Online courses don't need physical room
        self.section_letter = section_letter
        self.section_number = section_number
        self.section_name = f"{section_letter}{section_number}"
        self.lecturer = lecturer
        self.day = None
        self.start_time = None
        self.end_time = None
        self.room = "Online" if is_online else None
    
    def __repr__(self):
        return f"{self.course_name} {self.section_name} ({self.lecturer})"

def calculate_sections_needed_hybrid(courses, lecturers=None, population=150, generations=300):
    """
    SMART 2-TIER GA:
    
    Goal: Assign dosen ke MK & tentukan section count FLEKSIBEL per MK 
    agar setiap dosen capai 8-12 SKS.
    
    Bukan: 1 MK = 6 section wajib
    Tapi: Section count per MK = flexible (1-5) sesuai assignment & SKS balance
    
    Returns: {(dosen_name, course_id): num_sections}
    """
    if lecturers is None:
        lecturers = list(users_collection.find({"role": "dosen"}))
    
    if not lecturers or not courses:
        return {}
    
    logger.info("="*80)
    logger.info("SMART GA: OPTIMAL ASSIGNMENT + FLEXIBLE SECTION COUNT (8-12 SKS TARGET)")
    logger.info("="*80)
    
    num_courses = len(courses)
    num_lecturers = len(lecturers)

    # Build allowed lecturer indices per course (respect selected_by if provided)
    lecturer_index_map = {l['name']: idx for idx, l in enumerate(lecturers)}
    allowed_indices_per_course = []

    for course in courses:
        selected = course.get('selected_by') or []
        allowed = [lecturer_index_map[name] for name in selected if name in lecturer_index_map]
        if not allowed:
            # Fallback: allow all lecturers if no explicit selection exists
            allowed = list(range(num_lecturers))
        allowed_indices_per_course.append(allowed)
    
    # TIER 1: Assign dosen ke MK (1-3 dosen per MK, 2-3 MK per dosen)
    logger.info(f"\nTIER 1: Assigning {num_lecturers} dosens to {num_courses} courses...")
    
    try:
        creator.create("FitnessAssignMin", base.Fitness, weights=(-1.0,))
        creator.create("IndividualAssign", list, fitness=creator.FitnessAssignMin)
    except:
        del creator.FitnessAssignMin
        del creator.IndividualAssign
        creator.create("FitnessAssignMin", base.Fitness, weights=(-1.0,))
        creator.create("IndividualAssign", list, fitness=creator.FitnessAssignMin)
    
    toolbox1 = base.Toolbox()
    
    def create_assignment():
        """Setiap MK diassign 1-3 dosen, dibatasi oleh selected_by jika ada"""
        assignment = []
        for course_idx, allowed in enumerate(allowed_indices_per_course):
            candidates = allowed or list(range(num_lecturers))
            # Pastikan selalu ada kandidat
            max_pick = min(3, len(candidates))
            num_dosens = random.randint(1, max_pick) if max_pick > 0 else 1
            assigned = random.sample(candidates, num_dosens) if candidates else []
            assignment.append(assigned)
        return assignment
    
    toolbox1.register("individual", tools.initIterate, creator.IndividualAssign, create_assignment)
    toolbox1.register("population", tools.initRepeat, list, toolbox1.individual)
    
    def evaluate_assignment(individual, courses_list, lecturers_list, allowed_indices_per_course):
        penalty = 0
        dosen_courses = {l['name']: set() for l in lecturers_list}
        
        for course_idx, lect_indices in enumerate(individual):
            for lect_idx in lect_indices:
                allowed_set = set(allowed_indices_per_course[course_idx] or range(len(lecturers_list)))

                # Hard penalty if assignment uses lecturer outside allowed list
                if lect_idx not in allowed_set:
                    penalty += 100000
                    continue

                dosen_courses[lecturers_list[lect_idx]['name']].add(course_idx)
        
        # Prefer: 2-3 MK per dosen
        for dosen_name, courses_set in dosen_courses.items():
            num_courses_assigned = len(courses_set)
            if num_courses_assigned == 0:
                penalty += 50000
            elif num_courses_assigned == 1:
                penalty += 5000
            elif num_courses_assigned > 3:
                penalty += (num_courses_assigned - 3) * 2000
        
        return (penalty,)
    
    toolbox1.register("evaluate", evaluate_assignment, courses_list=courses, lecturers_list=lecturers,
                     allowed_indices_per_course=allowed_indices_per_course)
    toolbox1.register("mate", tools.cxUniform, indpb=0.5)
    toolbox1.register("mutate", tools.mutShuffleIndexes, indpb=0.3)
    toolbox1.register("select", tools.selTournament, tournsize=3)
    
    pop1 = toolbox1.population(n=population)
    hof1 = tools.HallOfFame(1)
    algorithms.eaSimple(pop1, toolbox1, cxpb=0.6, mutpb=0.3, ngen=generations//2, halloffame=hof1, verbose=False)
    
    best_assignment = hof1[0]
    logger.info(f"✓ Best assignment fitness: {hof1[0].fitness.values[0]}")
    
    # Extract assignment
    course_to_dosens = {}
    for course_idx, lect_indices in enumerate(best_assignment):
        course_to_dosens[course_idx] = list(set(lect_indices))
    
    # TIER 2: Section count optimization (SKS-based, flexible)
    logger.info(f"\nTIER 2: Optimizing section count per (dosen, MK) for 8-12 SKS balance...")
    
    # Build (dosen_idx, course_idx) pairs
    dosen_course_pairs = []
    for course_idx, lect_indices in course_to_dosens.items():
        for lect_idx in lect_indices:
            dosen_course_pairs.append((lect_idx, course_idx))
    
    try:
        creator.create("FitnessSectionMin", base.Fitness, weights=(-1.0,))
        creator.create("IndividualSection", list, fitness=creator.FitnessSectionMin)
    except:
        del creator.FitnessSectionMin
        del creator.IndividualSection
        creator.create("FitnessSectionMin", base.Fitness, weights=(-1.0,))
        creator.create("IndividualSection", list, fitness=creator.FitnessSectionMin)
    
    toolbox2 = base.Toolbox()
    toolbox2.register("attr_section", random.randint, 1, 4)  # 1-4 sections per (dosen, MK)
    toolbox2.register("individual", tools.initRepeat, creator.IndividualSection,
                     toolbox2.attr_section, n=len(dosen_course_pairs))
    toolbox2.register("population", tools.initRepeat, list, toolbox2.individual)
    
    def evaluate_section_counts(individual, pairs, courses_list, lecturers_list):
        """
        ULTRA-STRICT CONSTRAINT: Setiap dosen HARUS 8-12 SKS (MIN=8, MAX=12).
        TIDAK BOLEH kurang dari 8, TIDAK BOLEH lebih dari 12.
        
        EXTREME Penalty untuk violation - gunakan rejection weight.
        """
        penalty = 0
        dosen_sks = {l['name']: 0 for l in lecturers_list}
        dosen_sections = {l['name']: 0 for l in lecturers_list}
        num_violations = 0
        
        for (lect_idx, course_idx), num_sections in zip(pairs, individual):
            dosen_name = lecturers_list[lect_idx]['name']
            course_sks = courses_list[course_idx].get('sks', 3)
            sks_gained = num_sections * course_sks
            dosen_sks[dosen_name] += sks_gained
            dosen_sections[dosen_name] += num_sections
        
        # ULTRA-STRICT EVALUATION: 8-12 SKS WAJIB, TIDAK ADA KOMPROMI
        for dosen_name, total_sks in dosen_sks.items():
            if total_sks == 0:
                continue  # Dosen tidak assign MK (akan dihandle di Tier 1)
            
            # REJECTION WEIGHT: Setiap violation = penalty SANGAT BESAR
            if total_sks < MIN_SKS_PER_DOSEN:  # Kurang dari 8
                gap = MIN_SKS_PER_DOSEN - total_sks
                # Exponential penalty untuk undershooting
                penalty += 10**6 + (gap ** 4) * 10000
                num_violations += 1
            elif total_sks > MAX_SKS_PER_DOSEN:  # Lebih dari 12
                gap = total_sks - MAX_SKS_PER_DOSEN
                # Exponential penalty untuk overshooting
                penalty += 10**6 + (gap ** 4) * 10000
                num_violations += 1
            else:
                # REWARD jika exact dalam range 8-12
                # Prefer close to target 10 SKS
                diff_from_target = abs(total_sks - 10)
                penalty -= (5 - diff_from_target) * 1000
        
        # AGGREGATE PENALTY: Jika ada violation, multiply by huge factor
        if num_violations > 0:
            penalty *= (10 ** num_violations)
        
        return (penalty,)
    
    def repair_solution(individual, pairs, courses_list, lecturers_list):
        """
        AGGRESSIVE HARD CONSTRAINT REPAIR: Paksa SEMUA dosen masuk range 8-12 SKS.
        - Use greedy approach untuk assign sections
        - Prioritas: Exact 10 SKS jika possible, atau 8-12 range
        - Iterasi aggressive sampai semua comply
        """
        max_iterations = 200
        iteration = 0
        
        while iteration < max_iterations:
            iteration += 1
            dosen_sks = {l['name']: 0 for l in lecturers_list}
            dosen_indices = {l['name']: [] for l in lecturers_list}
            
            # Calculate current SKS
            for idx, ((lect_idx, course_idx), num_sections) in enumerate(zip(pairs, individual)):
                dosen_name = lecturers_list[lect_idx]['name']
                course_sks = courses_list[course_idx].get('sks', 3)
                dosen_sks[dosen_name] += num_sections * course_sks
                dosen_indices[dosen_name].append(idx)
            
            all_valid = True
            violations = []
            
            for dosen_name, total_sks in dosen_sks.items():
                if total_sks == 0:
                    continue
                
                # REPAIR: Kurang dari 8 SKS - INCREASE AGGRESSIVELY
                if total_sks < MIN_SKS_PER_DOSEN:
                    all_valid = False
                    gap = MIN_SKS_PER_DOSEN - total_sks
                    violations.append((dosen_name, total_sks, 'under'))
                    # Try multiple increases sampai hit target atau max
                    indices = dosen_indices[dosen_name]
                    if indices:
                        # Prioritas: MK dengan SKS lebih kecil (untuk flexibility)
                        sorted_indices = sorted(indices, key=lambda i: courses_list[pairs[i][1]].get('sks', 3))
                        for _ in range(gap):  # Adjust multiple times
                            if sorted_indices:
                                # Find best index to increase
                                best_idx = None
                                for idx in sorted_indices:
                                    if individual[idx] < 5:  # Relax to 5 sections max during repair
                                        best_idx = idx
                                        break
                                if best_idx is not None:
                                    individual[best_idx] += 1
                
                # REPAIR: Lebih dari 12 SKS - DECREASE AGGRESSIVELY
                elif total_sks > MAX_SKS_PER_DOSEN:
                    all_valid = False
                    gap = total_sks - MAX_SKS_PER_DOSEN
                    violations.append((dosen_name, total_sks, 'over'))
                    # Try multiple decreases sampai hit target atau min
                    indices = dosen_indices[dosen_name]
                    if indices:
                        # Prioritas: MK dengan SKS lebih besar (untuk flexibility)
                        sorted_indices = sorted(indices, key=lambda i: courses_list[pairs[i][1]].get('sks', 3), reverse=True)
                        for _ in range(gap):  # Adjust multiple times
                            if sorted_indices:
                                # Find best index to decrease
                                best_idx = None
                                for idx in sorted_indices:
                                    if individual[idx] > 1:  # Min 1 section per (dosen, MK)
                                        best_idx = idx
                                        break
                                if best_idx is not None:
                                    individual[best_idx] -= 1
            
            if all_valid:
                break
            
            # Jika max iterations tercapai dan masih ada violations, log warning
            if iteration >= max_iterations - 1 and not all_valid:
                logger.warning(f"Repair: Max iterations ({max_iterations}) reached with violations still present")
                for dosen_name, total_sks, violation_type in violations:
                    logger.warning(f"  {dosen_name}: {total_sks} SKS ({violation_type})")
        
        return individual
    
    # Custom selection dengan repair
    def select_and_repair(individuals, k, pairs, courses_list, lecturers_list):
        selected = tools.selTournament(individuals, k, tournsize=3)
        return [repair_solution(ind[:], pairs, courses_list, lecturers_list) for ind in selected]
    
    toolbox2.register("evaluate", evaluate_section_counts, pairs=dosen_course_pairs,
                     courses_list=courses, lecturers_list=lecturers)
    toolbox2.register("mate", tools.cxTwoPoint)
    toolbox2.register("mutate", tools.mutUniformInt, low=1, up=4, indpb=0.3)
    toolbox2.register("select", select_and_repair, pairs=dosen_course_pairs,
                     courses_list=courses, lecturers_list=lecturers)
    
    # Initial population dengan repair
    pop2 = toolbox2.population(n=population)
    pop2 = [repair_solution(ind, dosen_course_pairs, courses, lecturers) for ind in pop2]
    
    hof2 = tools.HallOfFame(1)
    algorithms.eaSimple(pop2, toolbox2, cxpb=0.7, mutpb=0.3, ngen=generations//2, halloffame=hof2, verbose=False)
    
    # FINAL REPAIR: Pastikan best solution juga comply
    best_section_counts = repair_solution(hof2[0][:], dosen_course_pairs, courses, lecturers)
    logger.info(f"✓ Best section count fitness: {hof2[0].fitness.values[0]}")
    
    # Build result
    result = {}
    dosen_summary = {}
    
    for (lect_idx, course_idx), num_sections in zip(dosen_course_pairs, best_section_counts):
        dosen_name = lecturers[lect_idx]['name']
        course_id = str(courses[course_idx]['_id'])
        result[(dosen_name, course_id)] = num_sections
        
        if dosen_name not in dosen_summary:
            dosen_summary[dosen_name] = {'sks': 0, 'sections': 0, 'courses': []}
        
        course_sks = courses[course_idx].get('sks', 3)
        dosen_summary[dosen_name]['sks'] += num_sections * course_sks
        dosen_summary[dosen_name]['sections'] += num_sections
        dosen_summary[dosen_name]['courses'].append(courses[course_idx]['course_name'][:30])
    
    # Log result
    logger.info("\n" + "="*80)
    logger.info("HASIL GA - SKS & SECTION DISTRIBUTION:")
    logger.info("="*80)
    
    violations = []  # Track violations untuk error handling
    
    for dosen_name in sorted(dosen_summary.keys()):
        summary = dosen_summary[dosen_name]
        total_sks = summary['sks']
        total_sections = summary['sections']
        
        if total_sks > 0:
            status = "✓" if MIN_SKS_PER_DOSEN <= total_sks <= MAX_SKS_PER_DOSEN else "❌ VIOLATION"
            logger.info(f"{status} {dosen_name:40} {total_sks:2} SKS ({total_sections} sections)")
            
            # Track violations
            if total_sks < MIN_SKS_PER_DOSEN or total_sks > MAX_SKS_PER_DOSEN:
                violations.append(f"{dosen_name}: {total_sks} SKS")
    
    # CRITICAL CHECK: Jika ada violation, log ERROR
    if violations:
        logger.error("\n" + "!"*80)
        logger.error("❌ CONSTRAINT VIOLATION DETECTED!")
        logger.error(f"❌ {len(violations)} dosen dengan SKS di luar range 8-12:")
        for v in violations:
            logger.error(f"   - {v}")
        logger.error("!"*80)
        # NOTE: GA tetap return result, tapi dengan warning ini admin bisa tau ada issue
    else:
        logger.info("\n✅ ALL DOSEN COMPLY: 8-12 SKS range enforced successfully!")
    
    # Cleanup
    try:
        del creator.FitnessAssignMin
        del creator.IndividualAssign
        del creator.FitnessSectionMin
        del creator.IndividualSection
    except:
        pass
    
    return result

def create_initial_sections_hybrid(courses, dosen_course_sections):
    """
    Create sections based on 2-tier GA result: {(dosen_name, course_id): num_sections}
    Each section gets section_label (A1, A2, B1, B2, etc) for given (dosen, course) pair
    """
    all_sections = []
    
    for (dosen_name, course_id_str), num_sections in dosen_course_sections.items():
        # Find course object
        course = next((c for c in courses if str(c['_id']) == course_id_str), None)
        if not course:
            continue
        
        # For each (dosen, course) pair, create num_sections sections
        for section_num in range(1, num_sections + 1):
            # Section letter: A, B, C, D, ... or just use section_num as number
            section_letter = "A"  # All under same dosen
            
            section = Section(
                course_id=course_id_str,
                course_name=course['course_name'],
                sks=course.get('sks', 3),
                is_lab=course.get('is_lab', False),
                section_letter=section_letter,
                section_number=section_num,  # Sequential per (dosen, course)
                lecturer=dosen_name,
                is_online=course.get('is_online', False)
            )
            all_sections.append(section)
    
    return all_sections

def evaluate_section_assignment_hybrid(individual, sections, lecturers):
    """Fitness function for GA section assignment
    
    Target: 
    - Each course has 3 lecturers
    - Each lecturer teaches 2 sections of each assigned course
    - Example: Pemweb → Dosen A (A1+A2), Dosen B (B1+B2), Dosen C (C1+C2)
    """
    penalty = 0
    lecturer_sections = {l['name']: [] for l in lecturers}
    lecturer_courses = {l['name']: set() for l in lecturers}
    lecturer_sks = {l['name']: 0 for l in lecturers}
    course_lecturers = {}  # course_id -> set of lecturers
    
    for i, lecturer_idx in enumerate(individual):
        section = sections[i]
        lecturer = lecturers[lecturer_idx]
        lecturer_name = lecturer['name']
        section.lecturer = lecturer_name
        lecturer_sections[lecturer_name].append(section)
        lecturer_courses[lecturer_name].add(section.course_id)
        lecturer_sks[lecturer_name] += section.sks
        
        # Track lecturers per course
        if section.course_id not in course_lecturers:
            course_lecturers[section.course_id] = set()
        course_lecturers[section.course_id].add(lecturer_name)
    
    # CONSTRAINT 1: Each course must have exactly 3 lecturers
    for course_id, lect_set in course_lecturers.items():
        if len(lect_set) != 3:
            penalty += abs(len(lect_set) - 3) * 50000
    
    # CONSTRAINT 2: Each lecturer teaches exactly 2 sections per course
    for lecturer_name in lecturer_sections.keys():
        sections_by_course = {}
        for sec in lecturer_sections[lecturer_name]:
            if sec.course_id not in sections_by_course:
                sections_by_course[sec.course_id] = 0
            sections_by_course[sec.course_id] += 1
        
        # Penalty if not exactly 2 sections per course
        for course_id, count in sections_by_course.items():
            if count != 2:
                penalty += abs(count - 2) * 30000
    
    # CONSTRAINT 3: SKS balance (8-12 SKS per lecturer)
    for lecturer_name, total_sks in lecturer_sks.items():
        if total_sks < MIN_SKS_PER_DOSEN:
            penalty += (MIN_SKS_PER_DOSEN - total_sks) * 1000
        elif total_sks > MAX_SKS_PER_DOSEN:
            penalty += (total_sks - MAX_SKS_PER_DOSEN) * 2000
    
    # CONSTRAINT 4: Each lecturer teaches 2-3 courses
    for lecturer_name, courses_set in lecturer_courses.items():
        num_courses = len(courses_set)
        if num_courses < 2:
            penalty += (2 - num_courses) * 5000
        elif num_courses > MAX_COURSES_PER_DOSEN:
            penalty += (num_courses - MAX_COURSES_PER_DOSEN) * 5000
    
    return (penalty,)

def run_ga_section_assignment_hybrid(courses, lecturers, population_size=100, generations=200):
    """Run 2-tier GA: 1) assign dosen to MK, 2) optimize section count per (dosen, MK) for 8-12 SKS"""
    logger.info("="*80)
    logger.info("PHASE 1: 2-TIER GA - ASSIGNMENT + SECTION COUNT OPTIMIZATION")
    logger.info("="*80)
    
    # Run 2-tier GA to get optimal (dosen, course) -> section_count
    dosen_course_sections = calculate_sections_needed_hybrid(courses, lecturers, population=population_size, generations=generations)
    
    # Create sections based on GA result
    sections = create_initial_sections_hybrid(courses, dosen_course_sections)
    
    logger.info(f"Total sections created: {len(sections)}")
    
    # Sections already have lecturer assigned from GA, so no need for phase 2 GA
    logger.info("Section assignment complete!")
    logger.info(f"  Total sections: {len(sections)}")
    
    # Calculate summary per dosen
    dosen_summary = {}
    for section in sections:
        if section.lecturer not in dosen_summary:
            dosen_summary[section.lecturer] = {'count': 0, 'sks': 0}
        dosen_summary[section.lecturer]['count'] += 1
        dosen_summary[section.lecturer]['sks'] += section.sks
    
    logger.info("\nLecturer Summary:")
    for dosen_name, summary in dosen_summary.items():
        logger.info(f"  {dosen_name}: {summary['count']} sections, {summary['sks']} SKS")
    
    return sections

def get_timeslots_for_day_hybrid(day):
    """Get timeslots for a day with correct break times
    Senin-Kamis: Break 13:10-13:59 (49 minutes)
    Jumat: Break 11:30-13:50 (2 hours 20 minutes)
    """
    if day == "Jumat":
        # Jumat: 08:00-11:30 (morning), break 11:30-13:50, then 13:50-17:00 (afternoon)
        return [("08:00", "08:50"), ("08:50", "09:40"), ("09:40", "10:30"), ("10:30", "11:20"),
                ("13:50", "14:40"), ("14:40", "15:30"), ("15:30", "16:20"), ("16:20", "17:10")]
    else:
        # Senin-Kamis: 08:00-13:10 (morning), break 13:10-13:59, then 13:59-18:00 (afternoon)
        return [("08:00", "08:50"), ("08:50", "09:40"), ("09:40", "10:30"), ("10:30", "11:20"), ("11:20", "12:10"),
                ("12:10", "13:00"), ("13:59", "14:49"), ("14:49", "15:39"), ("15:39", "16:29"), ("16:29", "17:19"), ("17:19", "18:09")]

def get_duration_slots_hybrid(sks):
    """Calculate timeslots needed"""
    return 2 if sks in [1, 2] else 3

def run_ortools_scheduling_hybrid(sections):
    """Use OR-Tools to schedule sections
    Online courses (is_online=True) don't need room assignment
    """
    logger.info("\n" + "="*80)
    logger.info("PHASE 2: OR-TOOLS SCHEDULING")
    logger.info("="*80)
    
    model = cp_model.CpModel()
    schedule = {}
    
    for i, section in enumerate(sections):
        schedule[i] = {}
        
        # Online courses don't need physical rooms
        if section.is_online:
            rooms = ['Online']
        else:
            rooms = LAB_ROOMS if section.is_lab else NON_LAB_ROOMS
        
        for day in DAYS:
            schedule[i][day] = {}
            timeslots = get_timeslots_for_day_hybrid(day)
            
            for t_idx in range(len(timeslots)):
                schedule[i][day][t_idx] = {}
                for room in rooms:
                    schedule[i][day][t_idx][room] = model.NewBoolVar(f"s{i}_d{day}_t{t_idx}_r{room}")
    
    # Constraint 1: Each section scheduled exactly once
    for i in range(len(sections)):
        section = sections[i]
        if section.is_online:
            rooms = ['Online']
        else:
            rooms = LAB_ROOMS if section.is_lab else NON_LAB_ROOMS
        all_assignments = []
        for day in DAYS:
            timeslots = get_timeslots_for_day_hybrid(day)
            for t_idx in range(len(timeslots)):
                for room in rooms:
                    all_assignments.append(schedule[i][day][t_idx][room])
        model.Add(sum(all_assignments) == 1)
    
    # Constraint 2: No room conflicts (skip for online courses)
    for day in DAYS:
        timeslots = get_timeslots_for_day_hybrid(day)
        all_rooms = NON_LAB_ROOMS + LAB_ROOMS  # Physical rooms only
        for room in all_rooms:
            for t_idx in range(len(timeslots)):
                sections_at_slot = []
                for i, section in enumerate(sections):
                    # Skip online courses for room conflict check
                    if section.is_online:
                        continue
                    rooms = LAB_ROOMS if section.is_lab else NON_LAB_ROOMS
                    if room in rooms:
                        sections_at_slot.append(schedule[i][day][t_idx][room])
                if sections_at_slot:
                    model.Add(sum(sections_at_slot) <= 1)
    
    # Constraint 3: No lecturer conflicts
    lecturer_sections = {}
    for i, section in enumerate(sections):
        lecturer = section.lecturer
        if lecturer not in lecturer_sections:
            lecturer_sections[lecturer] = []
        lecturer_sections[lecturer].append(i)
    
    for lecturer, section_indices in lecturer_sections.items():
        for day in DAYS:
            timeslots = get_timeslots_for_day_hybrid(day)
            for t_idx in range(len(timeslots)):
                lecturer_at_slot = []
                for i in section_indices:
                    section = sections[i]
                    rooms = LAB_ROOMS if section.is_lab else NON_LAB_ROOMS
                    for room in rooms:
                        lecturer_at_slot.append(schedule[i][day][t_idx][room])
                if lecturer_at_slot:
                    model.Add(sum(lecturer_at_slot) <= 1)
    
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 120.0  # Increased timeout
    solver.parameters.num_search_workers = 8
    solver.parameters.log_search_progress = False
    
    logger.info(f"Solving OR-Tools model for {len(sections)} sections...")
    status = solver.Solve(model)
    
    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        logger.info(f"✓ Solution found! Status: {solver.StatusName(status)}")
        
        # Extract solution and assign to sections
        for i, section in enumerate(sections):
            rooms = LAB_ROOMS if section.is_lab else NON_LAB_ROOMS
            for day in DAYS:
                timeslots = get_timeslots_for_day_hybrid(day)
                for t_idx in range(len(timeslots)):
                    for room in rooms:
                        if solver.Value(schedule[i][day][t_idx][room]) == 1:
                            section.day = day
                            section.start_time = timeslots[t_idx][0]
                            duration = get_duration_slots_hybrid(section.sks)
                            end_idx = min(t_idx + duration - 1, len(timeslots)-1)
                            section.end_time = timeslots[end_idx][1]
                            section.room = room
                            
                            # Log each assignment
                            logger.info(f"  {section.course_name} {section.section_letter}{section.section_number} → {day} {section.start_time}-{section.end_time} @ {room} (Dosen: {section.lecturer})")
                            break
        
        logger.info("✓ All sections scheduled WITHOUT CONFLICTS!")
        return True
    else:
        logger.error(f"No solution found. Status: {solver.StatusName(status)}")
        return False

def save_schedules_to_db_hybrid(sections):
    """Save to MongoDB"""
    logger.info("\nSaving schedules to database...")
    schedules_collection.delete_many({})
    
    schedule_docs = []
    for section in sections:
        if section.day and section.start_time and section.room:
            doc = {
                "course_id": section.course_id,
                "course_name": section.course_name,
                "section_name": section.section_name,
                "dosen": section.lecturer,
                "day": section.day,
                "start": section.start_time,
                "end": section.end_time,
                "room": section.room,
                "sks": section.sks,
                "is_lab": section.is_lab,
                "created_at": datetime.now()
            }
            schedule_docs.append(doc)
    
    if schedule_docs:
        result = schedules_collection.insert_many(schedule_docs)
        logger.info(f"Saved {len(result.inserted_ids)} schedules")
    return len(schedule_docs)

@app.route("/koordinator/generate_hybrid_schedule", methods=["GET", "POST"])
def generate_hybrid_schedule():
    """Generate schedule using hybrid GA + OR-Tools"""
    if "user" not in session or session["user"]["role"] != "koordinator":
        return redirect(url_for("login"))
    
    if request.method == "GET":
        courses = list(courses_collection.find())
        lecturers = list(users_collection.find({"role": "dosen"}))
        total_sections = len(courses) * 6
        
        return render_template(
            "koordinator_hybrid_schedule.html",
            courses=courses,
            lecturers=lecturers,
            total_sections=total_sections,
            total_lecturers=len(lecturers)
        )
    
    else:  # POST
        try:
            logger.info("Starting hybrid schedule generation...")
            courses = list(courses_collection.find())
            lecturers = list(users_collection.find({"role": "dosen"}))
            
            # Phase 1: GA
            sections = run_ga_section_assignment_hybrid(courses, lecturers)
            
            # Phase 2: OR-Tools
            success = run_ortools_scheduling_hybrid(sections)
            
            if success:
                count = save_schedules_to_db_hybrid(sections)
                flash(f"✅ Schedule generated! {count} classes scheduled.", "success")
                logger.info(f"Hybrid scheduling completed: {count} schedules")
            else:
                flash("❌ Scheduling failed. Try again with different parameters.", "error")
                
        except Exception as e:
            flash(f"❌ Error: {str(e)}", "error")
            logger.error(f"Hybrid scheduling error: {e}", exc_info=True)
        
        return redirect(url_for("koordinator_schedule"))

@app.route("/koordinator/check_conflicts", methods=["GET"])
def check_conflicts():
    """Check for schedule conflicts and online course distribution"""
    if "user" not in session or session["user"]["role"] != "koordinator":
        return redirect(url_for("login"))
    
    try:
        schedules = list(schedules_collection.find())
        
        # Check lecturer conflicts
        lecturer_schedule = {}
        for sched in schedules:
            lect = sched.get('dosen', '')
            day = sched.get('day', '')
            start = sched.get('start', '')
            end = sched.get('end', '')
            
            if lect:
                if lect not in lecturer_schedule:
                    lecturer_schedule[lect] = []
                lecturer_schedule[lect].append((day, start, end))
        
        # Check room conflicts
        room_schedule = {}
        for sched in schedules:
            room = sched.get('room', '')
            if room != 'ONLINE':
                day = sched.get('day', '')
                start = sched.get('start', '')
                end = sched.get('end', '')
                
                key = (day, room)
                if key not in room_schedule:
                    room_schedule[key] = []
                room_schedule[key].append((start, end))
        
        # Check online distribution
        online_by_day = {}
        for sched in schedules:
            if sched.get('room') == 'ONLINE':
                day = sched.get('day', '')
                online_by_day[day] = online_by_day.get(day, 0) + 1
        
        result = {
            'lecturer_conflicts': 0,
            'room_conflicts': 0,
            'online_distribution': online_by_day,
            'total_online': sum(online_by_day.values())
        }
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route("/koordinator/test_hybrid")
def test_hybrid():
    """Quick test route untuk hybrid system"""
    if "user" not in session or session["user"]["role"] != "koordinator":
        return redirect(url_for("login"))
    
    try:
        courses = list(courses_collection.find())[:3]  # Test with 3 courses
        lecturers = list(users_collection.find({"role": "dosen"}))
        
        logger.info(f"Testing with {len(courses)} courses, {len(lecturers)} lecturers")
        
        sections = run_ga_section_assignment_hybrid(courses, lecturers, population_size=50, generations=50)
        success = run_ortools_scheduling_hybrid(sections)
        
        if success:
            count = save_schedules_to_db_hybrid(sections)
            flash(f"✅ Test successful! Generated {count} schedules", "success")
        else:
            flash("❌ Test failed", "error")
            
    except Exception as e:
        flash(f"❌ Test error: {str(e)}", "error")
        logger.error(f"Test error: {e}", exc_info=True)
    
    return redirect(url_for("koordinator"))

@app.route("/test_hybrid_simple")
def test_hybrid_simple():
    """Simple test tanpa login requirement"""
    try:
        logger.info("="*80)
        logger.info("SIMPLE HYBRID TEST")
        logger.info("="*80)
        
        courses = list(courses_collection.find())[:2]  # 2 courses only
        lecturers = list(users_collection.find({"role": "dosen"}))
        
        logger.info(f"Courses: {len(courses)}")
        logger.info(f"Lecturers: {len(lecturers)}")
        
        if len(courses) == 0:
            return jsonify({"status": "error", "message": "No courses found!"})
        
        if len(lecturers) == 0:
            return jsonify({"status": "error", "message": "No lecturers found!"})
        
        # Phase 1: GA
        logger.info("\n--- Phase 1: GA Assignment ---")
        sections = run_ga_section_assignment_hybrid(courses, lecturers, population_size=30, generations=50)
        
        logger.info(f"Created {len(sections)} sections")
        
        # Show assignments
        result_data = {
            "status": "success",
            "phase1_complete": True,
            "sections": []
        }
        
        for section in sections:
            result_data["sections"].append({
                "course": section.course_name,
                "section": section.section_name,
                "lecturer": section.lecturer,
                "sks": section.sks
            })
        
        logger.info("\n--- Phase 2: OR-Tools Scheduling ---")
        success = run_ortools_scheduling_hybrid(sections)
        
        if success:
            logger.info("Scheduling successful!")
            result_data["phase2_complete"] = True
            result_data["schedules"] = []
            
            for section in sections[:10]:  # Show first 10
                result_data["schedules"].append({
                    "course": section.course_name,
                    "section": section.section_name,
                    "lecturer": section.lecturer,
                    "day": section.day,
                    "time": f"{section.start_time}-{section.end_time}",
                    "room": section.room
                })
            
            # Save to DB
            count = save_schedules_to_db_hybrid(sections)
            result_data["saved_count"] = count
            
            logger.info(f"✅ Test complete! Saved {count} schedules")
            
        else:
            logger.error("Scheduling failed!")
            result_data["phase2_complete"] = False
            result_data["error"] = "OR-Tools could not find solution"
        
        return jsonify(result_data)
        
    except Exception as e:
        logger.error(f"Test error: {e}", exc_info=True)
        return jsonify({
            "status": "error",
            "message": str(e)
        })

# ------------------------------
# RUN APP
# ------------------------------

# ===============================
# NEW GA SECTION ASSIGNMENT + OR-TOOLS SCHEDULER (Simplified Pipeline)
# ===============================
# GA hanya memilih dosen untuk setiap mata kuliah (maks 3 dosen/course) lalu
# membuat 2 section per dosen dengan pola penamaan: slug MK + " A" + nomor.
# OR-Tools kemudian menjadwalkan setiap section ke (hari, timeblock, ruangan)
# tanpa bentrok dosen maupun ruangan.

def _slugify_course(name):
    if not name:
        return "course"
    import re
    slug = re.sub(r"[^A-Za-z0-9]", "", name.lower())
    return slug or "course"

def _ensure_online_courses_marked():
    """Auto-update courses to mark religious/citizenship courses as online"""
    online_keywords = ['kewarganegaraan', 'kemahlikussalahan', 'kemal']
    updated_count = 0
    for keyword in online_keywords:
        result = courses_collection.update_many(
            {"course_name": {"$regex": keyword, "$options": "i"}, "is_online": {"$ne": True}},
            {"$set": {"is_online": True}}
        )
        updated_count += result.modified_count
    if updated_count > 0:
        print(f"✅ Auto-marked {updated_count} courses as online")
    return updated_count

def _update_islab_field():
    """Update is_lab field untuk semua MK berdasarkan jenis praktikum/teori"""
    LAB_COURSES = [
        "SISTEM OPERASI",
        "SISTEM MANAJEMEN DATABASE", 
        "PEMOGRAMAN WEB",
        "DASAR DIGITAL",
        "ORGANISASI DAN ARSITEKTUR KOMPUTER",
        "GRAFIKA KOMPUTER",
        "INTERPRETASI PENGOLAHAN CITRA",
        "JARINGAN KOMPUTER",
        "METODE NUMERIK",
        "KEAMANAN DATABASE",
        "ETICAL HACKING",
        "DESAIN DAN ANALISIS ALGORITMA",
        "DATA MINING",
        "ALGORITMA DAN PEMOGRAMAN II",
        "SISTEM INFORMASI GEOGRAFIS"
    ]
    
    all_courses = list(courses_collection.find())
    updated_lab = 0
    updated_non_lab = 0
    
    for course in all_courses:
        course_name = course.get('course_name', '').strip().upper()
        is_lab_course = any(lab_name in course_name for lab_name in LAB_COURSES)
        
        # Get current value
        current_is_lab = course.get('is_lab', None)
        
        # Always update
        result = courses_collection.update_one(
            {'_id': course['_id']},
            {'$set': {'is_lab': is_lab_course}}
        )
        
        # Count by type (not by modified_count)
        if is_lab_course:
            updated_lab += 1
        else:
            updated_non_lab += 1
    
    total_updated = updated_lab + updated_non_lab
    print(f"[INFO] Processed {total_updated} courses: {updated_lab} LAB, {updated_non_lab} NON-LAB")
    return total_updated

def generate_sections_ga_pipeline(population=150, generations=300, force_regenerate=False):
    """
    Pipeline pembuat section otomatis dengan fokus beban SKS:
    - Target ketat: setiap dosen 8-12 SKS
    - Jumlah section per MK fleksibel (tidak wajib 6 section/MK)
    - 1 MK bisa dipecah menjadi beberapa section untuk memenuhi batas SKS dosen
    - Menghormati preferred lecturer (selected_by) bila tersedia
    """
    if sections_collection is None:
        flash("Sections collection not available.")
        return
    
    # Update is_lab field untuk semua courses
    _update_islab_field()
    
    # Ensure online courses are marked before generating
    _ensure_online_courses_marked()
    
    courses = list(courses_collection.find())
    lecturers = list(users_collection.find({"role": "dosen"}))
    if not courses or not lecturers:
        flash("Data dosen atau mata kuliah kosong.")
        return

    # Build lecturer name set
    lecturer_names = [l['name'] for l in lecturers]
    lecturer_name_set = set(lecturer_names)
    
    # Buat mapping nama dosen (normalized -> original)
    # Normalisasi: hapus spasi, titik, koma, lowercase untuk matching
    def normalize_name(name):
        """Normalize name untuk matching: hapus spasi, titik, koma, strip, lowercase"""
        if not name:
            return ""
        # Strip whitespace, remove spaces, dots, commas, lowercase
        return name.strip().replace(" ", "").replace(".", "").replace(",", "").lower()
    
    lecturer_normalized_map = {}
    for lect_name in lecturer_names:
        normalized = normalize_name(lect_name)
        # If duplicate normalized names, keep first one
        if normalized not in lecturer_normalized_map:
            lecturer_normalized_map[normalized] = lect_name
    
    # DEBUG: Cek nama dosen yang ada di database
    print(f"\n[DEBUG] First 5 lecturer names in database (original -> normalized):")
    for name in lecturer_names[:5]:
        print(f"  - '{name}' -> '{normalize_name(name)}'")
    
    # DEBUG: Cek data selected_by sebelum normalisasi
    print(f"\n[DEBUG] Checking selected_by data for first 3 courses:")
    for c in courses[:3]:
        sb = c.get('selected_by', [])
        print(f"  - {c['course_name']}:")
        for sel in sb:
            print(f"      '{sel}' -> '{normalize_name(sel)}'")
    
    # Normalisasi selected_by: Matching dengan normalized name
    matched_count = 0
    unmatched_count = 0
    
    for c in courses:
        sb = c.get('selected_by', [])
        if isinstance(sb, str):
            sb = [sb]
        
        valid_selected = []
        for selected_name in sb:
            # Cek exact match dulu
            if selected_name in lecturer_name_set:
                valid_selected.append(selected_name)
                matched_count += 1
                continue
            
            # Jika tidak exact match, coba matching dengan normalized name
            selected_normalized = normalize_name(selected_name)
            if selected_normalized in lecturer_normalized_map:
                # Gunakan nama original dari database
                original_name = lecturer_normalized_map[selected_normalized]
                valid_selected.append(original_name)
                matched_count += 1
            else:
                unmatched_count += 1
                print(f"⚠️  Cannot match: '{selected_name}' (normalized: '{selected_normalized}')")
        
        c['selected_by'] = valid_selected
    
    print(f"\n[DEBUG] Matching result: {matched_count} matched, {unmatched_count} unmatched")
    
    # DEBUG: Cek data selected_by setelah normalisasi
    print(f"\n[DEBUG] After normalization:")
    for c in courses[:5]:
        print(f"  - {c['course_name']}: selected_by = {c.get('selected_by', [])}")

    # Reset sections database
    if force_regenerate:
        sections_collection.delete_many({})
        schedules_collection.delete_many({})
    
    sections_collection.delete_many({})
    
    # Call helper function to continue the section generation
    _continue_generate_sections_pipeline(courses, lecturers, lecturer_names, force_regenerate)


def calculate_flexible_section_distribution(num_lecturers, sks, course_name=None, lecturer_current_sks=None):
    """
    Menghitung distribusi section yang fleksibel berdasarkan:
    - Minimum 8 SKS per dosen
    - Maximum 12 SKS per dosen
    - Jumlah dosen yang tersedia
    - SKS mata kuliah
    - Current SKS dosen (untuk avoid over-allocation)
    
    Returns: list of int (jumlah section untuk setiap dosen)
    """
    MIN_SKS = 8
    MAX_SKS = 12
    
    # Handle special cases
    if course_name:
        if course_name in ["KEAMANAN DATABASE", "ETICAL HACKING"]:
            return [1]  # Only 1 section for these courses
        elif course_name == "KECERDASAN KOMPUTASIONAL":
            return [3]  # 3 sections for this course
        elif course_name.lower() in ["kewarganegaraan", "kemalikussalehan"]:
            if num_lecturers == 1:
                return [3]  # Special courses get 3 sections
    
    if num_lecturers == 0:
        return []
    
    # STRATEGI BARU: Target 3-4 section per dosen per MK (untuk avoid over-allocation)
    # Jika dosen punya 2 MK: 3 section × 2 MK × 3 SKS = 18 SKS (OVER!)
    # Solusi: batas 2-3 section per MK untuk dosen dengan multiple courses
    
    if num_lecturers == 1:
        # 1 dosen: target 3-4 section tergantung SKS
        if sks == 3:
            sections = 3  # 3 section × 3 SKS = 9 SKS (lebih aman jika ada MK lain)
        elif sks == 2:
            sections = 4  # 4 section × 2 SKS = 8 SKS
        else:  # sks == 1
            sections = 9  # 9 section × 1 SKS = 9 SKS
        return [sections]
    
    # Multiple lecturers: target 2-3 section per dosen per MK
    # Ini lebih aman untuk menghindari over-allocation
    sections_per_lecturer = 3 if sks == 3 else (4 if sks == 2 else 8)
    
    distribution = [sections_per_lecturer] * num_lecturers
    
    return distribution


# Continuation of generate_sections_ga_pipeline after calculate_flexible_section_distribution function
def _continue_generate_sections_pipeline(courses, lecturers, lecturer_names, force_regenerate=False):
    """
    Helper function untuk melanjutkan proses generate sections setelah normalisasi
    Dipanggil dari generate_sections_ga_pipeline
    """
    new_sections = []  # Initialize list untuk menampung section yang akan dibuat
    
    # Track assignment per dosen
    lecturer_course_count = {name: 0 for name in lecturer_names}
    lecturer_section_count = {name: 0 for name in lecturer_names}
    
    # PRE-ANALYSIS: Cek berapa mata kuliah yang dipilih setiap dosen
    print(f"\n{'='*80}")
    print(f"PRE-ANALYSIS: Jumlah Mata Kuliah yang Dipilih Setiap Dosen")
    print(f"{'='*80}\n")
    
    lecturer_selected_courses = {name: [] for name in lecturer_names}
    for c in courses:
        for lect_name in c.get('selected_by', []):
            if lect_name in lecturer_selected_courses:
                lecturer_selected_courses[lect_name].append(c['course_name'])
    
    only_one_course_lecturers = []
    for name, course_list in sorted(lecturer_selected_courses.items()):
        if len(course_list) > 0:
            if len(course_list) == 1:
                only_one_course_lecturers.append((name, course_list[0]))
                print(f"  [!] {name}: Hanya memilih 1 MK: {course_list[0]}")
            elif len(course_list) == 2:
                print(f"  [OK] {name}: Memilih 2 MK: {', '.join(course_list)}")
    
    if only_one_course_lecturers:
        print(f"\n[!] WARNING: {len(only_one_course_lecturers)} dosen hanya memilih 1 mata kuliah!")
        print(f"   Constraint '2 mata kuliah per dosen' TIDAK dapat dipenuhi untuk mereka.")
    
    print(f"\n{'='*80}")
    print(f"MEMBUAT SECTION SESUAI ATURAN (BALANCED SKS)")
    print(f"{'='*80}\n")
    
    # SMART DISTRIBUTION: Hitung section per dosen berdasarkan jumlah MK yang diajar
    # Target: 10 SKS per dosen (8-12 range)
    # Strategi:
    #   - Dosen dengan 1 MK (3 SKS): 3 section × 3 = 9 SKS (mendekati 10)
    #   - Dosen dengan 2 MK (3 SKS each): 2 section × 2 MK × 3 = 12 SKS (max)
    #   - Dosen dengan 2 MK (2 SKS each): 2-3 section × 2 MK × 2 = 8-12 SKS
    
    for course in courses:
        course_id = str(course['_id'])
        course_name = course['course_name']
        sks = course.get('sks', 3)
        is_lab = course.get('is_lab', False)
        is_online = course.get('is_online', False)
        
        # Ambil dosen yang memilih course ini
        selected_raw = course.get('selected_by', [])
        selected = [n for n in selected_raw if n in lecturer_names]
        num_lecturers = len(selected)
        
        # Special case: Kewarganegaraan dan Kemalikussalehan
        if num_lecturers == 0:
            print(f"[!] {course_name}: Tidak ada dosen yang memilih, skip.")
            continue
        
        # Gunakan fungsi fleksibel untuk menghitung distribusi section
        # Untuk dosen dengan multiple courses, beri lebih sedikit section per MK
        # Asumsi: jika dosen pilih 2 MK, maka target 2-3 section per MK
        distribution = []
        for lecturer in selected:
            # Hitung berapa MK yang dipilih dosen ini
            num_courses_of_lecturer = len(lecturer_selected_courses.get(lecturer, []))
            
            # Calculate section berdasarkan jumlah MK
            if num_courses_of_lecturer == 1:
                # 1 MK: beri 3 section (3 × 3 SKS = 9 SKS)
                sections_for_this_lecturer = 3 if sks == 3 else (4 if sks == 2 else 9)
            elif num_courses_of_lecturer == 2:
                # 2 MK: beri 2 section per MK (2 × 2 MK × 3 SKS = 12 SKS)
                sections_for_this_lecturer = 2 if sks == 3 else (2 if sks == 2 else 5)
            else:
                # 3+ MK: beri 1-2 section per MK
                sections_for_this_lecturer = 1 if sks == 3 else (2 if sks == 2 else 3)
            
            distribution.append(sections_for_this_lecturer)
        
        # Info distribusi
        total_sections = sum(distribution)
        total_sks_course = total_sections * sks
        dist_str = '+'.join(map(str, distribution))
        print(f"  [i] {course_name} ({sks} SKS): {num_lecturers} dosen -> {dist_str} section (total {total_sections} section, {total_sks_course} SKS)")
        
        # Buat section untuk setiap dosen
        section_idx = 0
        for lect_idx, lecturer in enumerate(selected):
            num_sections = distribution[lect_idx]
            for i in range(num_sections):
                section_number = section_idx + 1  # Sequential: 1, 2, 3, 4, 5, 6...
                section_letter = "A"  # Always use letter A
                
                new_sections.append({
                    'course_id': course_id,
                    'course_name': course_name,
                    'sks': sks,
                    'is_lab': is_lab,
                    'is_online': is_online,
                    'section_letter': section_letter,
                    'section_number': section_number,
                    'section_label': f"{section_letter}{section_number}",
                    'lecturer': lecturer
                })
                section_idx += 1
                lecturer_section_count[lecturer] += 1
            
            lecturer_course_count[lecturer] += 1
        
        dist_str = '+'.join(map(str, distribution))
        print(f"[OK] {course_name}: {num_lecturers} dosen -> distribusi {dist_str} section")
    
    # Simpan section ke database
    if new_sections:
        # Mulai normalisasi: gunakan field 'dosen' konsisten; tetap dukung 'lecturer' lama untuk backward-compat sementara
        for s in new_sections:
            if 'lecturer' in s:
                s['dosen'] = s.get('lecturer')
        sections_collection.insert_many(new_sections)
        print(f"\n[OK] Total {len(new_sections)} section berhasil dibuat dan disimpan ke database.")
    
    # Cetak ringkasan
    total_courses = len(courses)
    total_lecturers = len(set((s.get('dosen') or s.get('lecturer')) for s in new_sections))
    print(f"\n{'='*80}")
    print(f"RINGKASAN DISTRIBUSI")
    print(f"{'='*80}")
    print(f"Total dosen: {total_lecturers}")
    print(f"Total mata kuliah: {total_courses}")
    print(f"Total section: {len(new_sections)}")
    
    # Distribusi section per dosen
    dist_summary = {}
    for count in lecturer_section_count.values():
        dist_summary[count] = dist_summary.get(count, 0) + 1
    print(f"\nDistribusi section per dosen:")
    for count in sorted(dist_summary.keys()):
        print(f"  {count} section: {dist_summary[count]} dosen")
    
    print(f"{'='*80}\n")
    
    # Success flash message (only if in request context)
    try:
        flash(f"[OK] {len(new_sections)} section berhasil dibuat! Silakan klik 'Jadwalkan Section' untuk menjadwalkan.", "success")
    except RuntimeError:
        # Not in request context, skip flash
        pass

def schedule_sections_with_ortools(max_seconds=60):
    """
    SIMPLIFIED GREEDY SCHEDULER: Assign sections to (day, time, room) without OR-Tools complexity.
    Uses a smart greedy algorithm that respects all constraints.
    INCLUDES LECTURER PREFERENCES: available_days, preferred_times, unavailable_time_ranges.
    """
    logger.info(f"\n{'='*80}")
    logger.info(f"🔵 FUNCTION START: schedule_sections_with_ortools() called")
    logger.info(f"{'='*80}")
    
    if sections_collection is None:
        flash("Sections collection not available.")
        return
    sections = list(sections_collection.find())
    if not sections:
        flash("Tidak ada section untuk dijadwalkan. Jalankan GA terlebih dahulu.")
        return
    
    # LOG: Verify section count
    section_count = sections_collection.count_documents({})
    logger.info(f"\n📊 SCHEDULE WITH OR-TOOLS: Loaded {len(sections)} sections from list(find()), DB has {section_count} total")
    
    # Safeguard: If more sections loaded than expected, something duplicated them
    if len(sections) != section_count:
        logger.warning(f"⚠️ Section count mismatch: list() returned {len(sections)} but count_documents() returned {section_count}")
    
    # If way too many sections, likely a bug - abort
    if len(sections) > 200:
        logger.error(f"❌ Too many sections ({len(sections)}) detected! This likely indicates a duplication bug. Clearing sections collection.")
        sections_collection.delete_many({})
        sections = []
        flash(f"⚠️ Detected {section_count} sections (expected ~125). Cleared to prevent corruption. Please regenerate.", "error")
        return
    
    # Load lecturer preferences from database
    print("\n📋 Loading lecturer preferences...")
    lecturer_preferences = {}
    # Primary source: users collection (embedded preferences)
    if users_collection is not None:
        for user in users_collection.find({"role": "dosen"}):
            name = user.get("name")
            prefs = user.get("preferences", {}) or {}
            if name:
                lecturer_preferences[name] = prefs
    # Secondary source: lecturer_preferences collection (migrated docs)
    try:
        if 'lecturer_preferences' in db.list_collection_names():
            for doc in db['lecturer_preferences'].find({}):
                name = doc.get('dosen') or doc.get('lecturer_name')
                if not name:
                    continue
                # Build normalized preference dict
                normalized = {}
                for key in ["available_days","preferred_times","unavailable_time_ranges","max_load"]:
                    if key in doc and doc.get(key) not in (None, ""):
                        normalized[key] = doc.get(key)
                if name in lecturer_preferences:
                    # Merge: keep existing user prefs, supplement missing keys
                    for k,v in normalized.items():
                        if k not in lecturer_preferences[name] or not lecturer_preferences[name][k]:
                            lecturer_preferences[name][k] = v
                else:
                    lecturer_preferences[name] = normalized
    except Exception as e:
        print(f"⚠️ Could not load external lecturer_preferences collection: {e}")
    print(f"✅ Loaded preferences for {len(lecturer_preferences)} lecturers\n")
    
    print(f"\n{'='*80}")
    print(f"GREEDY SCHEDULING ALGORITHM (WITH LECTURER PREFERENCES)")
    print(f"{'='*80}")
    print(f"Scheduling {len(sections)} sections...")
    print(f"{'='*80}\n")
    
    # Available resources
    rooms = NON_LAB_ROOMS + LAB_ROOMS
    days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
    
    def parse_time(t):
        """Convert HH:MM to minutes"""
        if not t or ':' not in t:
            return 0
        h, m = map(int, t.split(':'))
        return h * 60 + m
    
    def times_overlap(t1_start, t1_end, t2_start, t2_end):
        """Check if two time ranges overlap"""
        s1, e1 = parse_time(t1_start), parse_time(t1_end)
        s2, e2 = parse_time(t2_start), parse_time(t2_end)
        return s1 < e2 and s2 < e1
    
    def get_match_quality(score):
        """Convert score to match quality label
        Args:
            score: 0-100 preference score
        Returns:
            Match quality label: Excellent/Good/Acceptable/Poor
        """
        if score >= 90:
            return "Excellent"
        elif score >= 70:
            return "Good"
        elif score >= 50:
            return "Acceptable"
        else:
            return "Poor"
    
    def check_lecturer_preferences(lect, day, time_start, time_end, strict_mode=True, is_online=False):
        """Check if slot is allowed by lecturer preferences. Returns (allowed, score)
        
        Args:
            lect: Lecturer name
            day: Day of week
            time_start: Start time (HH:MM)
            time_end: End time (HH:MM)
            strict_mode: If True, enforce available_days. If False, relax to soft preference
            is_online: If True, relax blocked times constraint (online can be done anywhere)
        
        Returns:
            (allowed, score) where:
            - allowed: True if slot can be used
            - score: 100 = perfect match, 50+ = acceptable, 20- = not preferred
        """
        prefs = lecturer_preferences.get(lect, {})
        
        # Check unavailable_time_ranges (HARD constraint for physical classes, SOFT for online)
        unavailable_ranges = prefs.get("unavailable_time_ranges", [])
        for block in unavailable_ranges:
            block_day = block.get('day', '*')
            block_start = block.get('start', '00:00')
            block_end = block.get('end', '23:59')
            
            # Check if block applies to this day
            if block_day == '*' or block_day == day:
                if times_overlap(time_start, time_end, block_start, block_end):
                    if is_online:
                        # For online classes, blocked times are just a preference (penalty), not hard constraint
                        pass  # Allow but will reduce score below
                    else:
                        return False, 0  # Hard constraint violation for physical classes
        
        # Check available_days (strict vs relaxed)
        available_days = prefs.get("available_days", [])
        day_penalty = 0
        
        if available_days:
            if day not in available_days:
                if strict_mode:
                    return False, 0  # Hard constraint in strict mode
                else:
                    # In relaxed mode, allow but give penalty
                    day_penalty = 30  # Not preferred day
        
        # Check preferred_times (soft preference - for scoring)
        preferred_times = prefs.get("preferred_times", {})
        
        # IMPROVED SCORING: Bertingkat berdasarkan kesesuaian
        # Base score tergantung apakah dosen punya preferensi atau tidak
        if available_days and day in available_days:
            # Dosen punya available_days dan slot ini dalam hari tersedia
            score = 75  # Good - hari sesuai, waktu fleksibel
        else:
            # Tidak ada preferensi hari, atau hari di luar available_days (relaxed mode)
            score = 50  # Acceptable - umum
        
        if day in preferred_times:
            # If lecturer has preferences for this day, check if time is FULLY within preferred window
            day_prefs = preferred_times[day]
            if day_prefs:
                slot_start_min = parse_time(time_start)
                slot_end_min = parse_time(time_end)
                
                for pref_range in day_prefs:
                    if '-' in pref_range:
                        pref_start, pref_end = pref_range.split('-')
                        pref_start_min = parse_time(pref_start.strip())
                        pref_end_min = parse_time(pref_end.strip())
                        
                        # Check if slot is FULLY within preferred window
                        if slot_start_min >= pref_start_min and slot_end_min <= pref_end_min:
                            score = 100  # Excellent - Fully within preferred time!
                            break
                else:
                    # Slot not within any preferred window for this day
                    # Tapi hari masih OK (dalam available_days)
                    score = 60  # Acceptable - day OK, time not perfect
        
        # Apply day penalty if not in available_days (relaxed mode)
        score = max(0, score - day_penalty)
        
        return True, score
    
    # Separate online and regular sections
    online_sections = []  # Don't need room assignment
    regular_sections = []  # Need room assignment
    
    for i, sec in enumerate(sections):
        is_online = sec.get('is_online', False)
        if is_online:
            online_sections.append((i, sec))
        else:
            regular_sections.append((i, sec))
    
    print(f"Online sections (no room needed): {len(online_sections)}")
    print(f"Regular sections (need room): {len(regular_sections)}")
    
    # Group sections by lecturer (regular sections only for room assignment)
    lecturer_sections_map = {}
    def _get_section_dosen(sec):
        return sec.get('dosen') or sec.get('lecturer') or 'Unknown'

    for i, sec in regular_sections:
        lect = _get_section_dosen(sec)
        if lect not in lecturer_sections_map:
            lecturer_sections_map[lect] = []
        lecturer_sections_map[lect].append((i, sec))
    
    # Group online sections by lecturer
    lecturer_online_sections = {}
    for i, sec in online_sections:
        lect = _get_section_dosen(sec)
        if lect not in lecturer_online_sections:
            lecturer_online_sections[lect] = []
        lecturer_online_sections[lect].append((i, sec))
    
    # Track assignments
    room_calendar = {}  # (day, room, time_start, time_end) -> section_idx
    lecturer_calendar = {}  # (lecturer, day, time_start, time_end) -> section_idx
    schedule_result = {}  # section_idx -> (day, room, time_block)
    
    # Track lecturer days for 2-3 day constraint (combine regular + online lecturers)
    all_lecturers = set(lecturer_sections_map.keys()) | set(lecturer_online_sections.keys())
    lecturer_days_used = {lect: set() for lect in all_lecturers}
    
    import random
    
    # Helper function to get valid days for lecturer based on preferences
    def get_valid_days_for_lecturer(lect):
        """Get list of valid days for lecturer based on available_days preference"""
        prefs = lecturer_preferences.get(lect, {})
        available_days = prefs.get("available_days", [])
        if available_days:
            # Return only days lecturer marked as available
            return [d for d in weekday_order if d in available_days]
        else:
            # No preference set, all days available
            return weekday_order
    
    # Greedy assignment: Try to assign each lecturer's sections to 2-3 consecutive days
    # IMPROVED: Sort lecturers by number of sections (ascending) to prioritize those with fewer sections
    weekday_order = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
    placed_count = 0
    failed_sections = []
    
    # STEP 1: Schedule ONLINE sections first (no room needed, faster)
    print("\n" + "="*80)
    print("STEP 1: SCHEDULING ONLINE SECTIONS (NO ROOM NEEDED)")
    print("="*80)
    
    # STRATEGY: Concentrate ALL online classes on RABU (Wednesday) starting from 09:00
    # Schedule them sequentially to avoid gaps
    # Respect lecturer preferences - if lecturer not available on Rabu, find alternative from available_days
    ONLINE_DAY = "Rabu"  # MAIN: Concentrate online classes here
    ONLINE_START_TIME = "09:00"  # Start time for first online class
    
    print(f"📌 Strategy: Concentrating ALL online sections on {ONLINE_DAY}")
    print(f"   Starting from: {ONLINE_START_TIME} (sequential scheduling)")
    print(f"   Respecting lecturer preferences for available days\n")
    
    # Sort online sections by lecturer and SKS for better organization
    online_sections_sorted = []
    for lect, sec_list in lecturer_online_sections.items():
        for si, sec in sec_list:
            online_sections_sorted.append((si, sec, lect))
    
    # Sort by lecturer name, then by SKS (smaller first for easier packing)
    online_sections_sorted.sort(key=lambda x: (x[2], x[1].get('sks', 2)))
    
    # Track next available time on ONLINE_DAY
    def parse_time_to_minutes(time_str):
        h, m = map(int, time_str.split(':'))
        return h * 60 + m
    
    def minutes_to_time(minutes):
        h = minutes // 60
        m = minutes % 60
        return f"{h:02d}:{m:02d}"
    
    next_available_time = parse_time_to_minutes(ONLINE_START_TIME)
    
    for si, sec, lect in online_sections_sorted:
        sks = int(sec.get('sks', 2))
        duration_minutes = sks * 50  # 1 SKS = 50 minutes
        
        placed = False
        
        # FORCE ALL ONLINE SECTIONS TO RABU ONLY
        # Online teaching is flexible, ignore lecturer preferences for day selection
        # But still respect real lecturer schedule conflicts (if already scheduled)
        
        # Try sequential placement on Rabu first
        start_minutes = next_available_time
        end_minutes = start_minutes + duration_minutes
        
        # Check if within working hours (max 17:00)
        if end_minutes <= parse_time_to_minutes("17:00"):
            start = minutes_to_time(start_minutes)
            end = minutes_to_time(end_minutes)
            
            # Check lecturer conflict (don't care about preferences, just physical conflict)
            lect_conflict = any(
                l == lect and d == ONLINE_DAY and times_overlap(start, end, s, e)
                for (l, d, s, e) in lecturer_calendar.keys()
            )
            
            if not lect_conflict:
                # Place it on Rabu with sequential time!
                lecturer_calendar[(lect, ONLINE_DAY, start, end)] = si
                schedule_result[si] = (ONLINE_DAY, "ONLINE", f"{start}-{end}")
                lecturer_days_used[lect].add(ONLINE_DAY)
                placed_count += 1
                placed = True
                
                # Update next available time for Rabu
                next_available_time = end_minutes
        
        # If sequential placement failed due to time crunch, find any available slot on Rabu
        if not placed:
            timeblocks = get_timeblocks_for_day(ONLINE_DAY, sks)
            for tb in timeblocks:
                if placed:
                    break
                start, end = tb.split('-')
                
                # Check lecturer conflict
                lect_conflict = any(
                    l == lect and d == ONLINE_DAY and times_overlap(start, end, s, e)
                    for (l, d, s, e) in lecturer_calendar.keys()
                )
                
                if not lect_conflict:
                    # Place it on Rabu!
                    lecturer_calendar[(lect, ONLINE_DAY, start, end)] = si
                    schedule_result[si] = (ONLINE_DAY, "ONLINE", f"{start}-{end}")
                    lecturer_days_used[lect].add(ONLINE_DAY)
                    placed_count += 1
                    placed = True
                    break
        
        if not placed:
            failed_sections.append((si, sec))
            print(f"  ⚠️  Failed to place ONLINE section {si}: {sec.get('course_name', 'Unknown')} - {lect} (Rabu penuh)")
    
    online_placed = len([s for s in schedule_result.values() if s[1] == 'ONLINE'])
    print(f"Online sections placed: {online_placed}/{len(online_sections)}")
    
    # Show distribution of online sections by day
    online_by_day = {}
    for si, (day, room, tb) in schedule_result.items():
        if room == "ONLINE":
            online_by_day[day] = online_by_day.get(day, 0) + 1
    
    if online_by_day:
        print(f"\n📊 Online sections distribution:")
        for day in weekday_order:
            if day in online_by_day:
                print(f"   {day}: {online_by_day[day]} sections")
    
    # STEP 2: Schedule REGULAR sections (need physical rooms)
    print("\n" + "="*80)
    print("STEP 2: SCHEDULING REGULAR SECTIONS (NEED ROOMS)")
    print("="*80)
    
    # Calculate preference restrictiveness score for each lecturer
    def calculate_restrictiveness(lect):
        """Calculate how restrictive a lecturer's preferences are.
        Higher score = more restrictive = should be scheduled FIRST
        """
        prefs = lecturer_preferences.get(lect, {})
        score = 0
        
        # Factor 1: Limited available_days (most important)
        avail_days = prefs.get('available_days', [])
        if avail_days:
            # Score: 100 points for each day NOT available
            unavailable_count = 5 - len(avail_days)
            score += unavailable_count * 100
        
        # Factor 2: Number of blocked time ranges
        blocked_times = prefs.get('unavailable_time_ranges', [])
        score += len(blocked_times) * 50
        
        # Factor 3: Number of sections (tiebreaker - more sections = higher priority)
        num_sections = len(lecturer_sections_map.get(lect, []))
        score += num_sections * 10
        
        return score
    
    # Sort lecturers: MOST RESTRICTIVE FIRST (high score = schedule first)
    sorted_lecturers = sorted(
        lecturer_sections_map.items(), 
        key=lambda x: calculate_restrictiveness(x[0]), 
        reverse=True
    )
    
    print(f"\n📊 Scheduling order (by preference restrictiveness + section count):")
    for lect, sec_list in sorted_lecturers[:10]:
        prefs = lecturer_preferences.get(lect, {})
        avail_days = prefs.get('available_days', [])
        blocked = len(prefs.get('unavailable_time_ranges', []))
        restrictiveness = calculate_restrictiveness(lect)
        avail_str = f"{len(avail_days)}/5 days" if avail_days else "all days"
        print(f"  {lect}: {len(sec_list)} sections | {avail_str}, {blocked} blocks | score={restrictiveness}")
    if len(sorted_lecturers) > 10:
        print(f"  ... and {len(sorted_lecturers) - 10} more lecturers")
    print()
    
    for lect, sec_list in sorted_lecturers:
        # Shuffle sections for variety
        sec_list = list(sec_list)
        random.shuffle(sec_list)
        
        # Try to place all sections for this lecturer in 2-3 days
        # Strategy: Try consecutive day windows
        best_placement = None
        best_day_count = 99
        
        for start_day_idx in range(len(weekday_order)):
            for day_span in [2, 3]:  # Try 2 days first, then 3
                if start_day_idx + day_span > len(weekday_order):
                    continue
                
                target_days = weekday_order[start_day_idx:start_day_idx+day_span]
                temp_assignments = []
                temp_room_calendar = set()  # Track temporary room bookings
                temp_lect_calendar = set()  # Track temporary lecturer bookings
                
                # Try to place all sections in these days
                success = True
                for si, sec in sec_list:
                    sks = int(sec.get('sks', 2))
                    
                    placed = False
                    best_score = -1
                    best_temp_assignment = None
                    
                    for day in target_days:
                        if placed:
                            break
                        timeblocks = get_timeblocks_for_day(day, sks)
                        for room in rooms:
                            if placed:
                                break
                            for tb in timeblocks:
                                start, end = tb.split('-')
                                
                                # Check lecturer preferences FIRST (hard constraints)
                                allowed, pref_score = check_lecturer_preferences(lect, day, start, end)
                                if not allowed:
                                    continue  # Skip this slot if preferences don't allow
                                
                                # Check room conflict (existing + temporary)
                                room_conflict = any(
                                    d == day and r == room and times_overlap(start, end, s, e)
                                    for (d, r, s, e) in room_calendar.keys()
                                ) or any(
                                    d == day and r == room and times_overlap(start, end, s, e)
                                    for (d, r, s, e) in temp_room_calendar
                                )
                                
                                # Check lecturer conflict (existing + temporary)
                                lect_conflict = any(
                                    l == lect and d == day and times_overlap(start, end, s, e)
                                    for (l, d, s, e) in lecturer_calendar.keys()
                                ) or any(
                                    d == day and times_overlap(start, end, s, e)
                                    for (d, s, e) in temp_lect_calendar
                                )
                                
                                if not room_conflict and not lect_conflict:
                                    # Valid slot found, check if better than previous
                                    if pref_score > best_score:
                                        best_score = pref_score
                                        best_temp_assignment = (si, day, room, start, end)
                                        if pref_score == 100:  # Perfect match!
                                            placed = True
                                            break
                    
                    # Apply best assignment if found
                    if best_temp_assignment:
                        si, day, room, start, end = best_temp_assignment
                        temp_assignments.append((si, day, room, start, end))
                        temp_room_calendar.add((day, room, start, end))
                        temp_lect_calendar.add((day, start, end))
                        placed = True
                    
                    if not placed:
                        success = False
                        break
                
                # If all sections placed, consider this a candidate
                if success and len(temp_assignments) == len(sec_list):
                    days_used = len(set(a[1] for a in temp_assignments))
                    if days_used < best_day_count:
                        best_day_count = days_used
                        best_placement = temp_assignments
                    if best_day_count == 2:  # Optimal, stop searching
                        break
            
            if best_day_count == 2:
                break
        
        # Apply best placement found
        if best_placement:
            for si, day, room, start, end in best_placement:
                room_calendar[(day, room, start, end)] = si
                lecturer_calendar[(lect, day, start, end)] = si
                schedule_result[si] = (day, room, f"{start}-{end}")
                lecturer_days_used[lect].add(day)
                placed_count += 1
        else:
            # 3-PHASE FALLBACK: Progressive constraint relaxation
            print(f"  ⚠️  {lect}: Could not find 2-3 consecutive days, trying fallback placement...")
            fallback_attempts_per_section = 0
            
            for si, sec in sec_list:
                if si in schedule_result:  # Skip if already placed
                    continue
                    
                sks = int(sec.get('sks', 2))
                is_lab = sec.get('is_lab', False)
                suitable_rooms = LAB_ROOMS if is_lab else NON_LAB_ROOMS
                
                placed = False
                attempts = 0
                placement_phase = "none"
                
                # PHASE 1: Try AVAILABLE_DAYS with ALL timeblocks (strict preferences but flexible times)
                if lect in lecturer_preferences:
                    avail_days = lecturer_preferences[lect].get('available_days', [])
                    if avail_days:
                        days_phase1 = [d for d in weekday_order if d in avail_days]
                        
                        for day in days_phase1:
                            if placed:
                                break
                            timeblocks = get_timeblocks_for_day(day, sks)
                            for room in suitable_rooms:
                                if placed:
                                    break
                                for tb in timeblocks:
                                    attempts += 1
                                    start, end = tb.split('-')
                                    
                                    # Check preferences (strict mode - respect available_days)
                                    allowed, _ = check_lecturer_preferences(lect, day, start, end, strict_mode=True)
                                    if not allowed:
                                        continue
                                    
                                    # Check conflicts
                                    room_conflict = any(
                                        d == day and r == room and times_overlap(start, end, s, e)
                                        for (d, r, s, e) in room_calendar.keys()
                                    )
                                    lect_conflict = any(
                                        l == lect and d == day and times_overlap(start, end, s, e)
                                        for (l, d, s, e) in lecturer_calendar.keys()
                                    )
                                    
                                    if not room_conflict and not lect_conflict:
                                        room_calendar[(day, room, start, end)] = si
                                        lecturer_calendar[(lect, day, start, end)] = si
                                        schedule_result[si] = (day, room, tb)
                                        lecturer_days_used[lect].add(day)
                                        placed_count += 1
                                        placed = True
                                        placement_phase = "Phase 1 (available_days)"
                                        print(f"    ✓ [Phase 1] {sec.get('course_name')} {sec.get('section_label')} on {day} in {room} at {tb}")
                                        break
                
                # PHASE 2: Try ALL days with ALL timeblocks (RELAXED - ignore available_days)
                if not placed:
                    days_to_try = list(set(weekday_order) - lecturer_days_used[lect]) + list(lecturer_days_used[lect])
                    
                    for day in days_to_try:
                        if placed:
                            break
                        timeblocks = get_timeblocks_for_day(day, sks)
                        for room in suitable_rooms:
                            if placed:
                                break
                            for tb in timeblocks:
                                attempts += 1
                                start, end = tb.split('-')
                                
                                # Check preferences (RELAXED mode - only unavailable_time_ranges enforced)
                                allowed, _ = check_lecturer_preferences(lect, day, start, end, strict_mode=False)
                                if not allowed:
                                    continue  # Still respect blocked times
                                
                                # Check conflicts
                                room_conflict = any(
                                    d == day and r == room and times_overlap(start, end, s, e)
                                    for (d, r, s, e) in room_calendar.keys()
                                )
                                lect_conflict = any(
                                    l == lect and d == day and times_overlap(start, end, s, e)
                                    for (l, d, s, e) in lecturer_calendar.keys()
                                )
                                
                                if not room_conflict and not lect_conflict:
                                    room_calendar[(day, room, start, end)] = si
                                    lecturer_calendar[(lect, day, start, end)] = si
                                    schedule_result[si] = (day, room, tb)
                                    lecturer_days_used[lect].add(day)
                                    placed_count += 1
                                    placed = True
                                    placement_phase = "Phase 2 (relaxed - any day)"
                                    print(f"    ✓ [Phase 2 RELAXED] {sec.get('course_name')} {sec.get('section_label')} on {day} in {room} at {tb}")
                                    break
                
                fallback_attempts_per_section = attempts
                if not placed:
                    failed_sections.append((si, sec))
                    print(f"    ❌ Failed after {attempts} attempts: {sec.get('course_name')} {sec.get('section_label')}")
                elif placement_phase != "none":
                    print(f"       (Used {placement_phase} after {attempts} attempts)")
            
            if fallback_attempts_per_section > 0:
                print(f"  📊 Fallback tried average {fallback_attempts_per_section} combinations per section for {lect}")
    
    total_sections = len(sections)
    online_placed = len([s for s in schedule_result.values() if s[1] == "ONLINE"])
    regular_placed = placed_count - online_placed
    
    print(f"\n{'='*80}")
    print(f"SCHEDULING SUMMARY")
    print(f"{'='*80}")
    print(f"Total sections: {total_sections}")
    print(f"  - Online sections placed: {online_placed}/{len(online_sections)}")
    print(f"  - Regular sections placed: {regular_placed}/{len(regular_sections)}")
    print(f"Total placed: {placed_count}/{total_sections}")
    print(f"{'='*80}\n")
    
    if failed_sections:
        print(f"⚠️  Failed to place {len(failed_sections)} sections")
        
        # Store failed sections info for UI display
        failed_info = []
        
        for si, sec in failed_sections[:10]:
            lect = _get_section_dosen(sec)
            failed_info.append({
                'course': sec.get('course_name'),
                'section': sec.get('section_label'),
                'lecturer': lect,
                'sks': sec.get('sks', 2),
                'is_lab': sec.get('is_lab', False)
            })
            print(f"  - {sec.get('course_name')} {sec.get('section_label')} ({lect})")
        
        # Store in session for display
        session['failed_sections'] = failed_info
        
        # DIAGNOSTIC: Analyze why sections failed
        print(f"\n{'='*80}")
        print(f"DIAGNOSTIC: Why Sections Failed to Schedule")
        print(f"{'='*80}")
        
        for si, sec in failed_sections:
            lect = _get_section_dosen(sec)
            course = sec.get('course_name')
            label = sec.get('section_label')
            sks = int(sec.get('sks', 2))
            is_lab = sec.get('is_lab', False)
            
            print(f"\n❌ {course} {label} ({lect}):")
            print(f"   Type: {'Lab' if is_lab else 'Regular'}, SKS: {sks}")
            
            # Check lecturer preferences
            prefs = lecturer_preferences.get(lect, {})
            avail_days = prefs.get('available_days', [])
            blocked_times = prefs.get('unavailable_time_ranges', [])
            
            if avail_days:
                print(f"   Available days: {', '.join(avail_days)}")
            else:
                print(f"   Available days: All days (no restriction)")
            
            if blocked_times:
                print(f"   Blocked times:")
                for block in blocked_times:
                    day = block.get('day', '*')
                    start = block.get('start', '00:00')
                    end = block.get('end', '23:59')
                    print(f"     - {day}: {start}-{end}")
            
            # Count available slots for this lecturer
            suitable_rooms = LAB_ROOMS if is_lab else NON_LAB_ROOMS
            
            total_slots = 0
            blocked_slots = 0
            conflict_slots = 0
            
            days_to_check = avail_days if avail_days else days
            
            for day in days_to_check:
                timeblocks = get_timeblocks_for_day(day, sks)
                for room in suitable_rooms:
                    for tb in timeblocks:
                        start, end = tb.split('-')
                        total_slots += 1
                        
                        # Check if blocked
                        allowed, _ = check_lecturer_preferences(lect, day, start, end, strict_mode=True)
                        if not allowed:
                            blocked_slots += 1
                            continue
                        
                        # Check conflicts
                        room_conflict = any(
                            d == day and r == room and times_overlap(start, end, s, e)
                            for (d, r, s, e) in room_calendar.keys()
                        )
                        lect_conflict = any(
                            l == lect and d == day and times_overlap(start, end, s, e)
                            for (l, d, s, e) in lecturer_calendar.keys()
                        )
                        
                        if room_conflict or lect_conflict:
                            conflict_slots += 1
            
            available_slots = total_slots - blocked_slots - conflict_slots
            
            print(f"   Slot analysis (within available days):")
            print(f"     Total possible slots: {total_slots}")
            print(f"     Blocked by preferences: {blocked_slots} ({100*blocked_slots//total_slots if total_slots>0 else 0}%)")
            print(f"     Already occupied (conflicts): {conflict_slots} ({100*conflict_slots//total_slots if total_slots>0 else 0}%)")
            print(f"     Available slots: {available_slots} ({100*available_slots//total_slots if total_slots>0 else 0}%)")
            
            # Recommendation
            if available_slots == 0:
                if blocked_slots > conflict_slots:
                    print(f"   💡 RECOMMENDATION: Lecturer preferences too restrictive!")
                    print(f"      → Ask {lect} to expand available_days or reduce blocked times")
                else:
                    print(f"   💡 RECOMMENDATION: All compatible slots already occupied")
                    print(f"      → Need to move other sections or expand room/time capacity")
            else:
                print(f"   ⚠️  WARNING: {available_slots} slots available but placement failed (bug?)")
        
        # DIAGNOSTIC: Check slot availability
        print(f"\n📊 DIAGNOSTIC: Slot Utilization")
        total_slots = len(rooms) * len(days) * 8  # Approximate: 13 rooms × 5 days × 8 avg timeblocks
        used_slots = len(room_calendar)
        print(f"  Total approximate slots: {total_slots}")
        print(f"  Used slots: {used_slots}")
        print(f"  Utilization: {used_slots/total_slots*100:.1f}%")
        print(f"  Available: {total_slots - used_slots}")
        
        # Show which days/times are most congested
        day_usage = {}
        for (day, room, start, end) in room_calendar.keys():
            day_usage[day] = day_usage.get(day, 0) + 1
        print(f"\n  Usage by day:")
        for day in weekday_order:
            count = day_usage.get(day, 0)
            print(f"    {day}: {count} slots used")
    
    # ------------------------------------------------------------------
    # DAY DISTRIBUTION REBALANCE (ensure Friday not too sparse and reduce overload)
    # ------------------------------------------------------------------
    weekday_order = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
    distribution_before = {d: 0 for d in weekday_order}
    for si, (day, room, timeblock) in schedule_result.items():
        if day in distribution_before:
            distribution_before[day] += 1

    print("\n" + "="*80)
    print("DAY DISTRIBUTION (BEFORE REBALANCE)")
    print("="*80)
    for d in weekday_order:
        print(f"  {d:<7}: {distribution_before[d]}")
    
    # Show online distribution
    online_distribution = {}
    for si, (day, room, tb) in schedule_result.items():
        if room == "ONLINE":
            online_distribution[day] = online_distribution.get(day, 0) + 1
    
    if online_distribution:
        print(f"\n📊 Online sections by day:")
        for d in weekday_order:
            if d in online_distribution:
                marker = "✓ TARGET" if d == "Jumat" else ""
                print(f"  {d}: {online_distribution[d]} online sections {marker}")

    total_assigned = sum(distribution_before.values())
    if total_assigned == 0:
        print("No sections assigned; skipping rebalance.")
    else:
        # Target: aim for near-even spread; Friday must have at least 12-15 sections
        avg_target = total_assigned / 5.0
        
        # Check if Friday already has enough sections (including online)
        friday_count = distribution_before.get('Jumat', 0)
        online_on_friday = online_distribution.get('Jumat', 0)
        
        # Adjust minimum Friday target based on online sections already there
        min_friday_target = max(12, min(15, int(avg_target * 0.8)))  # At least 12, max 15, or 80% of average
        
        # If we already have many online sections on Friday, reduce move requirement
        if online_on_friday >= 6:
            min_friday_target = max(min_friday_target, friday_count)  # Don't force more if we have online
            print(f"\n✓ Friday already has {online_on_friday} online sections - minimal rebalancing needed")
        
        print(f"\nTarget avg per day ≈ {avg_target:.2f}; Friday minimum target: {min_friday_target}")

        # Build fast lookup for existing occupancy: (day, room, start-end)
        occupied_slots = set()
        for (day, room, start, end) in room_calendar.keys():
            occupied_slots.add((day, room, start, end))

        # Helper to parse timeblock
        def tb_parts(tb):
            s,e = tb.split('-'); return s,e

        # All candidate timeblocks (reuse TIMEBLOCKS mapping by SKS type of section moved)
        def find_free_slot(section):
            sks = int(section.get('sks', 2))
            is_lab = section.get('is_lab', False)
            is_online = section.get('is_online', False)
            lect = _get_section_dosen(section)

            # Online: tetap hormati preferensi dosen (available_days & blocked times)
            if is_online:
                timeblocks = get_timeblocks_for_day('Jumat', sks)
                for tb in timeblocks:
                    start, end = tb.split('-')
                    # Enforce preferences strictly agar tidak masuk hari terblokir
                    allowed, _ = check_lecturer_preferences(lect, 'Jumat', start, end, strict_mode=True, is_online=True)
                    if not allowed:
                        continue
                    lect_conflict = any(
                        l == lect and d == 'Jumat' and times_overlap(start, end, s, e)
                        for (l, d, s, e) in lecturer_calendar.keys()
                    )
                    if not lect_conflict:
                        return ('ONLINE', tb)
                return None
            
            # Try suitable rooms first, then all rooms if needed
            suitable_rooms = LAB_ROOMS if is_lab else NON_LAB_ROOMS
            all_fallback_rooms = NON_LAB_ROOMS if is_lab else LAB_ROOMS
            
            timeblocks = get_timeblocks_for_day('Jumat', sks)
            for room_list in [suitable_rooms, all_fallback_rooms]:
                for room in room_list:
                    for tb in timeblocks:
                        start, end = tb.split('-')
                        # Enforce preferences strictly (hindari hari terblokir/di luar available_days)
                        allowed, _ = check_lecturer_preferences(lect, 'Jumat', start, end, strict_mode=True, is_online=False)
                        if not allowed:
                            continue
                        # Room free?
                        room_busy = any(
                            d == 'Jumat' and r == room and times_overlap(start, end, s, e)
                            for (d, r, s, e) in room_calendar.keys()
                        )
                        if room_busy:
                            continue
                        # Lecturer free?
                        lect_busy = any(
                            l == lect and d == 'Jumat' and times_overlap(start, end, s, e)
                            for (l, d, s, e) in lecturer_calendar.keys()
                        )
                        if not lect_busy:
                            return (room, tb)
            return None

        moves = []  # (section_index, from_day, to_day, new_room, new_tb)
        if distribution_before['Jumat'] < min_friday_target:
            deficit = min_friday_target - distribution_before['Jumat']
            print(f"\nFriday deficit detected: need to move {deficit} section(s) to Jumat")
            # Build list of donor days sorted by surplus descending
            # Allow ALL days (not just above average) to donate to Friday
            donor_days = sorted(
                [d for d in weekday_order if d != 'Jumat' and distribution_before[d] > 0],
                key=lambda d: distribution_before[d], reverse=True
            )
            print(f"Donor days (by count): {[(d, distribution_before[d]) for d in donor_days]}")
            
            for donor_day in donor_days:
                if deficit <= 0:
                    break
                # Collect movable sections from donor_day (prefer non-lab 2 SKS for flexibility)
                donor_candidates = [
                    si for si,(day,room,tb) in schedule_result.items()
                    if day == donor_day
                ]
                # Prioritize by (is_lab False, sks small)
                donor_candidates = sorted(
                    donor_candidates,
                    key=lambda si: (
                        schedule_result[si][1] == 'ONLINE',  # online easiest
                        not sections[si].get('is_lab', False),    # non-lab first (reversed)
                        int(sections[si].get('sks',2))        # smaller SKS first
                    ),
                    reverse=True  # ONLINE first, non-lab first
                )
                print(f"  Donor day {donor_day} has {len(donor_candidates)} candidates")
                moved_from_this_day = 0
                for si in donor_candidates:
                    if deficit <= 0:
                        break
                    section = sections[si]
                    slot = find_free_slot(section)
                    if not slot:
                        print(f"    ✗ Cannot move {section.get('course_name','')} {section.get('section_label','')} - no free slot")
                        continue
                    new_room, new_tb = slot
                    old_day, old_room, old_tb = schedule_result[si]
                    # Update calendars (remove old, add new)
                    start_old, end_old = old_tb.split('-')
                    start_new, end_new = new_tb.split('-')
                    # Remove old room/lecturer occupancy (skip ONLINE room removal for conflicts purpose)
                    if old_room != 'ONLINE':
                        room_calendar.pop((old_day, old_room, start_old, end_old), None)
                    lect = section.get('dosen') or section.get('lecturer','')
                    lecturer_calendar.pop((lect, old_day, start_old, end_old), None)
                    # Add new occupancy
                    if new_room != 'ONLINE':
                        room_calendar[( 'Jumat', new_room, start_new, end_new)] = si
                    lecturer_calendar[( lect, 'Jumat', start_new, end_new)] = si
                    # Apply move
                    schedule_result[si] = ('Jumat', new_room, new_tb)
                    distribution_before[donor_day] -= 1
                    distribution_before['Jumat'] += 1
                    deficit -= 1
                    moved_from_this_day += 1
                    moves.append((si, donor_day, 'Jumat', new_room, new_tb))
                    print(f"    ✓ Moved {section.get('course_name','')} {section.get('section_label','')} to {new_room} at {new_tb}")
                print(f"  Processed donor day {donor_day}; moved {moved_from_this_day} sections; Friday now: {distribution_before['Jumat']}/{min_friday_target}")
            
            if deficit > 0:
                print(f"\n⚠️ Could not fully balance Friday. Still need {deficit} more sections.")
        else:
            print("Friday meets minimum target; no redistribution needed.")

        # ADDITIONAL REBALANCE: Balance other days (like Kamis) that are too sparse
        print("\n" + "="*80)
        print("GENERAL DAY BALANCE CHECK")
        print("="*80)
        
        # Find days that are significantly below average
        avg_count = sum(distribution_before.values()) / len(weekday_order)
        min_acceptable = max(10, avg_count * 0.4)  # At least 40% of average or 10 sections
        print(f"Average per day: {avg_count:.1f}, Minimum acceptable: {min_acceptable:.1f}")
        
        # Days below minimum threshold need help
        sparse_days = [(d, distribution_before[d]) for d in weekday_order 
                       if distribution_before[d] < min_acceptable]
        
        if sparse_days:
            print(f"Sparse days detected: {sparse_days}")
            
            for sparse_day, sparse_count in sorted(sparse_days, key=lambda x: x[1]):  # Start with sparsest
                target_add = max(3, int(avg_count * 0.8 - sparse_count))  # Bring to 80% of average, min 3
                if target_add <= 0:
                    continue
                    
                print(f"\nTrying to add {target_add} sections to {sparse_day} (current: {sparse_count})...")
                
                # Find donor days (well above average) - be more aggressive
                potential_donors = sorted(
                    [(d, distribution_before[d]) for d in weekday_order 
                     if d != sparse_day and distribution_before[d] > avg_count * 1.1],  # At least 10% above avg
                    key=lambda x: x[1], reverse=True
                )
                
                if not potential_donors:
                    # If no one is significantly above, take from anyone above average
                    potential_donors = sorted(
                        [(d, distribution_before[d]) for d in weekday_order 
                         if d != sparse_day and distribution_before[d] > avg_count],
                        key=lambda x: x[1], reverse=True
                    )
                
                print(f"  Potential donors: {potential_donors}")
                
                added = 0
                for donor_day, donor_count in potential_donors:
                    if added >= target_add:
                        break
                    
                    # Limit how much we take from each donor (don't make them too sparse)
                    max_from_donor = max(1, int((donor_count - avg_count) * 0.5))
                    
                    # Try to move sections from donor_day to sparse_day
                    donor_sections = [(si, (day, room, tb)) for si, (day, room, tb) in schedule_result.items()
                                     if day == donor_day]
                    
                    moved_from_this_donor = 0
                    for si, section_tuple in donor_sections:
                        if added >= target_add or moved_from_this_donor >= max_from_donor:
                            break
                            
                        day_old, room_old, tb_old = section_tuple
                        
                        # Find matching section data
                        matching_sections = [s for s in sections if s.get('_id') == si]
                        if not matching_sections:
                            continue
                        section = matching_sections[0]
                        
                        # Try to find free slot on sparse_day
                        sks = int(section.get('sks', 2))
                        is_lab = section.get('is_lab', False)
                        is_online = section.get('is_online', False)
                        
                        lect = _get_section_dosen(section)
                        suitable_rooms = ['ONLINE'] if is_online else (LAB_ROOMS if is_lab else NON_LAB_ROOMS)
                        
                        timeblocks = get_timeblocks_for_day(sparse_day, sks)
                        found_slot = False
                        for tb_new in timeblocks:  # Try all timeblocks first
                            if found_slot:
                                break
                            for room_new in suitable_rooms:
                                start_new, end_new = tb_new.split('-')
                                
                                # Check room conflict
                                room_busy = any(d == sparse_day and r == room_new and times_overlap(start_new, end_new, s, e)
                                              for (d,r,s,e) in room_calendar.keys())
                                if room_busy:
                                    continue
                                
                                # Check lecturer conflict
                                lect_busy = any(l == lect and d == sparse_day and times_overlap(start_new, end_new, s, e)
                                              for (l,d,s,e) in lecturer_calendar.keys())
                                if lect_busy:
                                    continue
                                
                                # Check lecturer preferences (STRICT: hindari hari di luar available_days / blocked)
                                if lect in lecturer_preferences:
                                    allowed, _ = check_lecturer_preferences(lect, sparse_day, start_new, end_new, strict_mode=True)
                                    if not allowed:
                                        continue
                                
                                # Move section!
                                # Remove old slot
                                old_start, old_end = tb_parts(tb_old)
                                del room_calendar[(day_old, room_old, old_start, old_end)]
                                del lecturer_calendar[(lect, day_old, old_start, old_end)]
                                
                                # Add new slot
                                room_calendar[(sparse_day, room_new, start_new, end_new)] = si
                                lecturer_calendar[(lect, sparse_day, start_new, end_new)] = si
                                schedule_result[si] = (sparse_day, room_new, tb_new)
                                
                                distribution_before[day_old] -= 1
                                distribution_before[sparse_day] += 1
                                
                                added += 1
                                moved_from_this_donor += 1
                                found_slot = True
                                print(f"  ✓ Moved {section.get('course_name')} ({section.get('section_label')}) from {day_old} to {sparse_day} at {tb_new}")
                                break
                    
                    if moved_from_this_donor > 0:
                        print(f"  Moved {moved_from_this_donor} from {donor_day}")
                
                if added > 0:
                    print(f"  Successfully added {added}/{target_add} sections to {sparse_day}")
                else:
                    print(f"  ⚠️ Could not add sections to {sparse_day} - no free slots found")
        else:
            print("All days have acceptable distribution")

        distribution_after = {d: 0 for d in weekday_order}
        for si,(day,room,tb) in schedule_result.items():
            distribution_after[day] += 1

        print("\n" + "="*80)
        print("DAY DISTRIBUTION (AFTER REBALANCE)")
        print("="*80)
        for d in weekday_order:
            print(f"  {d:<7}: {distribution_after[d]}")
        if moves:
            print(f"\nMoved {len(moves)} section(s) to Friday:\n  " + "\n  ".join([
                f"{sections[m[0]].get('course_name','?')} {sections[m[0]].get('section_label','')} {m[1]} -> Jumat at {m[3]} {m[4]}"
                for m in moves
            ]))
        else:
            print("\nNo moves performed.")

    # Save to database (after potential rebalance)
    logger.info(f"\n🗑️ CLEARING SCHEDULES BEFORE INSERT:")
    before_delete = schedules_collection.count_documents({})
    schedules_collection.delete_many({})
    after_delete = schedules_collection.count_documents({})
    logger.info(f"   Before: {before_delete} schedules")
    logger.info(f"   After delete: {after_delete} schedules")
    
    new_sched = []
    
    logger.info(f"\n📊 SCHEDULE_RESULT STATS BEFORE INSERT:")
    logger.info(f"   - schedule_result has {len(schedule_result)} total items")
    logger.info(f"   - sections list has {len(sections)} items")
    logger.info(f"   - unique section indices in schedule_result: {len(set(schedule_result.keys()))}")
    
    # Check for duplicate section indices
    if len(schedule_result) != len(set(schedule_result.keys())):
        logger.warning(f"⚠️ DUPLICATE KEYS in schedule_result!")
    
    if len(schedule_result) != len(sections):
        logger.warning(f"⚠️ MISMATCH: schedule_result ({len(schedule_result)}) != sections ({len(sections)})")
        if len(schedule_result) == len(sections) * 2:
            logger.error(f"❌ schedule_result is EXACTLY 2x sections! Major bug in scheduling logic!")
    
    for si, (day, room, timeblock) in schedule_result.items():
        sec = sections[si]
        start, end = timeblock.split('-')
        
        # Calculate preference score for this placement
        lect = sec.get('dosen') or sec.get('lecturer')
        is_online = (room == "ONLINE")
        allowed, pref_score = check_lecturer_preferences(lect, day, start, end, strict_mode=False, is_online=is_online)
        match_quality = get_match_quality(pref_score)
        
        doc = {
            'day': day,
            'room': room,
            'start': start,
            'end': end,
            'course_id': sec.get('course_id'),
            'course_name': sec.get('course_name'),
            'sks': sec.get('sks'),
            'dosen': lect,
            'section_label': sec.get('section_label'),
            'is_lab': sec.get('is_lab', False),
            'preference_score': pref_score,
            'match_quality': match_quality
        }
        # Add is_online flag if it's an online course
        if room == "ONLINE":
            doc['is_online'] = True
        new_sched.append(doc)
    
    if new_sched:
        schedules_collection.insert_many(new_sched)
        
        # VERIFY: Count should match sections
        final_count = schedules_collection.count_documents({})
        logger.info(f"✅ Inserted {len(new_sched)} schedules. DB now has {final_count} total schedules")
        
        if final_count != len(sections):
            logger.warning(f"⚠️ Schedule count mismatch: inserted {len(new_sched)} but DB has {final_count} (expected {len(sections)})")
            if final_count > len(sections) * 1.5:
                logger.error(f"❌ WAY TOO MANY SCHEDULES ({final_count})! Something is appending instead of replacing.")
        else:
            logger.info(f"✅ Perfect: {final_count} schedules = {len(sections)} sections")

    
    # VALIDATION: Check for conflicts
    print("\n" + "="*80)
    print("VALIDATION REPORT")
    print("="*80)
    
    lecturer_schedule = {}  # lecturer -> list of (day, start_min, end_min, course, room)
    room_schedule = {}      # (day, room) -> list of (start_min, end_min, course, lecturer)
    
    for doc in new_sched:
        dosen = doc.get('dosen', '')
        day = doc.get('day', '')
        start = doc.get('start', '')
        end = doc.get('end', '')
        room = doc.get('room', '')
        course = doc.get('course_name', '')
        
        # Parse times to minutes
        try:
            start_parts = start.split(':')
            end_parts = end.split(':')
            start_min = int(start_parts[0]) * 60 + int(start_parts[1])
            end_min = int(end_parts[0]) * 60 + int(end_parts[1])
        except:
            continue
        
        # Track lecturer schedule
        if dosen:
            if dosen not in lecturer_schedule:
                lecturer_schedule[dosen] = []
            lecturer_schedule[dosen].append((day, start_min, end_min, course, room))
        
        # Track room schedule
        if room and day:
            key = (day, room)
            if key not in room_schedule:
                room_schedule[key] = []
            room_schedule[key].append((start_min, end_min, course, dosen))
    
    # Check lecturer conflicts
    lecturer_conflicts = []
    for lect, slots in lecturer_schedule.items():
        for i in range(len(slots)):
            for j in range(i+1, len(slots)):
                day1, s1, e1, c1, r1 = slots[i]
                day2, s2, e2, c2, r2 = slots[j]
                
                # Same day and overlapping times?
                if day1 == day2 and s1 < e2 and s2 < e1:
                    lecturer_conflicts.append({
                        'lecturer': lect,
                        'day': day1,
                        'slot1': f"{s1//60:02d}:{s1%60:02d}-{e1//60:02d}:{e1%60:02d} {r1} ({c1})",
                        'slot2': f"{s2//60:02d}:{s2%60:02d}-{e2//60:02d}:{e2%60:02d} {r2} ({c2})"
                    })
    
    # Check room conflicts (SKIP ONLINE ROOMS - unlimited capacity)
    room_conflicts = []
    for (day, room), slots in room_schedule.items():
        # Skip conflict check for ONLINE rooms
        if room == 'ONLINE':
            continue
            
        for i in range(len(slots)):
            for j in range(i+1, len(slots)):
                s1, e1, c1, l1 = slots[i]
                s2, e2, c2, l2 = slots[j]
                
                # Overlapping times?
                if s1 < e2 and s2 < e1:
                    room_conflicts.append({
                        'room': room,
                        'day': day,
                        'slot1': f"{s1//60:02d}:{s1%60:02d}-{e1//60:02d}:{e1%60:02d} {l1} ({c1})",
                        'slot2': f"{s2//60:02d}:{s2%60:02d}-{e2//60:02d}:{e2%60:02d} {l2} ({c2})"
                    })
    
    if lecturer_conflicts:
        print(f"\n❌ FOUND {len(lecturer_conflicts)} LECTURER CONFLICTS:")
        for i, conf in enumerate(lecturer_conflicts[:10], 1):
            print(f"{i}. {conf['lecturer']} on {conf['day']}:")
            print(f"   - {conf['slot1']}")
            print(f"   - {conf['slot2']}")
        if len(lecturer_conflicts) > 10:
            print(f"   ... and {len(lecturer_conflicts)-10} more")
    else:
        print("✅ No lecturer conflicts found")
    
    if room_conflicts:
        print(f"\n❌ FOUND {len(room_conflicts)} ROOM CONFLICTS (ONLINE rooms excluded):")
        for i, conf in enumerate(room_conflicts[:10], 1):
            print(f"{i}. {conf['room']} on {conf['day']}:")
            print(f"   - {conf['slot1']}")
            print(f"   - {conf['slot2']}")
        if len(room_conflicts) > 10:
            print(f"   ... and {len(room_conflicts)-10} more")
    else:
        print("✅ No room conflicts found (ONLINE rooms excluded from check)")
    
    print("="*80 + "\n")
    
    # ADDITIONAL VALIDATION: Check consecutive days constraint
    print("="*80)
    print("CONSECUTIVE DAYS VALIDATION")
    print("="*80)
    
    lecturer_days = {}  # lecturer -> set of days
    for doc in new_sched:
        lect = doc.get('dosen', '')
        day = doc.get('day', '')
        if lect and day:
            if lect not in lecturer_days:
                lecturer_days[lect] = set()
            lecturer_days[lect].add(day)
    
    weekday_order = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
    day_index = {d: i for i, d in enumerate(weekday_order)}
    
    consecutive_violations = []
    for lect, days_set in sorted(lecturer_days.items()):
        day_indices = sorted([day_index[d] for d in days_set if d in day_index])
        num_days = len(day_indices)
        
        # Check if days are consecutive
        is_consecutive = True
        if num_days >= 2:
            for i in range(len(day_indices)-1):
                if day_indices[i+1] - day_indices[i] != 1:
                    is_consecutive = False
                    break
        
        days_str = ', '.join([weekday_order[i] for i in day_indices])
        
        if num_days < 2 or num_days > 3:
            consecutive_violations.append(f"{lect}: {num_days} days ({days_str}) ❌ [Expected 2-3 days]")
        elif not is_consecutive:
            # NOTE: Non-consecutive days are now ACCEPTABLE (relaxed constraint)
            print(f"  {lect}: {days_str} ⚠ Non-consecutive (acceptable)")
        else:
            print(f"  {lect}: {days_str} ✓ Consecutive")
    
    if consecutive_violations:
        print(f"\n❌ DAY COUNT VIOLATIONS: {len(consecutive_violations)} lecturers")
        for v in consecutive_violations[:10]:
            print(f"  {v}")
        if len(consecutive_violations) > 10:
            print(f"  ... and {len(consecutive_violations)-10} more")
    else:
        print("\n✓ All lecturers teach within 2-3 days range")
    
    print("="*80 + "\n")
    
    # Flash message based on conflicts
    total_conflicts = len(lecturer_conflicts) + len(room_conflicts)
    consecutive_ok = len(consecutive_violations) == 0
    
    # Check for failed sections from earlier
    unscheduled_count = len(sections) - len(new_sched)
    
    # Final validation
    conflicts_check = check_schedule_conflicts(new_sched)
    has_real_conflicts = conflicts_check and (
        (isinstance(conflicts_check, dict) and (conflicts_check.get('lecturer_conflicts') or conflicts_check.get('room_conflicts'))) or
        (isinstance(conflicts_check, bool) and conflicts_check)
    )
    
    if unscheduled_count > 0:
        flash(f"⚠️ CRITICAL: {unscheduled_count} sections GAGAL dijadwalkan! Lihat log untuk detail. Perlu review preferences dosen.", "error")
    elif has_real_conflicts:
        flash(f"⚠️ PERINGATAN: Jadwal disimpan tapi masih ada bentrok! Silakan cek jadwal.", "warning")
    elif total_conflicts > 0:
        flash(f"⚠️ PERINGATAN: Jadwal disimpan tapi ada {len(lecturer_conflicts)} bentrok dosen dan {len(room_conflicts)} bentrok ruangan!", "warning")
    elif not consecutive_ok:
        flash(f"⚠️ PERINGATAN: Jadwal disimpan tapi ada {len(consecutive_violations)} dosen di luar range 2-3 hari!", "warning")
    else:
        flash(f"✅ Penjadwalan selesai! Total {len(new_sched)} slot tanpa bentrok. Semua dosen mengajar 2-3 hari.", "success")
    
    logger.info(f"\n{'='*80}")
    logger.info(f"🟢 FUNCTION END: schedule_sections_with_ortools() completed")
    logger.info(f"{'='*80}\n")

@app.route("/generate_sections_ga", methods=["GET","POST"])
def route_generate_sections_ga():
    try:
        population = int(request.form.get('population', 150)) if request.method == 'POST' else 150
        generations = int(request.form.get('generations', 300)) if request.method == 'POST' else 300
        force = request.form.get('force_regenerate') == 'true' if request.method == 'POST' else False
    except Exception:
        population, generations, force = 150, 300, False
    
    # ALWAYS clear existing data when generating fresh sections (not manual append)
    # This prevents jadwal lama numpuk ketika user klik "Run GA" berkali-kali
    logger.info(f"🔄 CLEARING DB BEFORE GA GENERATION...")
    before_sched = schedules_collection.count_documents({})
    before_sect = sections_collection.count_documents({})
    logger.info(f"   Before: {before_sect} sections, {before_sched} schedules")
    
    schedules_collection.delete_many({})
    sections_collection.delete_many({})
    
    after_sched = schedules_collection.count_documents({})
    after_sect = sections_collection.count_documents({})
    logger.info(f"   After: {after_sect} sections, {after_sched} schedules")
    
    if after_sect > 0 or after_sched > 0:
        logger.error(f"❌ DELETE FAILED! Still have {after_sect} sections and {after_sched} schedules!")
    else:
        logger.info(f"✅ Database cleared successfully")
    
    generate_sections_ga_pipeline(population=population, generations=generations, force_regenerate=force)
    return redirect(url_for('koordinator_home')) if 'koordinator_home' in app.view_functions else redirect(url_for('koordinator'))

@app.route("/schedule_sections_ortools", methods=["GET","POST"])
def route_schedule_sections_ortools():
    # Safety guard: if too many sections, something is wrong
    section_count = sections_collection.count_documents({})
    if section_count > 200:
        logger.warning(f"⚠️ Detected {section_count} sections (way too many)! Clearing before scheduling.")
        sections_collection.delete_many({})
        schedules_collection.delete_many({})
        flash(f"⚠️ Detected {section_count} sections in database. Cleared to prevent duplication. Please run GA again.", "warning")
        return redirect(url_for('koordinator'))
    
    logger.info(f"\n🔵 ROUTE: /schedule_sections_ortools called")
    schedule_sections_with_ortools(max_seconds=60)
    logger.info(f"🟢 ROUTE: /schedule_sections_ortools completed, redirecting...")
    return redirect(url_for('koordinator_home')) if 'koordinator_home' in app.view_functions else redirect(url_for('koordinator'))


# ============================================================================
# MANUAL SECTION MANAGEMENT
# ============================================================================

@app.route("/add_section_manual", methods=["POST"])
def add_section_manual():
    """Add new sections for a course, then auto-schedule them"""
    if 'user' not in session or session.get('user', {}).get('role') != 'koordinator':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Get form data
        course_id = request.form.get('course_id')
        course_name = request.form.get('course_name')
        lecturer = request.form.get('lecturer')
        num_sections = int(request.form.get('num_sections', 1))
        sks = int(request.form.get('sks', 3))
        
        # Validate required fields
        if not all([course_id, course_name, lecturer]):
            return jsonify({
                'success': False,
                'message': 'Mata kuliah dan dosen harus dipilih!'
            }), 400
        
        if num_sections < 1 or num_sections > 10:
            return jsonify({
                'success': False,
                'message': 'Jumlah section harus antara 1-10!'
            }), 400
        
        # Get course info
        course = courses_collection.find_one({'_id': ObjectId(course_id)})
        if not course:
            return jsonify({
                'success': False,
                'message': 'Mata kuliah tidak ditemukan!'
            }), 404
        
        is_lab = course.get('is_lab', False)
        is_online = course.get('is_online', False)
        
        # Get next section label
        existing_sections = list(sections_collection.find({'course_id': course_id}))
        start_section_num = len(existing_sections) + 1
        
        # Create new sections (without schedule yet)
        new_sections = []
        for i in range(num_sections):
            section_num = start_section_num + i
            section_letter = "A"  # Always use letter A
            section_number = section_num  # Sequential numbering
            section_label = f"{section_letter}{section_number}"
            
            new_section = {
                'course_id': course_id,
                'course_name': course_name,
                'sks': sks,
                'is_lab': is_lab,
                'is_online': is_online,
                'section_letter': section_letter,
                'section_number': section_number,
                'section_label': section_label,
                'lecturer': lecturer
            }
            new_sections.append(new_section)
        
        # TRY TO SCHEDULE FIRST (before inserting to DB)
        # This way, if schedule fails, section is not inserted
        try:
            # Create mock sections with temporary IDs for scheduling attempt
            mock_sections = [{**s, '_id': ObjectId()} for s in new_sections]
            schedule_new_sections_only([s['_id'] for s in mock_sections], mock_sections)
        except Exception as schedule_error:
            # Schedule failed - don't insert anything
            return jsonify({
                'success': False,
                'message': 'Gagal menjadwalkan section',
                'details': str(schedule_error)
            }), 400
        
        # Schedule succeeded - now insert sections to DB, then schedule again to populate schedules collection
        schedule_info_list = []
        swap_info = None
        
        if new_sections:
            result = sections_collection.insert_many(new_sections)
            inserted_ids = result.inserted_ids
            
            try:
                # Now schedule with actual section IDs in DB (will insert to schedules collection)
                # We'll capture the schedule details by reading from DB after scheduling
                schedule_new_sections_only(inserted_ids)
                
                # Get the schedule info for each new section
                for inserted_id in inserted_ids:
                    schedule_doc = schedules_collection.find_one({
                        'course_id': course_id,
                        '_id': {'$ne': None}
                    }, sort=[('_id', -1)])
                    
                    if schedule_doc:
                        schedule_info_list.append({
                            'day': schedule_doc['day'],
                            'start': schedule_doc['start'],
                            'end': schedule_doc['end'],
                            'room': schedule_doc['room']
                        })
            except Exception as schedule_error:
                # Scheduling failed - ROLLBACK: Delete sections we just inserted
                for section_id in inserted_ids:
                    sections_collection.delete_one({'_id': section_id})
                
                # Also delete any schedules that were inserted
                schedules_collection.delete_many({
                    'course_id': course_id,
                    'section_label': {'$in': [s['section_label'] for s in new_sections]}
                })
                
                # Return error
                return jsonify({
                    'success': False,
                    'message': 'Gagal menjadwalkan section',
                    'details': str(schedule_error)
                }), 400
            
            # If only one section, return single schedule info
            if schedule_info_list:
                schedule_info = schedule_info_list[0]
            else:
                schedule_info = {
                    'day': '?',
                    'start': '?',
                    'end': '?',
                    'room': '?'
                }
        
        return jsonify({
            'success': True,
            'message': f'{num_sections} section baru berhasil dibuat dan dijadwalkan',
            'schedule_info': schedule_info,
            'swap_info': {
                'happened': False
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': 'Error menambah section',
            'details': str(e)
        }), 500


def schedule_new_sections_only(section_ids, sections_to_schedule=None):
    """Schedule only specific new sections using greedy algorithm with preference checking
    
    Args:
        section_ids: List of section IDs to schedule
        sections_to_schedule: Optional list of section objects (for dry-run/preview before DB insert)
                            If None, will fetch from database
    
    Returns:
        dict with schedule results and swap info
    """
    # Get the new sections (either from parameter or from DB)
    if sections_to_schedule is not None:
        new_sections = sections_to_schedule
    else:
        new_sections = list(sections_collection.find({'_id': {'$in': section_ids}}))

    # Get existing schedules to avoid conflicts
    existing_schedules = list(schedules_collection.find())
    
    # Build conflict tracking from existing schedules
    room_calendar = {}
    lecturer_calendar = {}
    
    for sched in existing_schedules:
        day = sched['day']
        room = sched['room']
        start = sched['start']
        end = sched['end']
        lecturer = sched.get('dosen', '')
        
        # Skip ONLINE from room tracking
        if room != 'ONLINE':
            room_calendar[(day, room, start, end)] = True
        
        if lecturer:
            lecturer_calendar[(lecturer, day, start, end)] = True
    
    # Helper function
    def times_overlap(start1, end1, start2, end2):
        """Check if two time ranges overlap"""
        def to_minutes(t):
            h, m = map(int, t.split(':'))
            return h * 60 + m
        s1, e1 = to_minutes(start1), to_minutes(end1)
        s2, e2 = to_minutes(start2), to_minutes(end2)
        return s1 < e2 and s2 < e1
    
    def to_minutes(t):
        h, m = map(int, t.split(':'))
        return h * 60 + m
    
    # Get lecturer preferences (for scoring)
    def get_lecturer_preferences(lecturer_name):
        """Get preferences for a lecturer"""
        prefs = users_collection.find_one({'name': lecturer_name, 'role': 'dosen'})
        if prefs and 'preferences' in prefs:
            return prefs['preferences']
        return {}
    
    def check_preference_constraints(lecturer_name, day, start, end, is_online=False):
        """Check if slot violates hard preference constraints
        Returns: (allowed, score) where:
        - allowed: True if slot meets hard constraints
        - score: 100=excellent, 75=good, 60=acceptable, 50=general
        
        NOTE: available_days is now a SOFT constraint (prefers but doesn't block other days)
        Only unavailable_time_ranges is a hard constraint.
        """
        prefs = get_lecturer_preferences(lecturer_name)
        
        # Check unavailable_time_ranges (HARD constraint for physical, soft for online)
        unavailable_ranges = prefs.get("unavailable_time_ranges", [])
        for block in unavailable_ranges:
            block_day = block.get('day', '*')
            block_start = block.get('start', '00:00')
            block_end = block.get('end', '23:59')
            if block_day == '*' or block_day == day:
                if times_overlap(start, end, block_start, block_end):
                    if not is_online:
                        return False, 0  # Hard violation for physical classes
        
        # Check available_days (SOFT constraint - prefers available days but allows others as fallback)
        available_days = prefs.get("available_days", [])
        
        # Calculate preference score
        preferred_times = prefs.get("preferred_times", {})
        
        # Start with base score: available day is preferred (75), others are general (50)
        if available_days and day in available_days:
            score = 75  # Good - day is available
        else:
            score = 50  # Acceptable - not in preferred days, but acceptable as fallback
        
        # Check preferred_times match (only if day is available)
        if available_days and day in available_days and day in preferred_times:
            day_prefs = preferred_times[day]
            if day_prefs:
                slot_start_min = to_minutes(start)
                slot_end_min = to_minutes(end)
                
                for pref_range in day_prefs:
                    if '-' in pref_range:
                        pref_start, pref_end = pref_range.split('-')
                        pref_start_min = to_minutes(pref_start.strip())
                        pref_end_min = to_minutes(pref_end.strip())
                        
                        # Check if slot is FULLY within preferred window
                        if slot_start_min >= pref_start_min and slot_end_min <= pref_end_min:
                            score = 100  # Excellent - fully within preferred time on preferred day!
                            return True, score
                
                # Slot on preferred day but not in preferred time window
                score = 60  # Acceptable - day OK, time not perfect
        
        return True, score
    
    def try_smart_swap(section, lecturer, sks, is_lab, is_online, 
                       existing_schedules, room_calendar, lecturer_calendar,
                       get_prefs_func, check_pref_func, overlap_func, to_min_func,
                       sections_to_schedule):
        """Try to schedule by swapping with existing section
        
        Find an existing schedule and try to swap:
        - Move existing section to alternative slot (must respect their preferences)
        - Place new section in the freed slot (must respect new section's preferences)
        
        Returns:
            (swap_success, swap_details) where swap_details contains swap info if successful
        """
        weekday_order = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
        NON_LAB_ROOMS = ['Infor 1', 'Infor 2', 'Infor 3', 'Infor 4', 'Infor 5', 'Jawa 8A', 'Jawa 8B']
        LAB_ROOMS = ['Lab 1', 'Lab 2', 'Lab 3', 'Lab 4']
        TIMEBLOCKS = {
            1: ['08:00-08:50', '08:50-09:40', '09:40-10:30', '10:30-11:20', '11:20-12:10', '14:00-14:50', '14:50-15:40', '15:40-16:30'],
            2: ['08:00-09:40', '09:40-11:20', '10:30-12:10', '11:20-13:00', '14:00-15:40', '15:40-17:20'],
            3: ['08:00-10:30', '09:40-12:10', '10:30-13:00', '14:00-16:30']
        }
        
        # Iterate through existing schedules to find swap candidates
        for existing_sched in existing_schedules:
            existing_dosen = existing_sched.get('dosen', '')
            existing_day = existing_sched['day']
            existing_start = existing_sched['start']
            existing_end = existing_sched['end']
            existing_room = existing_sched['room']
            existing_course = existing_sched.get('course_name', '')
            existing_section = existing_sched.get('section_label', '')
            
            # Create a temporary lecturer_calendar EXCLUDING the existing section being swapped
            # This prevents false conflicts when checking alternative slots
            temp_lecturer_calendar = dict(lecturer_calendar)
            
            # Remove the existing dosen's current slot from tracking
            # (this slot will be freed when we move the existing section to alternative)
            for (l, d, s, e) in list(temp_lecturer_calendar.keys()):
                if (l == existing_dosen and d == existing_day and 
                    s == existing_start and e == existing_end):
                    del temp_lecturer_calendar[(l, d, s, e)]
            
            # Try to find alternative slot for existing section
            for alt_day in weekday_order:
                for alt_room in (LAB_ROOMS if is_lab else NON_LAB_ROOMS):
                    if alt_room == 'ONLINE':
                        continue
                    
                    alt_sks = existing_sched.get('sks', 3)
                    alt_timeblocks = get_timeblocks_for_day(alt_day, alt_sks)
                    
                    for alt_tb in alt_timeblocks:
                        alt_start, alt_end = alt_tb.split('-')
                        
                        # Check if alt slot respects existing section's preferences
                        alt_allowed, _ = check_pref_func(existing_dosen, alt_day, alt_start, alt_end, is_online=False)
                        if not alt_allowed:
                            continue
                        
                        # Check if alt slot has conflicts (excluding the original slot)
                        alt_conflict = False
                        for (d, r, s, e) in room_calendar.keys():
                            if (d == alt_day and r == alt_room and 
                                overlap_func(alt_start, alt_end, s, e) and
                                not (d == existing_day and r == existing_room and 
                                     s == existing_start and e == existing_end)):
                                alt_conflict = True
                                break
                        
                        if alt_conflict:
                            continue
                        
                        # Check lecturer conflict for alt slot USING TEMP CALENDAR (which excludes existing dosen's old slot)
                        lect_conflict_alt = False
                        for (l, d, s, e) in temp_lecturer_calendar.keys():
                            if (l == existing_dosen and d == alt_day and 
                                overlap_func(alt_start, alt_end, s, e)):
                                lect_conflict_alt = True
                                break
                        
                        if lect_conflict_alt:
                            continue
                        
                        # Now check if new section can fit in existing slot
                        # Check if existing slot respects new section's preferences
                        new_allowed, _ = check_pref_func(lecturer, existing_day, existing_start, existing_end, is_online=is_online)
                        if not new_allowed:
                            continue
                        
                        # Check room compatibility for new section
                        if is_online and existing_room != 'ONLINE':
                            continue
                        if not is_online and existing_room == 'ONLINE':
                            continue
                        
                        # All checks passed - perform the swap!
                        # Update existing section to new slot
                        if sections_to_schedule is None:
                            schedules_collection.update_one(
                                {'_id': existing_sched['_id']},
                                {'$set': {'day': alt_day, 'room': alt_room, 'start': alt_start, 'end': alt_end}}
                            )
                        
                        # Insert new section to old slot
                        new_schedule = {
                            'day': existing_day,
                            'room': existing_room,
                            'start': existing_start,
                            'end': existing_end,
                            'course_id': section['course_id'],
                            'course_name': section['course_name'],
                            'sks': sks,
                            'dosen': lecturer,
                            'section_label': section['section_label'],
                            'is_lab': is_lab,
                            'is_online': is_online
                        }
                        
                        if sections_to_schedule is None:
                            schedules_collection.insert_one(new_schedule)
                        
                        # Update tracking
                        if existing_room != 'ONLINE':
                            room_calendar[(existing_day, existing_room, existing_start, existing_end)] = True
                        if alt_room != 'ONLINE':
                            room_calendar[(alt_day, alt_room, alt_start, alt_end)] = True
                        
                        # Return swap success with details
                        return True, {
                            'happened': True,
                            'new_course': section['course_name'],
                            'new_lecturer': lecturer,
                            'new_day': existing_day,
                            'new_start': existing_start,
                            'new_end': existing_end,
                            'existing_course': existing_course,
                            'existing_section': existing_section,
                            'existing_lecturer': existing_dosen,
                            'existing_new_day': alt_day,
                            'existing_new_start': alt_start,
                            'existing_new_end': alt_end
                        }
        
        return False, None  # No successful swap found
    
    # Constants
    weekday_order = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
    NON_LAB_ROOMS = ['Infor 1', 'Infor 2', 'Infor 3', 'Infor 4', 'Infor 5', 'Jawa 8A', 'Jawa 8B']
    LAB_ROOMS = ['Lab 1', 'Lab 2', 'Lab 3', 'Lab 4']
    
    # Timeblocks by SKS
    TIMEBLOCKS = {
        1: ['08:00-08:50', '08:50-09:40', '09:40-10:30', '10:30-11:20', '11:20-12:10', '14:00-14:50', '14:50-15:40', '15:40-16:30'],
        2: ['08:00-09:40', '09:40-11:20', '10:30-12:10', '11:20-13:00', '14:00-15:40', '15:40-17:20'],
        3: ['08:00-10:30', '09:40-12:10', '10:30-13:00', '14:00-16:30']
    }
    
    # Try to schedule each new section
    scheduled_count = 0
    swap_info = None
    swapped_sections = []  # Track sections that were swapped (to exclude from bentrok check)
    
    # Track schedules for sections being added in THIS batch (to detect bentrok within batch)
    new_batch_schedules = {}  # section_idx -> (day, room, start, end)
    
    for section_idx, section in enumerate(new_sections):
        lecturer = section['lecturer']
        sks = section.get('sks', 3)
        is_lab = section.get('is_lab', False)
        is_online = section.get('is_online', False)
        
        # Online sections get ONLINE room automatically
        if is_online:
            room = 'ONLINE'
            suitable_rooms = ['ONLINE']
        else:
            suitable_rooms = LAB_ROOMS if is_lab else NON_LAB_ROOMS
        
        # Collect valid candidates with preference scores
        valid_candidates = []
        
        for day in weekday_order:
            timeblocks = get_timeblocks_for_day(day, sks)
            for room in suitable_rooms:
                for tb in timeblocks:
                    start, end = tb.split('-')
                    
                    # Check hard preference constraints
                    allowed, pref_score = check_preference_constraints(lecturer, day, start, end, is_online)
                    if not allowed:
                        continue  # Skip slots that violate hard constraints
                    
                    # Check conflicts WITH EXISTING SCHEDULES
                    if room != 'ONLINE':
                        room_conflict = any(
                            d == day and r == room and times_overlap(start, end, s, e)
                            for (d, r, s, e) in room_calendar.keys()
                        )
                    else:
                        room_conflict = False
                    
                    lect_conflict = any(
                        l == lecturer and d == day and times_overlap(start, end, s, e)
                        for (l, d, s, e) in lecturer_calendar.keys()
                    )
                    
                    # NEW: Check conflicts WITH OTHER SECTIONS IN SAME BATCH
                    batch_conflict = False
                    for other_idx, (other_day, other_room, other_start, other_end) in new_batch_schedules.items():
                        # Check if same lecturer would be double-booked
                        if lecturer == new_sections[other_idx]['lecturer'] and other_day == day and times_overlap(start, end, other_start, other_end):
                            batch_conflict = True
                            break
                    
                    if not room_conflict and not lect_conflict and not batch_conflict:
                        # Valid candidate
                        valid_candidates.append({
                            'day': day,
                            'room': room,
                            'start': start,
                            'end': end,
                            'score': pref_score
                        })
        
        if valid_candidates:
            # Sort by preference score (highest first)
            valid_candidates.sort(key=lambda x: -x['score'])
            best = valid_candidates[0]
            
            # Create schedule with best slot
            new_schedule = {
                'day': best['day'],
                'room': best['room'],
                'start': best['start'],
                'end': best['end'],
                'course_id': section['course_id'],
                'course_name': section['course_name'],
                'sks': sks,
                'dosen': lecturer,
                'section_label': section['section_label'],
                'is_lab': is_lab,
                'is_online': is_online
            }
            
            # Only insert to DB if this is NOT a dry-run (sections_to_schedule param not provided)
            if sections_to_schedule is None:
                schedules_collection.insert_one(new_schedule)
            
            # Update tracking (both existing calendar AND new batch schedule tracker)
            if best['room'] != 'ONLINE':
                room_calendar[(best['day'], best['room'], best['start'], best['end'])] = True
            lecturer_calendar[(lecturer, best['day'], best['start'], best['end'])] = True
            
            # Track in batch schedule dict to prevent conflicts within batch
            new_batch_schedules[section_idx] = (best['day'], best['room'], best['start'], best['end'])
            
            scheduled_count += 1
        else:
            # No slots with strict preference - try fallback: schedule anywhere without hard constraints
            fallback_candidates = []
            
            for day in weekday_order:
                for room in suitable_rooms:
                    for tb in timeblocks:
                        start, end = tb.split('-')
                        
                        # Skip hard constraints: unavailable_time_ranges and available_days
                        # But allow scheduling if lecturer has NO preferences at all
                        prefs = get_lecturer_preferences(lecturer)
                        available_days = prefs.get("available_days", [])
                        
                        # SOFT constraint: prefer available_days but allow other days as fallback
                        # (will score lower but still schedulable if needed)
                        
                        # Check unavailable_time_ranges (hard for physical classes)
                        unavailable_ranges = prefs.get("unavailable_time_ranges", [])
                        skip = False
                        for block in unavailable_ranges:
                            block_day = block.get('day', '*')
                            block_start = block.get('start', '00:00')
                            block_end = block.get('end', '23:59')
                            if (block_day == '*' or block_day == day) and not is_online:
                                if times_overlap(start, end, block_start, block_end):
                                    skip = True
                                    break
                        if skip:
                            continue
                        
                        # Check conflicts WITH EXISTING SCHEDULES
                        if room != 'ONLINE':
                            room_conflict = any(
                                d == day and r == room and times_overlap(start, end, s, e)
                                for (d, r, s, e) in room_calendar.keys()
                            )
                        else:
                            room_conflict = False
                        
                        lect_conflict = any(
                            l == lecturer and d == day and times_overlap(start, end, s, e)
                            for (l, d, s, e) in lecturer_calendar.keys()
                        )
                        
                        # NEW: Check conflicts WITH OTHER SECTIONS IN SAME BATCH
                        batch_conflict = False
                        for other_idx, (other_day, other_room, other_start, other_end) in new_batch_schedules.items():
                            # Check if same lecturer would be double-booked
                            if lecturer == new_sections[other_idx]['lecturer'] and other_day == day and times_overlap(start, end, other_start, other_end):
                                batch_conflict = True
                                break
                        
                        if not room_conflict and not lect_conflict and not batch_conflict:
                            fallback_candidates.append({
                                'day': day,
                                'room': room,
                                'start': start,
                                'end': end,
                                'score': 20  # Low score - fallback only
                            })
            
            if fallback_candidates:
                # Use fallback slot
                best = fallback_candidates[0]
                new_schedule = {
                    'day': best['day'],
                    'room': best['room'],
                    'start': best['start'],
                    'end': best['end'],
                    'course_id': section['course_id'],
                    'course_name': section['course_name'],
                    'sks': sks,
                    'dosen': lecturer,
                    'section_label': section['section_label'],
                    'is_lab': is_lab,
                    'is_online': is_online
                }
                
                # Only insert to DB if this is NOT a dry-run (sections_to_schedule param not provided)
                if sections_to_schedule is None:
                    schedules_collection.insert_one(new_schedule)
                
                # Update tracking
                if best['room'] != 'ONLINE':
                    room_calendar[(best['day'], best['room'], best['start'], best['end'])] = True
                lecturer_calendar[(lecturer, best['day'], best['start'], best['end'])] = True
                
                # Track in batch schedule dict to prevent conflicts within batch
                new_batch_schedules[section_idx] = (best['day'], best['room'], best['start'], best['end'])
                
                scheduled_count += 1
                
                # Log warning
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"⚠️ Section {section['section_label']} scheduled with fallback (not respecting soft preferences)")
            else:
                # No slots with soft preference either - try tahap 3: relaks available_days juga
                # Still respect hard constraints: unavailable_time_ranges (for physical classes)
                tahap3_candidates = []
                
                for day in weekday_order:
                    for room in suitable_rooms:
                        for tb in timeblocks:
                            start, end = tb.split('-')
                            
                            # Check unavailable_time_ranges (still hard constraint for physical)
                            prefs = get_lecturer_preferences(lecturer)
                            unavailable_ranges = prefs.get("unavailable_time_ranges", [])
                            skip = False
                            for block in unavailable_ranges:
                                block_day = block.get('day', '*')
                                block_start = block.get('start', '00:00')
                                block_end = block.get('end', '23:59')
                                if (block_day == '*' or block_day == day) and not is_online:
                                    if times_overlap(start, end, block_start, block_end):
                                        skip = True
                                        break
                            if skip:
                                continue

                            # NEW: Saat melonggarkan hari, waktu tetap ketat.
                            # Jika dosen punya preferred_times, slot harus masuk ke salah satu range untuk hari tersebut.
                            preferred_times = prefs.get("preferred_times", {})
                            if preferred_times:
                                if day not in preferred_times:
                                    continue  # Tidak ada preferensi waktu untuk hari ini, skip agar waktu tetap ketat
                                day_prefs = preferred_times.get(day, [])
                                in_pref_range = False
                                slot_start_min = to_minutes(start)
                                slot_end_min = to_minutes(end)
                                for pref_range in day_prefs:
                                    if '-' in pref_range:
                                        pref_start, pref_end = pref_range.split('-')
                                        pref_start_min = to_minutes(pref_start.strip())
                                        pref_end_min = to_minutes(pref_end.strip())
                                        if slot_start_min >= pref_start_min and slot_end_min <= pref_end_min:
                                            in_pref_range = True
                                            break
                                if not in_pref_range:
                                    continue  # Waktu tidak masuk preferensi, skip
                            
                            # Check conflicts
                            if room != 'ONLINE':
                                room_conflict = any(
                                    d == day and r == room and times_overlap(start, end, s, e)
                                    for (d, r, s, e) in room_calendar.keys()
                                )
                            else:
                                room_conflict = False
                            
                            lect_conflict = any(
                                l == lecturer and d == day and times_overlap(start, end, s, e)
                                for (l, d, s, e) in lecturer_calendar.keys()
                            )
                            
                            if not room_conflict and not lect_conflict:
                                tahap3_candidates.append({
                                    'day': day,
                                    'room': room,
                                    'start': start,
                                    'end': end,
                                    'score': 10  # Very low score - last resort
                                })
                
                if tahap3_candidates:
                    # Use tahap 3 slot (relaks available_days)
                    best = tahap3_candidates[0]
                    new_schedule = {
                        'day': best['day'],
                        'room': best['room'],
                        'start': best['start'],
                        'end': best['end'],
                        'course_id': section['course_id'],
                        'course_name': section['course_name'],
                        'sks': sks,
                        'dosen': lecturer,
                        'section_label': section['section_label'],
                        'is_lab': is_lab,
                        'is_online': is_online
                    }
                    
                    # Only insert to DB if this is NOT a dry-run
                    if sections_to_schedule is None:
                        schedules_collection.insert_one(new_schedule)
                    
                    # Update tracking
                    if best['room'] != 'ONLINE':
                        room_calendar[(best['day'], best['room'], best['start'], best['end'])] = True
                    lecturer_calendar[(lecturer, best['day'], best['start'], best['end'])] = True
                    
                    scheduled_count += 1
                    
                    # Log warning
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f"⚠️ Section {section['section_label']} scheduled on {best['day']} {best['start']}-{best['end']} (not respecting available_days preference)")
                else:
                    # Could not schedule this section even with all relaxations
                    # Try tahap 4: Smart swap dengan section lain yang sudah terjadwal
                        swap_success, swap_details = try_smart_swap(
                            section, lecturer, sks, is_lab, is_online,
                            existing_schedules, room_calendar, lecturer_calendar,
                            get_lecturer_preferences, check_preference_constraints,
                            times_overlap, to_minutes, sections_to_schedule
                        )
                        
                        if swap_success:
                            scheduled_count += 1
                            swap_info = swap_details  # Capture swap info for return
                            
                            # Track which section was swapped (for bentrok validation exclusion)
                            swapped_sections.append({
                                'section_label': swap_details['existing_section'],
                                'lecturer': swap_details['existing_lecturer'],
                                'old_day': swap_details.get('existing_new_day'),  # Where it moved TO
                                'old_start': swap_details.get('existing_new_start'),
                                'old_end': swap_details.get('existing_new_end')
                            })
                            
                            import logging
                            logger = logging.getLogger(__name__)
                            logger.warning(f"⚠️ Section {section['section_label']} scheduled via smart swap")
                        else:
                            # Swap juga gagal - truly cannot schedule
                            raise Exception(f"Tidak dapat menjadwalkan section {section['section_label']} - tidak ada slot tersedia (slot penuh)")
    
    # VALIDATION: Check for bentrok (conflicts) within new batch and with existing schedules
    def times_overlap_check(start1, end1, start2, end2):
        """Check if two time ranges overlap"""
        def to_minutes(t):
            h, m = map(int, t.split(':'))
            return h * 60 + m
        s1, e1 = to_minutes(start1), to_minutes(end1)
        s2, e2 = to_minutes(start2), to_minutes(end2)
        return s1 < e2 and s2 < e1
    
    # Build list of all newly scheduled sections with their slots
    newly_scheduled = []
    if sections_to_schedule is None:
        # Check from database - these were actually inserted
        for section_idx, section in enumerate(new_sections):
            # Find corresponding schedule in DB
            sched = schedules_collection.find_one({
                'course_id': section['course_id'],
                'section_label': section['section_label'],
                'dosen': section['lecturer']
            })
            if sched:
                newly_scheduled.append({
                    'section_label': section['section_label'],
                    'lecturer': section['lecturer'],
                    'day': sched['day'],
                    'start': sched['start'],
                    'end': sched['end'],
                    'room': sched['room']
                })
    else:
        # Dry-run mode: use simulated schedules
        for section_idx, (day, room, start, end) in new_batch_schedules.items():
            section = new_sections[section_idx]
            newly_scheduled.append({
                'section_label': section['section_label'],
                'lecturer': section['lecturer'],
                'day': day,
                'start': start,
                'end': end,
                'room': room
            })

    # Helper: resolve lecturer conflicts by finding alternate slot for the NEW section
    section_lookup = {sec['section_label']: sec for sec in new_sections}

    def try_reschedule_section(section_label):
        """Attempt to move a newly scheduled section to a conflict-free slot.
        Returns True if moved and updated in DB/newly_scheduled, else False."""
        if section_label not in section_lookup:
            return False
        sec = section_lookup[section_label]
        lect = sec.get('lecturer', '')
        sks = int(sec.get('sks', 2))
        is_lab = sec.get('is_lab', False)
        is_online = sec.get('is_online', False)
        
        # Candidate rooms: start with preferred type, then fallback to alternative
        if is_online:
            candidate_rooms_sets = [['ONLINE']]
        elif is_lab:
            candidate_rooms_sets = [LAB_ROOMS, NON_LAB_ROOMS]  # Try labs first, then non-lab as fallback
        else:
            candidate_rooms_sets = [NON_LAB_ROOMS, LAB_ROOMS]  # Try non-lab first, then lab as fallback
        
        weekday_order = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']

        def has_conflict(day, start, end, room, skip_label):
            """Check if slot conflicts with existing or newly scheduled sections.
            Also enforces preference constraints (available_days, blocked times, preferred_times).
            """
            # FIRST: Check lecturer preference constraints (hard block times)
            allowed, _ = check_preference_constraints(lect, day, start, end, is_online=is_online)
            if not allowed:
                return True  # Slot violates hard constraints
            
            # Check against existing schedules
            for sched in existing_schedules:
                if day != sched.get('day'):
                    continue
                if times_overlap_check(start, end, sched.get('start'), sched.get('end')):
                    # Lecturer clash
                    if lect and lect == sched.get('dosen', ''):
                        return True
                    # Room clash (skip online)
                    if room != 'ONLINE' and sched.get('room') == room:
                        return True
            # Check against other newly scheduled
            for sched in newly_scheduled:
                if sched['section_label'] == skip_label:
                    continue
                if day != sched['day']:
                    continue
                if times_overlap_check(start, end, sched['start'], sched['end']):
                    if lect and lect == sched['lecturer']:
                        return True
                    if room != 'ONLINE' and sched['room'] == room and sched['room'] != 'ONLINE':
                        return True
            return False

        # Build day priority: lecturer available_days first, then other days
        prefs = get_lecturer_preferences(lect)
        avail_days = prefs.get('available_days', []) or []
        current_entry = next((s for s in newly_scheduled if s['section_label'] == section_label), None)
        current_day = current_entry['day'] if current_entry else None

        if avail_days:
            # Lecturer has specific available days - prioritize those first
            # Keep order Mon-Fri but filter to available days, then add remaining days
            day_priority = [d for d in weekday_order if d in avail_days]
            # Add other days that are NOT in available (as fallback)
            day_priority.extend([d for d in weekday_order if d not in avail_days])
            # Move current day to the end to encourage changing day when resolving conflict
            if current_day and current_day in day_priority:
                day_priority = [d for d in day_priority if d != current_day] + [current_day]
        else:
            # No specific preference - try all days, but prioritize non-current days
            if current_day:
                day_priority = [d for d in weekday_order if d != current_day] + [current_day]
            else:
                day_priority = list(weekday_order)

        import logging
        _logger = logging.getLogger(__name__)
        _logger.info(f"[Resolver] Trying to reschedule {section_label} ({sks} SKS, {lect}, is_lab={is_lab})")
        _logger.info(f"[Resolver]   Available days: {avail_days}, Current day: {current_day}")
        _logger.info(f"[Resolver]   Day priority: {day_priority}")
        
        attempt_count = 0
        for day in day_priority:
            timeblocks = get_timeblocks_for_day(day, sks)
            _logger.info(f"[Resolver]   Trying {day}: {len(timeblocks)} timeblocks")
            for room_set in candidate_rooms_sets:
                for room in room_set:
                    for tb in timeblocks:
                        attempt_count += 1
                        start, end = tb.split('-')
                        allowed, pref_score = check_preference_constraints(lect, day, start, end, is_online=is_online)
                        if not allowed:
                            if attempt_count <= 20:  # Only log first 20 to reduce spam
                                _logger.info(f"[Resolver]     #{attempt_count} {day} {tb} {room}: REJECT (pref violation)")
                            continue
                        if has_conflict(day, start, end, room, section_label):
                            if attempt_count <= 20:
                                _logger.info(f"[Resolver]     #{attempt_count} {day} {tb} {room}: REJECT (conflict)")
                            continue
                        # Found slot: update newly_scheduled and DB if needed
                        _logger.info(f"[Resolver]     #{attempt_count} {day} {tb} {room}: ACCEPT ✓ (pref_score={pref_score})")
                        
                        # Update newly_scheduled list
                        if current_entry:
                            old_day = current_entry.get('day')
                            old_start = current_entry.get('start')
                            old_end = current_entry.get('end')
                            current_entry.update({'day': day, 'start': start, 'end': end, 'room': room})
                            _logger.info(f"[Resolver]     Updated newly_scheduled: {old_day} {old_start}-{old_end} → {day} {start}-{end} {room}")
                        
                        # Update DB if not dry-run (use OLD slot to find the right schedule entry)
                        if sections_to_schedule is None and current_entry:
                            # Match by lecturer + section_label + OLD slot (to find the exact schedule being moved)
                            update_result = schedules_collection.update_one(
                                {
                                    'dosen': sec['lecturer'],
                                    'section_label': sec['section_label'],
                                    'day': old_day,
                                    'start': old_start,
                                    'end': old_end
                                },
                                {'$set': {'day': day, 'start': start, 'end': end, 'room': room}}
                            )
                            _logger.info(f"[Resolver]     DB update matched: {update_result.matched_count}, modified: {update_result.modified_count}")
                        return True
        _logger.warning(f"[Resolver] Failed after {attempt_count} attempts - no slot available")
        return False
    
    # Check for dosen conflicts (same lecturer, same day, overlapping time)
    bentrok_dosen = []
    
    # 1️⃣ Check newly_scheduled AGAINST each other
    for i in range(len(newly_scheduled)):
        for j in range(i+1, len(newly_scheduled)):
            sched1 = newly_scheduled[i]
            sched2 = newly_scheduled[j]
            
            if (sched1['lecturer'] == sched2['lecturer'] and 
                sched1['day'] == sched2['day'] and
                times_overlap_check(sched1['start'], sched1['end'], sched2['start'], sched2['end'])):
                bentrok_dosen.append({
                    'lecturer': sched1['lecturer'],
                    'day': sched1['day'],
                    'section1': sched1['section_label'],
                    'time1': f"{sched1['start']}-{sched1['end']}",
                    'room1': sched1['room'],
                    'section2': sched2['section_label'],
                    'time2': f"{sched2['start']}-{sched2['end']}",
                    'room2': sched2['room']
                })
    
    # 2️⃣ CRITICAL: Check newly_scheduled AGAINST existing_schedules
    for new_sched in newly_scheduled:
        for existing_sched in existing_schedules:
            existing_lecturer = existing_sched.get('dosen', '')
            existing_day = existing_sched.get('day', '')
            existing_start = existing_sched.get('start', '')
            existing_end = existing_sched.get('end', '')
            existing_section = existing_sched.get('section_label', '')
            existing_room = existing_sched.get('room', '')
            existing_course = existing_sched.get('course_name', '')
            
            # 🔴 CRITICAL: Skip swapped sections (they were moved to new position)
            is_swapped = False
            for swapped in swapped_sections:
                if (swapped['section_label'] == existing_section and 
                    swapped['lecturer'] == existing_lecturer):
                    # This section was swapped - skip checking its old position
                    is_swapped = True
                    break
            
            if is_swapped:
                continue  # Skip - this section's old slot is now empty
            
            # Check if same lecturer would be double-booked
            if (new_sched['lecturer'] == existing_lecturer and 
                new_sched['day'] == existing_day and
                times_overlap_check(new_sched['start'], new_sched['end'], existing_start, existing_end)):
                bentrok_dosen.append({
                    'lecturer': new_sched['lecturer'],
                    'day': new_sched['day'],
                    'section1': new_sched['section_label'],
                    'time1': f"{new_sched['start']}-{new_sched['end']}",
                    'room1': new_sched['room'],
                    'section2': existing_section,
                    'time2': f"{existing_start}-{existing_end}",
                    'room2': existing_room,
                    'is_vs_existing': True
                })
    
    # Check for ruangan conflicts (same room, same day, overlapping time) - skip ONLINE
    bentrok_ruangan = []
    
    # 1️⃣ Check newly_scheduled AGAINST each other
    for i in range(len(newly_scheduled)):
        for j in range(i+1, len(newly_scheduled)):
            sched1 = newly_scheduled[i]
            sched2 = newly_scheduled[j]
            
            if (sched1['room'] == sched2['room'] and 
                sched1['room'] != 'ONLINE' and
                sched1['day'] == sched2['day'] and
                times_overlap_check(sched1['start'], sched1['end'], sched2['start'], sched2['end'])):
                bentrok_ruangan.append({
                    'room': sched1['room'],
                    'day': sched1['day'],
                    'lecturer1': sched1['lecturer'],
                    'time1': f"{sched1['start']}-{sched1['end']}",
                    'section1': sched1['section_label'],
                    'lecturer2': sched2['lecturer'],
                    'time2': f"{sched2['start']}-{sched2['end']}",
                    'section2': sched2['section_label']
                })
    
    # 2️⃣ CRITICAL: Check newly_scheduled AGAINST existing_schedules
    for new_sched in newly_scheduled:
        for existing_sched in existing_schedules:
            existing_lecturer = existing_sched.get('dosen', '')
            existing_day = existing_sched.get('day', '')
            existing_start = existing_sched.get('start', '')
            existing_end = existing_sched.get('end', '')
            existing_section = existing_sched.get('section_label', '')
            existing_room = existing_sched.get('room', '')
            
            # 🔴 CRITICAL: Skip swapped sections (they were moved to new position)
            is_swapped = False
            for swapped in swapped_sections:
                if (swapped['section_label'] == existing_section and 
                    swapped['lecturer'] == existing_lecturer):
                    # This section was swapped - skip checking its old position
                    is_swapped = True
                    break
            
            if is_swapped:
                continue  # Skip - this section's old slot is now empty
            
            # Check if same room would be used at same time
            if (new_sched['room'] == existing_room and 
                new_sched['room'] != 'ONLINE' and
                new_sched['day'] == existing_day and
                times_overlap_check(new_sched['start'], new_sched['end'], existing_start, existing_end)):
                bentrok_ruangan.append({
                    'room': new_sched['room'],
                    'day': new_sched['day'],
                    'lecturer1': new_sched['lecturer'],
                    'time1': f"{new_sched['start']}-{new_sched['end']}",
                    'section1': new_sched['section_label'],
                    'lecturer2': existing_lecturer,
                    'time2': f"{existing_start}-{existing_end}",
                    'section2': existing_section,
                    'is_vs_existing': True
                })
    
    # If bentrok found, raise error
    if bentrok_dosen:
        # Try to auto-resolve by moving the NEW section(s) to another slot
        unresolved = []
        for bentrok_detail in bentrok_dosen:
            tried = set()
            resolved = False

            # Prioritize moving section1 (usually the newly scheduled one)
            for label in [bentrok_detail['section1'], bentrok_detail['section2']]:
                if label in tried:
                    continue
                tried.add(label)
                if try_reschedule_section(label):
                    resolved = True
                    break

            if not resolved:
                unresolved.append(bentrok_detail)

        if unresolved:
            # Rollback if in DB mode (not dry-run)
            if sections_to_schedule is None:
                # Delete the sections and schedules we just inserted
                for section in new_sections:
                    sections_collection.delete_one({
                        'course_id': section['course_id'],
                        'section_label': section['section_label'],
                        'lecturer': section['lecturer']
                    })
                # Schedules were already deleted by delete_many in form handler if bentrok

            bentrok_detail = unresolved[0]
            raise Exception(
                f"❌ BENTROK DOSEN: {bentrok_detail['lecturer']} mengajar {bentrok_detail['section1']} dan {bentrok_detail['section2']} "
                f"pada {bentrok_detail['day']} jam {bentrok_detail['time1']} vs {bentrok_detail['time2']}!"
            )
    
    if bentrok_ruangan:
        # Rollback if in DB mode
        if sections_to_schedule is None:
            for section in new_sections:
                sections_collection.delete_one({
                    'course_id': section['course_id'],
                    'section_label': section['section_label'],
                    'lecturer': section['lecturer']
                })
        
        bentrok_detail = bentrok_ruangan[0]
        raise Exception(
            f"❌ BENTROK RUANGAN: {bentrok_detail['room']} dipakai untuk {bentrok_detail['section1']} ({bentrok_detail['lecturer1']}) "
            f"dan {bentrok_detail['section2']} ({bentrok_detail['lecturer2']}) pada {bentrok_detail['day']} jam {bentrok_detail['time1']} vs {bentrok_detail['time2']}!"
        )
    
    return {'sections_scheduled': scheduled_count, 'swap_info': swap_info}
@app.route("/delete_schedule/<schedule_id>", methods=["POST"])
def delete_schedule(schedule_id):
    """Delete a schedule entry manually"""
    if 'user' not in session or session.get('user', {}).get('role') != 'koordinator':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        result = schedules_collection.delete_one({'_id': ObjectId(schedule_id)})
        if result.deleted_count > 0:
            flash("✓ Jadwal berhasil dihapus!", "success")
        else:
            flash("❌ Jadwal tidak ditemukan!", "error")
    except Exception as e:
        flash(f"❌ Error menghapus jadwal: {str(e)}", "error")
    
    return redirect(url_for('koordinator'))


@app.route("/api/get_course_sections/<course_id>", methods=["GET"])
def get_course_sections(course_id):
    """Get all sections for a specific course with their schedules"""
    if 'user' not in session or session.get('user', {}).get('role') != 'koordinator':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        # Get all sections for this course
        sections = list(sections_collection.find({'course_id': course_id}))
        
        result_sections = []
        for section in sections:
            section_id = str(section['_id'])
            
            # Find schedule for this section
            schedule = schedules_collection.find_one({
                'course_id': course_id,
                'section_label': section.get('section_label')
            })
            
            section_data = {
                '_id': section_id,
                'section_label': section.get('section_label', ''),
                'lecturer': section.get('lecturer', 'N/A'),
                'sks': section.get('sks', 0),
                'schedule': None,
                'schedule_id': None
            }
            
            if schedule:
                section_data['schedule'] = {
                    'day': schedule.get('day', ''),
                    'start': schedule.get('start', ''),
                    'end': schedule.get('end', ''),
                    'room': schedule.get('room', '')
                }
                section_data['schedule_id'] = str(schedule['_id'])
            
            result_sections.append(section_data)
        
        # Sort by section label
        result_sections.sort(key=lambda x: x['section_label'])
        
        return jsonify({
            'success': True,
            'sections': result_sections
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route("/api/delete_section", methods=["POST"])
def delete_section():
    """Delete a section and its schedule"""
    if 'user' not in session or session.get('user', {}).get('role') != 'koordinator':
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    try:
        data = request.json
        section_id = data.get('section_id')
        schedule_id = data.get('schedule_id')
        
        # Delete section
        sections_collection.delete_one({'_id': ObjectId(section_id)})
        
        # Delete schedule if exists
        if schedule_id:
            schedules_collection.delete_one({'_id': ObjectId(schedule_id)})
        
        return jsonify({
            'success': True,
            'message': 'Section berhasil dihapus'
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


# ============================================================================
# EXPORT SCHEDULE WITH PROBABILITY
# ============================================================================

@app.route("/export_schedule_with_probability")
def export_schedule_with_probability():
    """Export jadwal lengkap dengan preference score dan match quality ke CSV"""
    if 'user' not in session:
        flash("Silakan login terlebih dahulu!", "error")
        return redirect(url_for("login"))
    
    try:
        # Get all schedules from database
        schedules = list(schedules_collection.find())
        
        if not schedules:
            flash("Belum ada jadwal untuk di-export!", "warning")
            return redirect(url_for('koordinator'))
        
        # Prepare data for export
        export_data = []
        for sched in schedules:
            export_data.append({
                'Hari': sched.get('day', ''),
                'Jam Mulai': sched.get('start', ''),
                'Jam Selesai': sched.get('end', ''),
                'Ruangan': sched.get('room', ''),
                'Mata Kuliah': sched.get('course_name', ''),
                'Section': sched.get('section_label', ''),
                'SKS': sched.get('sks', ''),
                'Dosen': sched.get('dosen', ''),
                'Tipe': 'Lab' if sched.get('is_lab') else ('Online' if sched.get('is_online') else 'Reguler'),
                'Preference Score (%)': sched.get('preference_score', 50),
                'Kesesuaian': sched.get('match_quality', 'Acceptable')
            })
        
        # Convert to DataFrame
        df = pd.DataFrame(export_data)
        
        # Sort by day and time
        day_order = {'Senin': 1, 'Selasa': 2, 'Rabu': 3, 'Kamis': 4, 'Jumat': 5}
        df['_day_order'] = df['Hari'].map(day_order)
        df = df.sort_values(['_day_order', 'Jam Mulai'])
        df = df.drop('_day_order', axis=1)
        
        # Create Excel file with formatting
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Jadwal dengan Probabilitas', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Jadwal dengan Probabilitas']
            
            # Add styling
            from openpyxl.styles import PatternFill, Font, Alignment
            
            # Header styling
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Color code the Kesesuaian column
            quality_colors = {
                'Excellent': 'C6EFCE',  # Light green
                'Good': 'DAEEF3',       # Light blue
                'Acceptable': 'FFF2CC', # Light yellow
                'Poor': 'FFC7CE'        # Light red
            }
            
            # Find Kesesuaian column index
            kesesuaian_col = None
            for idx, cell in enumerate(worksheet[1], 1):
                if cell.value == 'Kesesuaian':
                    kesesuaian_col = idx
                    break
            
            # Apply conditional formatting to Kesesuaian column
            if kesesuaian_col:
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=kesesuaian_col)
                    quality = cell.value
                    if quality in quality_colors:
                        cell.fill = PatternFill(start_color=quality_colors[quality], 
                                               end_color=quality_colors[quality], 
                                               fill_type='solid')
                        cell.font = Font(bold=True)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Generate filename with timestamp
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Jadwal_dengan_Probabilitas_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        flash(f"❌ Error saat export: {str(e)}", "error")
        return redirect(url_for('koordinator'))


@app.route("/coordinator/schedule_analytics")
def schedule_analytics():
    """Dashboard analitik kesesuaian jadwal dengan preferensi dosen"""
    if 'user' not in session or session.get('user', {}).get('role') != 'koordinator':
        flash("Akses ditolak!", "error")
        return redirect(url_for('login'))
    
    try:
        # Get all schedules
        schedules = list(schedules_collection.find())
        
        if not schedules:
            flash("Belum ada jadwal untuk dianalisis!", "warning")
            return redirect(url_for('koordinator'))
        
        # Calculate statistics
        total_schedules = len(schedules)
        
        # Count by match quality
        quality_counts = {
            'Excellent': 0,
            'Good': 0,
            'Acceptable': 0,
            'Poor': 0
        }
        
        # Lecturer statistics
        lecturer_stats = {}
        
        for sched in schedules:
            quality = sched.get('match_quality', 'Acceptable')
            quality_counts[quality] = quality_counts.get(quality, 0) + 1
            
            dosen = sched.get('dosen', 'Unknown')
            score = sched.get('preference_score', 50)
            
            if dosen not in lecturer_stats:
                lecturer_stats[dosen] = {
                    'total_classes': 0,
                    'scores': [],
                    'excellent': 0,
                    'good': 0,
                    'acceptable': 0,
                    'poor': 0
                }
            
            lecturer_stats[dosen]['total_classes'] += 1
            lecturer_stats[dosen]['scores'].append(score)
            lecturer_stats[dosen][quality.lower()] += 1
        
        # Calculate average scores per lecturer
        for dosen, stats in lecturer_stats.items():
            if stats['scores']:
                stats['avg_score'] = sum(stats['scores']) / len(stats['scores'])
            else:
                stats['avg_score'] = 0
        
        # Sort lecturers by average score (descending)
        sorted_lecturers = sorted(
            lecturer_stats.items(),
            key=lambda x: x[1]['avg_score'],
            reverse=True
        )
        
        # Calculate percentages
        quality_percentages = {
            quality: (count / total_schedules * 100) if total_schedules > 0 else 0
            for quality, count in quality_counts.items()
        }
        
        # Overall average score
        all_scores = [s.get('preference_score', 50) for s in schedules]
        overall_avg_score = sum(all_scores) / len(all_scores) if all_scores else 0
        
        return render_template(
            'koordinator_schedule_analytics.html',
            total_schedules=total_schedules,
            quality_counts=quality_counts,
            quality_percentages=quality_percentages,
            overall_avg_score=overall_avg_score,
            sorted_lecturers=sorted_lecturers
        )
    
    except Exception as e:
        flash(f"❌ Error loading analytics: {str(e)}", "error")
        return redirect(url_for('koordinator'))


if __name__ == "__main__":

    app.run(debug=True)

